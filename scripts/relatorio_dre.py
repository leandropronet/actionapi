#!/usr/bin/env python3
r"""Gera relatório gerencial de DRE + Balanço Patrimonial via ActionAPI.

O desenho parte da planilha-modelo:

    relatorios/contabilidade/SGA_DRE E BP 2021 a 2025 ANUAL_v2.xlsx

O relatório não acessa Oracle nem PostgreSQL diretamente. Os saldos vêm da
ActionAPI, endpoint /api/v1/executivo/contabilidade/sintetico.

Para DRE, o script solicita excluirEncerramento=true para remover o histórico
1000191 (zeramento anual), pois contas de resultado ficam zeradas depois do
encerramento contábil. Para BP, usa saldo acumulado normal.

Se --anos não for informado, o script pergunta interativamente o intervalo.

Exemplos:

    .\.venv\Scripts\python.exe scripts\relatorio_dre.py
    .\.venv\Scripts\python.exe scripts\relatorio_dre.py --anos 2021-2025
    .\.venv\Scripts\python.exe scripts\relatorio_dre.py --api-url http://127.0.0.1:3001
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

try:
    from gerar_relatorios_executivos import (
        BLUE,
        DATE,
        DECIMAL,
        GREEN,
        MONEY,
        NAVY,
        ORANGE,
        PERCENT,
        RED,
        WHITE,
        api_get,
        first_api_key,
        number,
        save,
        style_sheet,
        title_block,
    )
except ModuleNotFoundError:
    from scripts.gerar_relatorios_executivos import (
        BLUE,
        DATE,
        DECIMAL,
        GREEN,
        MONEY,
        NAVY,
        ORANGE,
        PERCENT,
        RED,
        WHITE,
        api_get,
        first_api_key,
        number,
        save,
        style_sheet,
        title_block,
    )

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:
    print(
        "Dependência ausente. Execute:\n"
        "  py -m pip install -r scripts/requirements-relatorio-contas-pagar.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_API_URL = "http://127.0.0.1:3000"
DEFAULT_MODEL = ROOT / "relatorios" / "contabilidade" / "SGA_DRE E BP 2021 a 2025 ANUAL_v2.xlsx"
DATA_INICIO_PLANO_ATUAL = "2008-01-01"
ENCERRAMENTO_HISTORICO = "1000191"

BRANCHES = [
    ("1", "Goiatuba"),
    ("9", "Piracanjuba"),
    ("8", "Alvorada"),
    ("3", "Gurupi"),
    ("4", "Lagoa"),
    ("5", "Porto"),
]


@dataclass(frozen=True)
class SourceSpec:
    account: str
    sign: str
    description: str = ""


@dataclass(frozen=True)
class DreLine:
    key: str
    label: str
    section: str
    level: int = 0
    source: SourceSpec | None = None
    formula: Callable[[dict[str, float]], float] | None = None
    bold: bool = False
    percent_base: str | None = "receita_liquida"


@dataclass(frozen=True)
class BpLine:
    account: str
    label: str
    level: int = 0
    bold: bool = False


DRE_LINES: list[DreLine] = [
    DreLine("receita_bruta", "Receita Bruta com Vendas", "Receitas", source=SourceSpec("3111", "credit"), bold=True),
    DreLine("vendas_mercadorias", "Vendas de Mercadorias em Geral", "Receitas", 1, SourceSpec("311101", "credit")),
    DreLine("venda_defensivos", "Venda de Defensivos", "Receitas", 1, SourceSpec("311102", "credit")),
    DreLine("venda_fertilizantes", "Venda de Fertilizantes", "Receitas", 1, SourceSpec("311103", "credit")),
    DreLine("venda_sementes", "Venda de Sementes", "Receitas", 1, SourceSpec("311104", "credit")),
    DreLine("prestacao_servico", "Prestação de Serviço/Locação", "Receitas", 1, SourceSpec("311105", "credit")),
    DreLine("avp_receitas", "AVP Receitas", "Receitas", 1, SourceSpec("311108", "credit")),
    DreLine("deducoes_receita", "Deduções da Receita Bruta", "Deduções", source=SourceSpec("3112", "credit"), bold=True),
    DreLine("devolucoes_geral", "Devoluções de Vendas", "Deduções", 1, SourceSpec("3112010002", "credit")),
    DreLine("impostos_geral", "Impostos nas Vendas em Geral", "Deduções", 1, formula=lambda v: v["deducoes_geral_total"] - v["devolucoes_geral"]),
    DreLine("deducoes_geral_total", "Deduções Sobre Vendas Mercadorias", "Auxiliar", source=SourceSpec("311201", "credit")),
    DreLine("devolucoes_defensivos", "Devoluções de Vendas Defensivos", "Deduções", 1, SourceSpec("3112020002", "credit")),
    DreLine(
        "impostos_defensivos",
        "Impostos nas Vendas de Defensivos",
        "Deduções",
        1,
        formula=lambda v: v["deducoes_defensivos_total"] - v["devolucoes_defensivos"],
    ),
    DreLine("deducoes_defensivos_total", "Deduções Sobre Vendas Defensivos", "Auxiliar", source=SourceSpec("311202", "credit")),
    DreLine("devolucoes_fertilizantes", "Devoluções de Vendas Fertilizantes", "Deduções", 1, SourceSpec("3112030002", "credit")),
    DreLine(
        "impostos_fertilizantes",
        "Impostos nas Vendas de Fertilizantes",
        "Deduções",
        1,
        formula=lambda v: v["deducoes_fertilizantes_total"] - v["devolucoes_fertilizantes"],
    ),
    DreLine("deducoes_fertilizantes_total", "Deduções Sobre Vendas Fertilizantes", "Auxiliar", source=SourceSpec("311203", "credit")),
    DreLine("devolucoes_sementes", "Devoluções de Vendas Sementes", "Deduções", 1, SourceSpec("3112040002", "credit")),
    DreLine(
        "impostos_sementes",
        "Impostos nas Vendas de Sementes",
        "Deduções",
        1,
        formula=lambda v: v["deducoes_sementes_total"] - v["devolucoes_sementes"],
    ),
    DreLine("deducoes_sementes_total", "Deduções Sobre Vendas Sementes", "Auxiliar", source=SourceSpec("311204", "credit")),
    DreLine("deducoes_servicos", "Deduções da Receita Bruta de Serviços e Locação", "Deduções", 1, SourceSpec("311205", "credit")),
    DreLine(
        "receita_liquida",
        "Receita Operacional Líquida",
        "Resultado",
        formula=lambda v: v["receita_bruta"] + v["deducoes_receita"],
        bold=True,
        percent_base=None,
    ),
    DreLine("custos_vendas", "Custos sobre Vendas", "Custos", source=SourceSpec("41", "debit"), bold=True),
    DreLine("custos_mercadorias", "Custos das Mercadorias em Geral Vendidas", "Custos", 1, SourceSpec("411101", "debit")),
    DreLine("custos_defensivos", "Custos dos Defensivos Vendidos", "Custos", 1, SourceSpec("411102", "debit")),
    DreLine("custos_fertilizantes", "Custos dos Fertilizantes Vendidos", "Custos", 1, SourceSpec("411103", "debit")),
    DreLine("custos_sementes", "Custos das Sementes Vendidas", "Custos", 1, SourceSpec("411104", "debit")),
    DreLine("custos_servicos", "Custos dos Serviços Prestados", "Custos", 1, SourceSpec("411105", "debit")),
    DreLine("avp_custo", "AVP Custo", "Custos", 1, SourceSpec("411107", "debit")),
    DreLine(
        "lucro_bruto",
        "Lucro Bruto / Margem Bruta",
        "Resultado",
        formula=lambda v: v["receita_liquida"] - v["custos_vendas"],
        bold=True,
        percent_base="receita_liquida",
    ),
    DreLine(
        "despesas_adm_com",
        "Despesas Administrativas e Comerciais",
        "Despesas",
        source=SourceSpec("4211", "debit"),
        formula=lambda v: v["despesas_adm_com_fonte"] - v["pcld_constituicao"] - v["pcld_reversao"] - v["perdas_estoque"],
        bold=True,
    ),
    DreLine("despesas_adm_com_fonte", "Despesas Administrativas e Comerciais (fonte 4211)", "Auxiliar", source=SourceSpec("4211", "debit")),
    DreLine("rh_diretores", "Despesas com RH Diretores", "Despesas", 1, SourceSpec("421101", "debit")),
    DreLine("rh_fixas", "Despesas com RH Fixas", "Despesas", 1, SourceSpec("421102", "debit")),
    DreLine("rh_variaveis", "Despesas com RH Variáveis", "Despesas", 1, SourceSpec("421103", "debit")),
    DreLine("ocupacao", "Ocupação", "Despesas", 1, SourceSpec("421104", "debit")),
    DreLine("utilidades", "Utilidades e Serviços", "Despesas", 1, SourceSpec("421105", "debit")),
    DreLine("funcionamento", "Despesas de Funcionamento", "Despesas", 1, SourceSpec("421108", "debit")),
    DreLine("servicos_profissionais", "Serviços Profissionais", "Despesas", 1, SourceSpec("421109", "debit")),
    DreLine("comunicacao", "Comunicação", "Despesas", 1, SourceSpec("421110", "debit")),
    DreLine("propaganda", "Propaganda e Publicidade", "Despesas", 1, SourceSpec("421112", "debit")),
    DreLine("frota", "Frota", "Despesas", 1, SourceSpec("421113", "debit")),
    DreLine("transporte", "Transporte / Logísticas", "Despesas", 1, SourceSpec("421114", "debit")),
    DreLine("tributos", "Tributos e Contribuições", "Despesas", 1, SourceSpec("421118", "debit")),
    DreLine("despesas_bancarias", "Despesas Bancárias", "Despesas", 1, SourceSpec("421120", "debit")),
    DreLine(
        "lucro_operacional",
        "Lucro Operacional antes do Resultado Financeiro e Provisões",
        "Resultado",
        formula=lambda v: v["lucro_bruto"] - v["despesas_adm_com"],
        bold=True,
    ),
    DreLine("resultado_financeiro", "Resultado Financeiro", "Financeiro", formula=lambda v: v["receitas_financeiras"] - v["despesas_financeiras"], bold=True),
    DreLine("receitas_financeiras", "Receitas Financeiras Totais", "Financeiro", 1, SourceSpec("4312", "credit")),
    DreLine("despesas_financeiras", "Despesas Financeiras Totais", "Financeiro", 1, SourceSpec("4311", "debit")),
    DreLine("pcld", "PCLD", "Provisões", formula=lambda v: v["pcld_perda"] + v["pcld_constituicao"] + v["pcld_reversao"], bold=True),
    DreLine("pcld_perda", "4211250060 - Despesa com Perda de PCLD", "Provisões", 1, SourceSpec("4211250060", "debit")),
    DreLine("pcld_constituicao", "4211250001 - Constituição do PCLD Contábil", "Provisões", 1, SourceSpec("4211250001", "debit")),
    DreLine("pcld_reversao", "4211250006 - Reversão do PCLD Contábil", "Provisões", 1, SourceSpec("4211250006", "debit")),
    DreLine("outras_rec_desp", "Outras Receitas e Despesas Operacionais", "Outros", source=SourceSpec("4331", "credit")),
    DreLine("perdas_estoque", "Constituição de Perdas Estimadas nos Estoques", "Outros", source=SourceSpec("4211250005", "debit")),
    DreLine(
        "resultado_contabil_antes_impostos",
        "Resultado Contábil antes dos Impostos (com PCLD)",
        "Resultado",
        # A Perda de PCLD (4211250060) já está dentro de despesas_adm_com (que só
        # devolve Constituição/Reversão/Perdas de Estoque). Subtrair o PCLD inteiro
        # aqui contaria a Perda em dobro (erro do modelo, célula r59). Por isso
        # descontamos apenas Constituição + Reversão = (pcld - pcld_perda).
        formula=lambda v: -((v["pcld"] - v["pcld_perda"]) - v["outras_rec_desp"] + v["perdas_estoque"]) + v["lucro_operacional"] + v["resultado_financeiro"],
        bold=True,
    ),
    DreLine(
        "resultado_gerencial_antes_impostos",
        "Resultado Gerencial antes dos Impostos",
        "Resultado",
        formula=lambda v: v["lucro_operacional"] + v["resultado_financeiro"] + v["outras_rec_desp"] - v["perdas_estoque"],
        bold=True,
    ),
    DreLine("provisoes_fiscais", "Provisões Fiscais", "Impostos", source=SourceSpec("4341", "debit"), bold=True),
    DreLine("ir_cs", "Imposto de Renda e Contribuição Social", "Impostos", 1, SourceSpec("434101", "debit")),
    DreLine(
        "resultado_exercicio",
        "Resultado do Exercício",
        "Resultado",
        formula=lambda v: v["resultado_gerencial_antes_impostos"] - v["provisoes_fiscais"],
        bold=True,
        percent_base="receita_liquida",
    ),
    DreLine("depreciacao", "Depreciação e Amortizações", "EBITDA", formula=lambda v: v["depreciacao_veiculos"] + v["depreciacao_equipamentos"], bold=True),
    DreLine("depreciacao_veiculos", "4211130006 - Depreciação e Amortizações Veículos", "EBITDA", 1, SourceSpec("4211130006", "debit")),
    DreLine("depreciacao_equipamentos", "4211080003 - Depreciação e Amortizações Móveis e Equip. Inform.", "EBITDA", 1, SourceSpec("4211080003", "debit")),
    DreLine("ebitda", "EBITDA", "Resultado", formula=lambda v: v["lucro_operacional"] + v["depreciacao"], bold=True, percent_base="receita_liquida"),
    DreLine("margem_bruta_valor", "Margem Bruta (valor)", "Resultado", formula=lambda v: v["lucro_bruto"], bold=True),
]

VISIBLE_DRE_KEYS = [
    line.key
    for line in DRE_LINES
    if line.section != "Auxiliar"
]

DRE_FORMULA_TEXT = {
    "deducoes_geral_total": "Conta 311201, valor = créditos - débitos",
    "impostos_geral": "Deduções Sobre Vendas Mercadorias - Devoluções de Vendas",
    "deducoes_defensivos_total": "Conta 311202, valor = créditos - débitos",
    "impostos_defensivos": "Deduções Sobre Vendas Defensivos - Devoluções de Vendas Defensivos",
    "deducoes_fertilizantes_total": "Conta 311203, valor = créditos - débitos",
    "impostos_fertilizantes": "Deduções Sobre Vendas Fertilizantes - Devoluções de Vendas Fertilizantes",
    "deducoes_sementes_total": "Conta 311204, valor = créditos - débitos",
    "impostos_sementes": "Deduções Sobre Vendas Sementes - Devoluções de Vendas Sementes",
    "receita_liquida": "Receita Bruta + Deduções da Receita Bruta",
    "lucro_bruto": "Receita Operacional Líquida - Custos sobre Vendas",
    "despesas_adm_com": "Despesas Adm/Com fonte 4211 - Constituição PCLD - Reversão PCLD - Perdas Estimadas nos Estoques",
    "despesas_adm_com_fonte": "Conta 4211, valor = débitos - créditos",
    "lucro_operacional": "Lucro Bruto - Despesas Administrativas e Comerciais",
    "resultado_financeiro": "Receitas Financeiras Totais - Despesas Financeiras Totais",
    "pcld": "Despesa com Perda de PCLD + Constituição PCLD + Reversão PCLD",
    "resultado_contabil_antes_impostos": "Lucro Operacional + Resultado Financeiro - Constituição PCLD - Reversão PCLD + Outras Receitas/Despesas - Perdas Estimadas nos Estoques",
    "resultado_gerencial_antes_impostos": "Lucro Operacional + Resultado Financeiro + Outras Receitas/Despesas - Perdas Estimadas nos Estoques",
    "resultado_exercicio": "Resultado Gerencial antes dos Impostos - Provisões Fiscais",
    "depreciacao": "Depreciação Veículos + Depreciação Móveis/Equipamentos",
    "ebitda": "Lucro Operacional antes do Resultado Financeiro e Provisões + Depreciação e Amortizações",
    "margem_bruta_valor": "Lucro Bruto / Margem Bruta em valor",
}

DRE_COMPONENT_KEYS = {
    "impostos_geral": ["deducoes_geral_total", "devolucoes_geral"],
    "impostos_defensivos": ["deducoes_defensivos_total", "devolucoes_defensivos"],
    "impostos_fertilizantes": ["deducoes_fertilizantes_total", "devolucoes_fertilizantes"],
    "impostos_sementes": ["deducoes_sementes_total", "devolucoes_sementes"],
    "receita_liquida": ["receita_bruta", "deducoes_receita"],
    "lucro_bruto": ["receita_liquida", "custos_vendas"],
    "despesas_adm_com": ["despesas_adm_com_fonte", "pcld_constituicao", "pcld_reversao", "perdas_estoque"],
    "lucro_operacional": ["lucro_bruto", "despesas_adm_com"],
    "resultado_financeiro": ["receitas_financeiras", "despesas_financeiras"],
    "pcld": ["pcld_perda", "pcld_constituicao", "pcld_reversao"],
    "resultado_contabil_antes_impostos": ["lucro_operacional", "resultado_financeiro", "pcld_constituicao", "pcld_reversao", "outras_rec_desp", "perdas_estoque"],
    "resultado_gerencial_antes_impostos": ["lucro_operacional", "resultado_financeiro", "outras_rec_desp", "perdas_estoque"],
    "resultado_exercicio": ["resultado_gerencial_antes_impostos", "provisoes_fiscais"],
    "depreciacao": ["depreciacao_veiculos", "depreciacao_equipamentos"],
    "ebitda": ["lucro_operacional", "depreciacao"],
    "margem_bruta_valor": ["lucro_bruto"],
}

INDICATOR_DEFINITIONS = [
    ("Receita Líquida", "receita_liquida", "money", "Receita Operacional Líquida da DRE"),
    ("Lucro Bruto", "lucro_bruto", "money", "Receita Líquida - Custos sobre Vendas"),
    ("Margem Bruta", "margem_bruta", "percent", "Lucro Bruto / Receita Líquida"),
    ("EBITDA", "ebitda", "money", "Lucro Operacional antes do Resultado Financeiro e Provisões + Depreciação/Amortização"),
    ("Margem EBITDA", "margem_ebitda", "percent", "EBITDA / Receita Líquida"),
    ("Resultado do Exercício", "resultado_exercicio", "money", "Resultado Gerencial antes dos Impostos - Provisões Fiscais"),
    ("Margem Líquida", "margem_liquida", "percent", "Resultado do Exercício / Receita Líquida"),
    ("Liquidez Corrente", "liquidez_corrente", "decimal", "Ativo Circulante / Passivo Circulante"),
    ("Liquidez Imediata", "liquidez_imediata", "decimal", "Disponível / Passivo Circulante"),
    ("Liquidez Seca", "liquidez_seca", "decimal", "(Ativo Circulante - Estoques) / Passivo Circulante"),
    ("Liquidez Geral — técnica: (AC + RLP) / (PC + PNC)", "liquidez_geral", "decimal", "(Ativo Circulante + Realizável a Longo Prazo) / (Passivo Circulante + Passivo Não Circulante)"),
    ("Liquidez Geral — critério modelo: (AC + ANC) / (PC + PNC)", "liquidez_geral_modelo", "decimal", "(Ativo Circulante + Ativo Não Circulante) / (Passivo Circulante + Passivo Não Circulante)"),
    ("Capital de Terceiros / PL: (PC + PNC) / PL", "endividamento", "decimal", "(Passivo Circulante + Passivo Não Circulante) / Patrimônio Líquido"),
    ("Endividamento Geral: (PC + PNC) / Ativo", "endividamento_geral", "percent", "(Passivo Circulante + Passivo Não Circulante) / Ativo Total"),
    ("Endividamento — critério modelo: Grupo 2 / PL", "endividamento_modelo", "decimal", "Grupo contábil 2 / Patrimônio Líquido; mantido para reconciliação com modelo"),
    ("Empréstimos / EBITDA", "emprestimos_ebitda", "decimal", "Empréstimos e Financiamentos Circulantes / EBITDA"),
    ("ROA — Resultado DRE / Ativo", "roa", "percent", "Resultado do Exercício / Ativo Total"),
    ("ROE — Resultado DRE / PL", "roe", "percent", "Resultado do Exercício / Patrimônio Líquido"),
    ("Ativo Total", "ativo_total", "money", "Conta 1 do Balanço Patrimonial"),
    ("Passivo Total", "passivo_total", "money", "Conta 2 do Balanço Patrimonial apresentada com sinal positivo"),
    ("Passivo Exigível", "passivo_exigivel", "money", "Passivo Circulante + Passivo Não Circulante"),
    ("Patrimônio Líquido", "patrimonio_liquido", "money", "Conta 23 do Balanço Patrimonial apresentada com sinal positivo"),
    ("Empréstimos e Financiamentos", "emprestimos", "money", "Conta 211104 do Balanço Patrimonial apresentada com sinal positivo"),
]

BP_LINES = [
    BpLine("1", "Ativo", 0, True),
    BpLine("11", "Circulante", 1, True),
    BpLine("111", "Disponível", 2),
    BpLine("1111", "Bens Numerários", 3),
    BpLine("1112", "Banco Cta Movimento", 3),
    BpLine("1113", "Aplicações Financeiras", 3),
    BpLine("1114", "Equivalentes de Caixa", 3),
    BpLine("112", "Clientes", 2),
    BpLine("1121", "Créditos a Receber", 3),
    BpLine("113", "Outros Créditos", 2),
    BpLine("1131", "Títulos a Receber", 3),
    BpLine("1132", "Créditos a Funcionários", 3),
    BpLine("1134", "Tributos a Compensar e Recuperar", 3),
    BpLine("114", "Investimentos Temporais", 2),
    BpLine("1141", "Investimentos Temporais", 3),
    BpLine("115", "Estoques", 2),
    BpLine("1151", "Estoque de Mercadorias", 3),
    BpLine("1153", "Estoque de Produtos", 3),
    BpLine("117", "Despesas do Exercício Seguinte Pagas Antecipadamente", 2),
    BpLine("1171", "Despesas a Apropriar Exercícios Seguintes", 3),
    BpLine("12", "Ativo Não Circulante", 1, True),
    BpLine("121", "Realizável a Longo Prazo", 2),
    BpLine("1211", "Créditos e Valores Longo Prazo", 3),
    BpLine("1213", "Despesas Antecipadas", 3),
    BpLine("1214", "Tributos Diferidos", 3),
    BpLine("123", "Imobilizado", 2),
    BpLine("1231", "Bens em Operação - Custo", 3),
    BpLine("1233", "Imobilizado em Andamento - Custo", 3),
    BpLine("124", "Intangível - Custo", 2),
    BpLine("1241", "Intangível Custo", 3),
    BpLine("1242", "(-) Amortização Acumulada / Perdas Estimadas", 3),
    BpLine("2", "Passivo + Patrimônio Líquido", 0, True),
    BpLine("21", "Passivo Circulante", 1, True),
    BpLine("211", "Circulante", 2),
    BpLine("2111", "Deveres e Obrigações", 3),
    BpLine("22", "Passivo Não Circulante", 1, True),
    BpLine("221", "Não Circulante", 2),
    BpLine("2211", "Não Circulante", 3),
    BpLine("23", "Patrimônio Líquido", 1, True),
    BpLine("231", "Patrimônio Líquido", 2),
    BpLine("2311", "Patrimônio Líquido", 3),
    BpLine("231101", "Capital Subscrito", 3),
    BpLine("2311010001", "Capital Autorizado", 4),
    BpLine("231104", "Reserva de Lucros", 3),
    BpLine("2311040004", "Reservas de Lucros a Realizar", 4),
    BpLine("2311040010", "Ajustes de Exercícios Anteriores", 4),
    BpLine("2311040090", "(-) Reserva de Subvenção", 4),
    BpLine("231105", "Lucros ou Prejuízos Acumulados", 3),
    BpLine("2311050001", "Lucros Acumulados", 4),
    BpLine("2311050002", "(-) Prejuízos Acumulados", 4),
    BpLine("2311050003", "(-) Lucros Distribuídos no Ano", 4),
    BpLine("231107", "Reserva Fiscal", 3),
    BpLine("2311070001", "Reserva de Incentivos Fiscais", 4),
    BpLine("211104", "Empréstimos e Financiamentos (Circulante)", 2, True),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera relatório DRE + BP anual via ActionAPI.")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument(
        "--anos",
        default=None,
        help="Intervalo/lista de anos. Ex.: 2021-2025 ou 2021,2022,2025. Se omitido, o script pergunta interativamente.",
    )
    parser.add_argument("--arquivo", help="Caminho do arquivo .xlsx de saída.")
    parser.add_argument("--modelo", default=str(DEFAULT_MODEL), help="Planilha-modelo para validação opcional.")
    parser.add_argument("--sem-validar-modelo", action="store_true", help="Não compara contra a planilha-modelo.")
    parser.add_argument("--historico-encerramento", default=ENCERRAMENTO_HISTORICO)
    parser.add_argument(
        "--permitir-api-sem-filtro",
        action="store_true",
        help="Permite continuar mesmo se a API em execução não aplicar excluirEncerramento=true.",
    )
    return parser.parse_args()


def prompt_for_years() -> str:
    default = f"2021-{datetime.now().year}"
    raw = input(f"Informe o intervalo de anos do relatório (ex.: 2021-2025) [Enter = {default}]: ").strip()
    return raw or default


def parse_years(raw: str) -> list[int]:
    text = str(raw or "").strip()
    if re.fullmatch(r"\d{4}\s*-\s*\d{4}", text):
        start, end = [int(item.strip()) for item in text.split("-", 1)]
        if end < start:
            raise ValueError("--anos inválido: ano final anterior ao inicial.")
        return list(range(start, end + 1))
    years = []
    for part in re.split(r"[,\s;]+", text):
        if not part:
            continue
        if not re.fullmatch(r"\d{4}", part):
            raise ValueError(f"--anos inválido: {part!r}")
        years.append(int(part))
    if not years:
        raise ValueError("--anos deve informar ao menos um ano.")
    return sorted(dict.fromkeys(years))


def value_from_source(row: dict[str, Any] | None, sign: str) -> float:
    if not row:
        return 0.0
    saldo = number(row.get("saldo"))
    if sign == "credit":
        return -saldo
    return saldo


def account_value(accounts: dict[str, dict[str, Any]], spec: SourceSpec) -> float:
    return value_from_source(accounts.get(spec.account), spec.sign)


def dre_line_formula_text(line: DreLine) -> str:
    if line.key in DRE_FORMULA_TEXT:
        return DRE_FORMULA_TEXT[line.key]
    if line.source:
        if line.source.sign == "credit":
            return f"Conta {line.source.account}, valor = créditos - débitos"
        return f"Conta {line.source.account}, valor = débitos - créditos"
    return "Fórmula não documentada"


def dre_line_components_text(line: DreLine, values: dict[str, float]) -> str:
    label_by_key = {item.key: item.label for item in DRE_LINES}
    components = DRE_COMPONENT_KEYS.get(line.key, [])
    if components:
        return "; ".join(
            f"{label_by_key.get(key, key)}={values.get(key, 0.0):,.2f}"
            for key in components
        )
    if line.source:
        return f"conta={line.source.account}; sinal={line.source.sign}"
    else:
        return ""


def format_formula_value(value: float) -> str:
    return f"{value:,.2f}"


def bp_value(accounts: dict[str, dict[str, Any]], account_id: str) -> float:
    row = accounts.get(account_id)
    if not row:
        return 0.0
    saldo = number(row.get("saldo"))
    if account_id.startswith("2"):
        return -saldo
    return saldo


def safe_ratio(num: float, den: float) -> float:
    return num / den if den else 0.0


def fetch_synthetic(
    api_url: str,
    api_key: str,
    *,
    data_inicio: str,
    data_fim: str,
    filial_id: str | None = None,
    excluir_encerramento: bool = False,
    historico_encerramento: str = ENCERRAMENTO_HISTORICO,
) -> dict[str, dict[str, Any]]:
    params: dict[str, Any] = {
        "dataInicio": data_inicio,
        "dataFim": data_fim,
    }
    if filial_id:
        params["filialId"] = filial_id
    if excluir_encerramento:
        params["excluirEncerramento"] = "true"
        params["historicoEncerramento"] = historico_encerramento
    payload = api_get(api_url, api_key, "/api/v1/executivo/contabilidade/sintetico", params)
    return {str(row.get("conta_id")): row for row in payload.get("contas", [])}


def calculate_dre(accounts: dict[str, dict[str, Any]]) -> dict[str, float]:
    values: dict[str, float] = {}
    for line in DRE_LINES:
        if line.source:
            values[line.key] = account_value(accounts, line.source)
    for line in DRE_LINES:
        if line.formula:
            values[line.key] = line.formula(values)
    return values


def calculate_indicators(
    year: int,
    dre: dict[str, float],
    bp_accounts: dict[str, dict[str, Any]],
) -> dict[str, float]:
    ativo = bp_value(bp_accounts, "1")
    ativo_circ = bp_value(bp_accounts, "11")
    disponivel = bp_value(bp_accounts, "111")
    estoques = bp_value(bp_accounts, "115")
    realizavel_lp = bp_value(bp_accounts, "121")
    passivo_total = bp_value(bp_accounts, "2")
    passivo_circ = bp_value(bp_accounts, "21")
    passivo_nao_circ = bp_value(bp_accounts, "22")
    ativo_nao_circ = bp_value(bp_accounts, "12")
    patrimonio = bp_value(bp_accounts, "23")
    passivo_exigivel = passivo_circ + passivo_nao_circ
    emprestimos = bp_value(bp_accounts, "211104")
    receita_liquida = dre.get("receita_liquida", 0.0)
    resultado = dre.get("resultado_exercicio", 0.0)
    ebitda = dre.get("ebitda", 0.0)

    return {
        "ano": float(year),
        "receita_liquida": receita_liquida,
        "lucro_bruto": dre.get("lucro_bruto", 0.0),
        "margem_bruta": safe_ratio(dre.get("lucro_bruto", 0.0), receita_liquida),
        "ebitda": ebitda,
        "margem_ebitda": safe_ratio(ebitda, receita_liquida),
        "resultado_exercicio": resultado,
        "margem_liquida": safe_ratio(resultado, receita_liquida),
        "liquidez_corrente": safe_ratio(ativo_circ, passivo_circ),
        "liquidez_imediata": safe_ratio(disponivel, passivo_circ),
        "liquidez_seca": safe_ratio(ativo_circ - estoques, passivo_circ),
        "liquidez_geral": safe_ratio(ativo_circ + realizavel_lp, passivo_exigivel),
        "liquidez_geral_modelo": safe_ratio(ativo_circ + ativo_nao_circ, passivo_exigivel),
        "endividamento": safe_ratio(passivo_exigivel, patrimonio),
        "endividamento_geral": safe_ratio(passivo_exigivel, ativo),
        "endividamento_modelo": safe_ratio(passivo_total, patrimonio),
        "emprestimos_ebitda": safe_ratio(emprestimos, ebitda),
        "roa": safe_ratio(resultado, ativo),
        "roe": safe_ratio(resultado, patrimonio),
        "ativo_total": ativo,
        "passivo_total": passivo_total,
        "passivo_exigivel": passivo_exigivel,
        "patrimonio_liquido": patrimonio,
        "emprestimos": emprestimos,
    }


def assert_api_filter_supported(accounts: dict[str, dict[str, Any]], allow_without_filter: bool) -> None:
    probe = accounts.get("311102")
    if not probe:
        return
    debits = number(probe.get("debitos"))
    credits = number(probe.get("creditos"))
    saldo = number(probe.get("saldo"))
    likely_not_filtered = debits > 1_000 and credits > 1_000 and abs(saldo) < 1 and abs(debits - credits) < 1
    if likely_not_filtered and not allow_without_filter:
        raise RuntimeError(
            "A API em execução ainda não aplicou excluirEncerramento=true. "
            "Reinicie o serviço/container da ActionAPI para carregar a alteração em "
            "packages/api/src/services/executivo.js, ou rode temporariamente com "
            "--permitir-api-sem-filtro para diagnóstico."
        )


def read_model_values(path: Path, years: list[int]) -> dict[str, dict[str, dict[int, float]]]:
    if not path.exists():
        return {"dre": {}, "bp": {}, "indicadores": {}}
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {"dre": {}, "bp": {}, "indicadores": {}}

    if "SGA_DRE Comparativa Exercicio" in wb.sheetnames:
        ws = wb["SGA_DRE Comparativa Exercicio"]
        year_cols = {int(ws.cell(8, col).value): col for col in (5, 8, 11, 14, 16) if ws.cell(8, col).value}
        for row in range(9, 72):
            label = ws.cell(row, 3).value or ws.cell(row, 2).value
            if not label:
                continue
            values: dict[int, float] = {}
            for year in years:
                col = year_cols.get(year)
                if col:
                    values[year] = number(ws.cell(row, col).value)
            if values:
                result["dre"][str(label)] = values

    if "SGA_BP" in wb.sheetnames:
        ws = wb["SGA_BP"]
        year_cols = {int(ws.cell(5, col).value): col for col in range(4, 9) if ws.cell(5, col).value}
        for row in list(range(6, 59)) + [66, 70, 73, 77, 80, 83, 86, 90, 94, 95]:
            label = ws.cell(row, 3).value or ws.cell(row, 2).value
            if not label:
                continue
            values = {}
            for year in years:
                col = year_cols.get(year)
                if col:
                    values[year] = number(ws.cell(row, col).value)
            if values:
                result["bp"][str(label)] = values
    wb.close()
    return result


def append_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_money_row(ws, row: int, start_col: int, end_col: int, percent_cols: set[int] | None = None) -> None:
    percent_cols = percent_cols or set()
    for col in range(start_col, end_col + 1):
        ws.cell(row, col).number_format = PERCENT if col in percent_cols else MONEY


def create_panel(wb: Workbook, years: list[int], indicators: dict[int, dict[str, float]], generated_at: str) -> None:
    ws = wb.active
    ws.title = "Painel"
    last = max(years)
    current = indicators[last]
    prev = indicators.get(last - 1, {})
    title_block(ws, "DRE E BALANÇO PATRIMONIAL", f"Análise anual {years[0]}–{years[-1]} via ActionAPI | gerado em {generated_at}")

    kpis = [
        ("Receita Operacional Líquida", current["receita_liquida"], "money"),
        ("Lucro Bruto", current["lucro_bruto"], "money"),
        ("Margem Bruta", current["margem_bruta"], "percent"),
        ("EBITDA", current["ebitda"], "money"),
        ("Margem EBITDA", current["margem_ebitda"], "percent"),
        ("Resultado do Exercício", current["resultado_exercicio"], "money"),
        ("Margem Líquida", current["margem_liquida"], "percent"),
        ("Liquidez Corrente", current["liquidez_corrente"], "decimal"),
        ("Liquidez Geral (técnica)", current["liquidez_geral"], "decimal"),
        ("Capital de Terceiros / PL", current["endividamento"], "decimal"),
        ("Endividamento Geral", current["endividamento_geral"], "percent"),
        ("Empréstimos / EBITDA", current["emprestimos_ebitda"], "decimal"),
        ("ROA (Resultado DRE / Ativo)", current["roa"], "percent"),
        ("ROE (Resultado DRE / PL)", current["roe"], "percent"),
    ]
    append_header(ws, ["Indicador", str(last), f"Variação vs {last - 1}" if prev else "Variação"])
    for label, value, kind in kpis:
        old = prev.get({
            "Receita Operacional Líquida": "receita_liquida",
            "Lucro Bruto": "lucro_bruto",
            "Margem Bruta": "margem_bruta",
            "EBITDA": "ebitda",
            "Margem EBITDA": "margem_ebitda",
            "Resultado do Exercício": "resultado_exercicio",
            "Margem Líquida": "margem_liquida",
            "Liquidez Corrente": "liquidez_corrente",
            "Liquidez Geral (técnica)": "liquidez_geral",
            "Capital de Terceiros / PL": "endividamento",
            "Endividamento Geral": "endividamento_geral",
            "Empréstimos / EBITDA": "emprestimos_ebitda",
            "ROA (Resultado DRE / Ativo)": "roa",
            "ROE (Resultado DRE / PL)": "roe",
        }[label], 0.0)
        variation = value / old - 1 if old and kind == "money" else value - old if old else 0.0
        ws.append([label, value, variation])
        row = ws.max_row
        ws.cell(row, 2).number_format = PERCENT if kind == "percent" else MONEY if kind == "money" else DECIMAL
        ws.cell(row, 3).number_format = PERCENT if kind in {"money", "percent"} else DECIMAL

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18

    start = ws.max_row + 3
    append_header(ws, ["Ano", "Receita Líquida", "Lucro Bruto", "EBITDA", "Resultado do Exercício"])
    chart_header = ws.max_row
    for year in years:
        row = indicators[year]
        ws.append([year, row["receita_liquida"], row["lucro_bruto"], row["ebitda"], row["resultado_exercicio"]])
        style_money_row(ws, ws.max_row, 2, 5)
    chart = LineChart()
    chart.title = "Evolução DRE"
    chart.height = 8
    chart.width = 18
    chart.add_data(Reference(ws, min_col=2, max_col=5, min_row=chart_header, max_row=ws.max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=chart_header + 1, max_row=ws.max_row))
    ws.add_chart(chart, f"G{start}")


def create_dre_exercicios(wb: Workbook, years: list[int], dre_by_year: dict[int, dict[str, float]]) -> None:
    ws = wb.create_sheet("DRE Exercícios")
    headers = ["Seção", "Linha"] + [str(year) for year in years]
    headers += [f"AH {year}" for year in years[1:]]
    headers += [f"AV {year}" for year in years]
    append_header(ws, headers)
    line_by_key = {line.key: line for line in DRE_LINES}
    for key in VISIBLE_DRE_KEYS:
        line = line_by_key[key]
        values = [dre_by_year[year].get(key, 0.0) for year in years]
        ah = [
            safe_ratio(values[idx], values[idx - 1]) - 1 if values[idx - 1] else 0.0
            for idx in range(1, len(values))
        ]
        av = [
            safe_ratio(dre_by_year[year].get(key, 0.0), dre_by_year[year].get(line.percent_base or "receita_liquida", 0.0))
            if line.percent_base else 0.0
            for year in years
        ]
        ws.append([line.section, ("  " * line.level) + line.label] + values + ah + av)
        row = ws.max_row
        for col in range(3, 3 + len(years)):
            ws.cell(row, col).number_format = MONEY
        for col in range(3 + len(years), 3 + len(years) + len(ah) + len(av)):
            ws.cell(row, col).number_format = PERCENT
        if line.bold:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).font = Font(bold=True, color=WHITE)
                ws.cell(row, col).fill = PatternFill("solid", fgColor=BLUE if line.section != "Resultado" else GREEN)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 58


def create_dre_filial(
    wb: Workbook,
    years: list[int],
    dre_by_year: dict[int, dict[str, float]],
    branch_dre_by_year: dict[int, dict[str, dict[str, float]]],
) -> None:
    # Uma linha por (ano, conta), com "Ano" na coluna A. Use o filtro nativo do
    # Excel no cabeçalho dessa coluna para selecionar o(s) ano(s) de exercício —
    # equivalente ao seletor de ano da planilha-modelo, sem depender de fórmulas
    # cruzando abas.
    ws = wb.create_sheet("DRE por Filial")
    headers = ["Ano", "Seção", "Linha", "Consolidado", "AV Consolidado"]
    for _branch_id, branch_name in BRANCHES:
        headers += [branch_name, f"% {branch_name}/Consolidado"]
    append_header(ws, headers)
    line_by_key = {line.key: line for line in DRE_LINES}
    for year in years:
        consolidated = dre_by_year[year]
        receita_liquida = consolidated.get("receita_liquida", 0.0)
        branch_dre = branch_dre_by_year[year]
        for key in VISIBLE_DRE_KEYS:
            line = line_by_key[key]
            value = consolidated.get(key, 0.0)
            row = [
                year,
                line.section,
                ("  " * line.level) + line.label,
                value,
                safe_ratio(value, receita_liquida) if line.percent_base else 0.0,
            ]
            for branch_id, _branch_name in BRANCHES:
                branch_value = branch_dre.get(branch_id, {}).get(key, 0.0)
                row += [branch_value, safe_ratio(branch_value, value)]
            ws.append(row)
            excel_row = ws.max_row
            for col in range(4, ws.max_column + 1, 2):
                ws.cell(excel_row, col).number_format = MONEY
            for col in range(5, ws.max_column + 1, 2):
                ws.cell(excel_row, col).number_format = PERCENT
            if line.bold:
                for col in range(1, ws.max_column + 1):
                    ws.cell(excel_row, col).font = Font(bold=True, color=WHITE)
                    ws.cell(excel_row, col).fill = PatternFill("solid", fgColor=BLUE if line.section != "Resultado" else GREEN)
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 58


def create_bp_sheet(wb: Workbook, years: list[int], bp_by_year: dict[int, dict[str, dict[str, Any]]]) -> None:
    ws = wb.create_sheet("Balanço Patrimonial")
    append_header(ws, ["Conta", "Nomenclatura"] + [str(year) for year in years])
    for line in BP_LINES:
        values = [bp_value(bp_by_year[year], line.account) for year in years]
        ws.append([line.account, ("  " * line.level) + line.label] + values)
        row = ws.max_row
        for col in range(3, 3 + len(years)):
            ws.cell(row, col).number_format = MONEY
        if line.bold:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).font = Font(bold=True, color=WHITE)
                ws.cell(row, col).fill = PatternFill("solid", fgColor=NAVY if len(line.account) == 1 else BLUE)
    diff_row = ws.max_row + 2
    ws.cell(diff_row, 1, "CHECK")
    ws.cell(diff_row, 2, "Ativo - Passivo")
    for idx, year in enumerate(years, start=3):
        ws.cell(diff_row, idx, bp_value(bp_by_year[year], "1") - bp_value(bp_by_year[year], "2"))
        ws.cell(diff_row, idx).number_format = MONEY
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 62


def create_indicators_sheet(wb: Workbook, years: list[int], indicators: dict[int, dict[str, float]]) -> None:
    ws = wb.create_sheet("Indicadores")
    append_header(ws, ["Indicador"] + [str(year) for year in years])
    for label, key, kind, _formula_text in INDICATOR_DEFINITIONS:
        ws.append([label] + [indicators[year].get(key, 0.0) for year in years])
        row = ws.max_row
        fmt = MONEY if kind == "money" else PERCENT if kind == "percent" else DECIMAL
        for col in range(2, 2 + len(years)):
            ws.cell(row, col).number_format = fmt
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 42


def create_validation_sheet(
    wb: Workbook,
    years: list[int],
    dre_by_year: dict[int, dict[str, float]],
    bp_by_year: dict[int, dict[str, dict[str, Any]]],
    model_values: dict[str, dict[str, dict[int, float]]],
    model_path: Path,
) -> None:
    ws = wb.create_sheet("Validação")
    ws.append(["Fonte", str(model_path) if model_path.exists() else "Planilha-modelo não encontrada"])
    ws.append(["Regra DRE", f"ActionAPI /executivo/contabilidade/sintetico com excluirEncerramento=true e HIST_HIS <> {ENCERRAMENTO_HISTORICO}"])
    ws.append(["Regra BP", "ActionAPI /executivo/contabilidade/sintetico acumulado até 31/12 de cada ano, com sinal invertido para Passivo/PL"])
    ws.append([])
    append_header(ws, ["Bloco", "Linha", "Ano", "Valor API", "Valor Modelo", "Diferença"])

    dre_model = model_values.get("dre", {})
    label_by_key = {line.key: line.label for line in DRE_LINES}
    comparable_labels = {
        "Receita Bruta com Vendas": "Receita Bruta Com Vendas",
        "Deduções da Receita Bruta": "Deducoes Da Receita Bruta De Vendas De Mercadorias",
        "Receita Operacional Líquida": "RECEITA OPERACIONAL LÍQUIDA",
        "Lucro Bruto / Margem Bruta": "LUCRO BRUTO (Margem Bruta)",
        "Despesas Administrativas e Comerciais": "Despesas Administrativas E Comerciais",
        "Resultado Financeiro": "Resultado Financeiro",
        "PCLD": "PCLD",
        "Resultado do Exercício": "RESULTADO DO EXERCÍCIO",
        "EBITDA": "EBITDA",
    }
    for key in [
        "receita_bruta",
        "deducoes_receita",
        "receita_liquida",
        "lucro_bruto",
        "despesas_adm_com",
        "resultado_financeiro",
        "pcld",
        "resultado_exercicio",
        "ebitda",
    ]:
        label = label_by_key[key]
        model_label = comparable_labels.get(label, label)
        for year in years:
            api_val = dre_by_year[year].get(key, 0.0)
            model_val = dre_model.get(model_label, {}).get(year)
            ws.append(["DRE", label, year, api_val, model_val, api_val - number(model_val)])
            for col in (4, 5, 6):
                ws.cell(ws.max_row, col).number_format = MONEY

    bp_model = model_values.get("bp", {})
    for account, label in [("1", "Ativo"), ("2", "Passivo"), ("11", "Circulante"), ("21", "Passivo Circulante"), ("23", "Patrimonio Liquido")]:
        for year in years:
            api_val = bp_value(bp_by_year[year], account)
            model_val = bp_model.get(label, {}).get(year)
            ws.append(["BP", f"{account} - {label}", year, api_val, model_val, api_val - number(model_val)])
            for col in (4, 5, 6):
                ws.cell(ws.max_row, col).number_format = MONEY

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A5:F{ws.max_row}"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 10
    for col in ("D", "E", "F"):
        ws.column_dimensions[col].width = 18


def create_mapping_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Mapeamento")
    append_header(ws, ["Tipo", "Linha", "Conta/Fórmula", "Sinal", "Observação"])
    for line in DRE_LINES:
        if line.section == "Auxiliar":
            continue
        source = line.source.account if line.source else "fórmula"
        sign = line.source.sign if line.source else ""
        note = "Crédito positivo (valor = crédito - débito)" if sign == "credit" else "Débito positivo (valor = débito - crédito)" if sign == "debit" else "Calculado a partir de linhas anteriores"
        ws.append(["DRE", line.label, source, sign, note])
    for line in BP_LINES:
        note = "Ativo usa saldo D-C; Passivo/PL inverte o sinal para apresentação gerencial."
        ws.append(["BP", line.label, line.account, "", note])
    style_sheet(ws)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 80


def create_formula_audit_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Auditoria Fórmulas")
    append_header(ws, ["Item", "Fórmula usada no relatório", "Fórmula observada no modelo", "Risco", "Decisão"])
    rows = [
        (
            "ROA",
            "Resultado do Exercício da DRE / Ativo Total",
            "Lucros ou Prejuízos Acumulados do BP / Ativo Total",
            "Alto",
            "Usar Resultado do Exercício da DRE. O saldo acumulado do BP pode ficar zerado por encerramento/distribuição e distorcer o retorno.",
        ),
        (
            "ROE",
            "Resultado do Exercício da DRE / Patrimônio Líquido",
            "Lucros ou Prejuízos Acumulados do BP / Patrimônio Líquido; rótulo do modelo menciona Ativo Total",
            "Alto",
            "Usar Resultado do Exercício da DRE / PL. A aba de indicadores mostra este critério como principal.",
        ),
        (
            "Liquidez Geral",
            "(Ativo Circulante + Realizável a Longo Prazo) / (Passivo Circulante + Passivo Não Circulante)",
            "(Ativo Circulante + Ativo Não Circulante) / (Passivo Circulante + Passivo Não Circulante)",
            "Médio",
            "Mostrar a fórmula técnica como principal e manter o critério do modelo em linha separada para reconciliação.",
        ),
        (
            "Endividamento",
            "Capital de Terceiros / PL = (Passivo Circulante + Passivo Não Circulante) / PL; também mostra Exigível / Ativo",
            "Grupo contábil 2 / PL, onde o grupo 2 inclui Passivo + Patrimônio Líquido",
            "Médio/Alto",
            "Mostrar o critério técnico como principal e manter o critério do modelo em linha separada.",
        ),
        (
            "Impostos nas Vendas em Geral",
            "Deduções Sobre Vendas Mercadorias - Devoluções de Vendas",
            "Deduções Sobre Vendas Mercadorias sem subtrair Devoluções de Vendas",
            "Médio",
            "Corrigir a sublinha para evitar dupla contagem visual. O total de Deduções da Receita permanece puxado da conta 3112.",
        ),
        (
            "Resultado antes dos impostos",
            "Duas linhas separadas: Resultado Contábil antes dos Impostos (com PCLD) e Resultado Gerencial antes dos Impostos",
            "Duas linhas com o mesmo rótulo no modelo",
            "Baixo/Médio",
            "Renomear a segunda linha para deixar claro que é o resultado gerencial sem efeito de PCLD.",
        ),
        (
            "Perda de PCLD (4211250060) no Resultado Contábil",
            "Resultado Contábil 'com PCLD' desconta apenas Constituição + Reversão; a Perda já está dentro das Despesas Adm/Comerciais",
            "Modelo (célula r59) desconta o PCLD inteiro (Perda + Constituição + Reversão), embora a Perda já esteja nas Despesas — dupla contagem quando a Perda ≠ 0 (ex.: 2021 = R$ 2,29 mi; zero em 2022-2025)",
            "Médio",
            "Corrigido: a Perda de PCLD deixa de ser subtraída em dobro na linha de Resultado Contábil. O Resultado do Exercício/Gerencial (e ROA/ROE/EBITDA) NÃO muda — mantém o critério do modelo de tratar a perda realizada como despesa real.",
        ),
    ]
    for row in rows:
        ws.append(row)
        risk = row[3]
        fill = RED if risk == "Alto" else ORANGE if "Médio" in risk else GREEN
        ws.cell(ws.max_row, 4).fill = PatternFill("solid", fgColor=fill)
        ws.cell(ws.max_row, 4).font = Font(color=WHITE, bold=True)
    style_sheet(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 72
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 90
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_calculation_memory_sheet(
    wb: Workbook,
    years: list[int],
    dre_by_year: dict[int, dict[str, float]],
    bp_by_year: dict[int, dict[str, dict[str, Any]]],
    indicators: dict[int, dict[str, float]],
    branch_dre_by_year: dict[int, dict[str, dict[str, float]]],
) -> None:
    ws = wb.create_sheet("Memória de Cálculo")
    headers = [
        "Bloco",
        "Tipo de valor",
        "Ano",
        "Filial",
        "Linha/Indicador",
        "Valor",
        "Fórmula / Critério",
        "Numerador / Componentes",
        "Denominador / Base",
        "Fonte de dados",
        "Observação",
    ]
    append_header(ws, headers)
    line_by_key = {line.key: line for line in DRE_LINES}

    def add_row(
        bloco: str,
        tipo: str,
        year: int,
        filial: str,
        label: str,
        value: float,
        formula: str,
        numerator: str,
        denominator: str,
        source: str,
        note: str = "",
        *,
        kind: str = "money",
    ) -> None:
        ws.append([bloco, tipo, year, filial, label, value, formula, numerator, denominator, source, note])
        row = ws.max_row
        ws.cell(row, 6).number_format = PERCENT if kind == "percent" else MONEY if kind == "money" else DECIMAL
        for col in range(7, 12):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")

    for year in years:
        dre_values = dre_by_year[year]
        dre_source = (
            "/api/v1/executivo/contabilidade/sintetico "
            f"dataInicio={year}-01-01; dataFim={year}-12-31; "
            f"excluirEncerramento=true; HIST_HIS<>{ENCERRAMENTO_HISTORICO}"
        )
        for key in VISIBLE_DRE_KEYS:
            line = line_by_key[key]
            value = dre_values.get(key, 0.0)
            add_row(
                "DRE",
                "Valor",
                year,
                "Consolidado",
                line.label,
                value,
                dre_line_formula_text(line),
                dre_line_components_text(line, dre_values),
                "",
                dre_source,
                "Valor apresentado na aba DRE Exercícios.",
                kind="money",
            )

            if line.percent_base:
                base_key = line.percent_base
                base_label = line_by_key.get(base_key, DreLine(base_key, base_key, "")).label
                base_value = dre_values.get(base_key, 0.0)
                add_row(
                    "DRE",
                    "Análise vertical",
                    year,
                    "Consolidado",
                    line.label,
                    safe_ratio(value, base_value),
                    "AV = valor da linha / base de comparação",
                    f"{line.label}={format_formula_value(value)}",
                    f"{base_label}={format_formula_value(base_value)}",
                    "Valores calculados da própria DRE anual",
                    "Base padrão: Receita Operacional Líquida. Linhas de resultado sem base ficam sem AV.",
                    kind="percent",
                )

        for idx, year_current in enumerate(years):
            if idx == 0 or year_current != year:
                continue
            previous_year = years[idx - 1]
            previous_values = dre_by_year[previous_year]
            for key in VISIBLE_DRE_KEYS:
                line = line_by_key[key]
                current_value = dre_values.get(key, 0.0)
                previous_value = previous_values.get(key, 0.0)
                add_row(
                    "DRE",
                    "Análise horizontal",
                    year,
                    "Consolidado",
                    line.label,
                    safe_ratio(current_value, previous_value) - 1 if previous_value else 0.0,
                    "AH = (valor do ano atual / valor do ano anterior) - 1",
                    f"{year}={format_formula_value(current_value)}",
                    f"{previous_year}={format_formula_value(previous_value)}",
                    "Valores calculados da própria DRE anual",
                    "Quando o ano anterior é zero, o relatório retorna 0 para evitar divisão inválida.",
                    kind="percent",
                )

    for year in years:
        branch_source = (
            "/api/v1/executivo/contabilidade/sintetico "
            f"dataInicio={year}-01-01; dataFim={year}-12-31; filialId={{filial}}; "
            f"excluirEncerramento=true; HIST_HIS<>{ENCERRAMENTO_HISTORICO}"
        )
        consolidated = dre_by_year[year]
        for branch_id, branch_name in BRANCHES:
            branch_values = branch_dre_by_year.get(year, {}).get(branch_id, {})
            for key in VISIBLE_DRE_KEYS:
                line = line_by_key[key]
                branch_value = branch_values.get(key, 0.0)
                consolidated_value = consolidated.get(key, 0.0)
                add_row(
                    "DRE por Filial",
                    "Valor",
                    year,
                    f"{branch_id} - {branch_name}",
                    line.label,
                    branch_value,
                    dre_line_formula_text(line),
                    dre_line_components_text(line, branch_values),
                    "",
                    branch_source.replace("{filial}", branch_id),
                    "Valor apresentado na aba DRE por Filial.",
                    kind="money",
                )
                add_row(
                    "DRE por Filial",
                    "% sobre consolidado",
                    year,
                    f"{branch_id} - {branch_name}",
                    line.label,
                    safe_ratio(branch_value, consolidated_value),
                    "% filial = valor da filial / valor consolidado da mesma linha",
                    f"{branch_name}={format_formula_value(branch_value)}",
                    f"Consolidado={format_formula_value(consolidated_value)}",
                    "Valores calculados da DRE por filial e DRE consolidada",
                    "Quando o consolidado é zero, o relatório retorna 0 para evitar divisão inválida.",
                    kind="percent",
                )

    for year in years:
        bp_source = (
            "/api/v1/executivo/contabilidade/sintetico "
            f"dataInicio={DATA_INICIO_PLANO_ATUAL}; dataFim={year}-12-31"
        )
        accounts = bp_by_year[year]
        for line in BP_LINES:
            value = bp_value(accounts, line.account)
            raw = accounts.get(line.account, {})
            formula = (
                "Ativo: saldo D-C. Passivo/PL: sinal invertido para apresentação positiva."
                if line.account.startswith("2")
                else "Saldo acumulado D-C da conta."
            )
            add_row(
                "Balanço Patrimonial",
                "Valor",
                year,
                "Consolidado",
                f"{line.account} - {line.label}",
                value,
                formula,
                f"débitos={format_formula_value(number(raw.get('debitos')))}; créditos={format_formula_value(number(raw.get('creditos')))}",
                f"saldo D-C={format_formula_value(number(raw.get('saldo')))}",
                bp_source,
                "Valor apresentado na aba Balanço Patrimonial.",
                kind="money",
            )

    for year in years:
        row = indicators[year]
        for label, key, kind, formula_text in INDICATOR_DEFINITIONS:
            value = row.get(key, 0.0)
            add_row(
                "Indicadores",
                "Valor",
                year,
                "Consolidado",
                label,
                value,
                formula_text,
                "",
                "",
                "DRE + Balanço Patrimonial calculados pela ActionAPI",
                "Valor apresentado na aba Indicadores.",
                kind=kind,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 20,
        "B": 22,
        "C": 10,
        "D": 22,
        "E": 58,
        "F": 18,
        "G": 72,
        "H": 62,
        "I": 52,
        "J": 68,
        "K": 78,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_api_accounts_sheet(wb: Workbook, year: int, accounts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Contas API DRE")
    append_header(ws, ["Conta", "Descrição", "Analítica", "Débitos", "Créditos", "Saldo D-C", "Valor DRE crédito+", "Valor DRE débito+"])
    for account_id, row in sorted(accounts.items()):
        if not (account_id.startswith("3") or account_id.startswith("4")):
            continue
        saldo = number(row.get("saldo"))
        ws.append([
            account_id,
            row.get("descricao"),
            bool(row.get("analitica")),
            number(row.get("debitos")),
            number(row.get("creditos")),
            saldo,
            -saldo,
            saldo,
        ])
        for col in range(4, 9):
            ws.cell(ws.max_row, col).number_format = MONEY
    style_sheet(ws)
    ws["J1"] = f"Base DRE {year}: lançamentos de encerramento excluídos"


def generate_report(args: argparse.Namespace) -> Path:
    years = parse_years(args.anos or prompt_for_years())

    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    output = (
        Path(args.arquivo).resolve()
        if args.arquivo
        else ROOT / "relatorios" / f"relatorio-dre-{years[0]}-{years[-1]}.xlsx"
    )
    model_path = Path(args.modelo).resolve()
    api_key = first_api_key()

    dre_accounts_by_year: dict[int, dict[str, dict[str, Any]]] = {}
    bp_accounts_by_year: dict[int, dict[str, dict[str, Any]]] = {}
    dre_by_year: dict[int, dict[str, float]] = {}

    print(f"[dre] consultando ActionAPI em {args.api_url}...", flush=True)
    for year in years:
        dre_accounts = fetch_synthetic(
            args.api_url,
            api_key,
            data_inicio=f"{year}-01-01",
            data_fim=f"{year}-12-31",
            excluir_encerramento=True,
            historico_encerramento=args.historico_encerramento,
        )
        if year == max(years):
            assert_api_filter_supported(dre_accounts, args.permitir_api_sem_filtro)
        bp_accounts = fetch_synthetic(
            args.api_url,
            api_key,
            data_inicio=DATA_INICIO_PLANO_ATUAL,
            data_fim=f"{year}-12-31",
        )
        dre_accounts_by_year[year] = dre_accounts
        bp_accounts_by_year[year] = bp_accounts
        dre_by_year[year] = calculate_dre(dre_accounts)
        print(f"[dre] {year}: {len(dre_accounts)} contas DRE, {len(bp_accounts)} contas BP", flush=True)

    branch_dre_by_year: dict[int, dict[str, dict[str, float]]] = {}
    for year in years:
        branch_dre_by_year[year] = {}
        for branch_id, branch_name in BRANCHES:
            accounts = fetch_synthetic(
                args.api_url,
                api_key,
                data_inicio=f"{year}-01-01",
                data_fim=f"{year}-12-31",
                filial_id=branch_id,
                excluir_encerramento=True,
                historico_encerramento=args.historico_encerramento,
            )
            branch_dre_by_year[year][branch_id] = calculate_dre(accounts)
            print(f"[dre] filial {branch_id} {branch_name} {year}: {len(accounts)} contas", flush=True)

    indicators = {
        year: calculate_indicators(year, dre_by_year[year], bp_accounts_by_year[year])
        for year in years
    }
    model_values = {"dre": {}, "bp": {}, "indicadores": {}}
    if not args.sem_validar_modelo:
        print(f"[dre] lendo modelo para validação: {model_path}", flush=True)
        model_values = read_model_values(model_path, years)

    wb = Workbook()
    create_panel(wb, years, indicators, generated_at)
    create_dre_exercicios(wb, years, dre_by_year)
    create_dre_filial(wb, years, dre_by_year, branch_dre_by_year)
    create_bp_sheet(wb, years, bp_accounts_by_year)
    create_indicators_sheet(wb, years, indicators)
    create_validation_sheet(wb, years, dre_by_year, bp_accounts_by_year, model_values, model_path)
    create_mapping_sheet(wb)
    create_formula_audit_sheet(wb)
    create_calculation_memory_sheet(wb, years, dre_by_year, bp_accounts_by_year, indicators, branch_dre_by_year)
    create_api_accounts_sheet(wb, max(years), dre_accounts_by_year[max(years)])

    required = [
        "Painel",
        "DRE Exercícios",
        "DRE por Filial",
        "Balanço Patrimonial",
        "Indicadores",
        "Validação",
        "Mapeamento",
        "Auditoria Fórmulas",
        "Memória de Cálculo",
        "Contas API DRE",
    ]
    save(wb, output, required)
    print(f"[dre] arquivo: {output}")
    return output


def main() -> None:
    try:
        generate_report(parse_args())
    except Exception as exc:
        print(f"[dre] erro: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
