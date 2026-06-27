#!/usr/bin/env python3
r"""Relatório DRE/BP no estilo do controller, 100% gerado a partir da ActionAPI.

Este script NÃO depende de nenhuma planilha-modelo: todas as abas são construídas
do zero com os saldos vindos da ActionAPI (endpoint
/api/v1/executivo/contabilidade/sintetico). O layout segue o conceito da planilha
do controller (DRE por exercício, comparativo por filial, balanço/indicadores,
planejamento e balancete geral), mas sem fórmulas vivas — os valores são números
estáticos calculados em Python.

Correções de critério já incorporadas (ver relatorio_dre.py):
  - DRE sem o lançamento de encerramento (HIST_HIS=1000191);
  - Resultado Contábil sem dupla contagem da perda de PCLD;
  - ROA/ROE pelo Resultado do Exercício da DRE;
  - Liquidez Geral e Endividamento no critério técnico;
  - A.V. (análise vertical) só nas linhas de subtotal/resultado.

Exemplos:

    .\.venv\Scripts\python.exe scripts\dre_controller.py
    .\.venv\Scripts\python.exe scripts\dre_controller.py --anos 2021-2025
"""

from __future__ import annotations

import argparse
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    print(
        "Dependência ausente. Execute:\n"
        "  py -m pip install -r scripts/requirements-relatorio-contas-pagar.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)

try:
    from relatorio_dre import (
        BP_LINES,
        BRANCHES,
        DATA_INICIO_PLANO_ATUAL,
        DEFAULT_API_URL,
        DRE_LINES,
        ENCERRAMENTO_HISTORICO,
        VISIBLE_DRE_KEYS,
        api_get,
        assert_api_filter_supported,
        bp_value,
        calculate_dre,
        calculate_indicators,
        dre_line_components_text,
        dre_line_formula_text,
        fetch_synthetic,
        first_api_key,
        format_formula_value,
        number,
        parse_years,
        safe_ratio,
    )
except ModuleNotFoundError:
    from scripts.relatorio_dre import (
        BP_LINES,
        BRANCHES,
        DATA_INICIO_PLANO_ATUAL,
        DEFAULT_API_URL,
        DRE_LINES,
        ENCERRAMENTO_HISTORICO,
        VISIBLE_DRE_KEYS,
        api_get,
        assert_api_filter_supported,
        bp_value,
        calculate_dre,
        calculate_indicators,
        dre_line_components_text,
        dre_line_formula_text,
        fetch_synthetic,
        first_api_key,
        format_formula_value,
        number,
        parse_years,
        safe_ratio,
    )


ROOT = Path(__file__).resolve().parents[1]

HEADER_FILL = PatternFill("solid", fgColor="143D59")
GROUP_HEADER_FILL = PatternFill("solid", fgColor="0B5D1E")
SUBTOTAL_FILL = PatternFill("solid", fgColor="DFF1E4")
ZEBRA_FILL = PatternFill("solid", fgColor="F7FAFC")
WARNING_FILL = PatternFill("solid", fgColor="FFF4CE")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN_SIDE = Side(style="thin", color="D7DEE8")
HEADER_SIDE = Side(style="thin", color="FFFFFF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_BORDER = Border(left=HEADER_SIDE, right=HEADER_SIDE, top=HEADER_SIDE, bottom=HEADER_SIDE)
BOLD_FONT = Font(name="Aptos", bold=True, color="17324D")
MUTED_FONT = Font(name="Aptos", color="64748B")
WHITE_FONT = Font(name="Aptos", color="FFFFFF", bold=True)
MONEY_FMT = '#,##0.00;[Red]-#,##0.00;""'
PERCENT_FMT = '0.00%;[Red]-0.00%;""'
NUMBER_FMT = '#,##0.00'

DRE_LINE_BY_KEY = {line.key: line for line in DRE_LINES}

# =============================================================================
# REGRA DE A.V. (ANÁLISE VERTICAL) — FONTE ÚNICA, NÃO DUPLICAR.
#
# No modelo do controller, a coluna "(%) A.V." (e a "(%) no Consolidado" da aba
# comparativa por filial) SÓ é calculada nas linhas de subtotal/resultado da
# DRE (as listadas em AV_KEYS abaixo, confirmado célula a célula na aba
# "SGA_DRE Comparativa Exercicio" da planilha-modelo). Em todas as linhas de
# detalhe (ex.: "Vendas de Mercadorias em Geral", "Despesas com RH Diretores")
# a célula correspondente fica em BRANCO no modelo — não é zero, é vazio.
#
# Essa MESMA regra/conjunto vale em TODAS as abas que mostram A.V. ou % de
# participação por linha da DRE: "DRE por Exercício" e "DRE Comparativa por
# Ano". Não recalcule a regra separadamente em cada aba — sempre chame
# av_for_dre_key(key, ...) e use seu retorno (None = deixar célula em branco).
# Se um dia for preciso adicionar uma nova aba com A.V./percentual por linha,
# reaproveite av_for_dre_key/AV_KEYS em vez de reescrever a condição.
# =============================================================================
AV_KEYS = {
    "receita_liquida",
    "custos_vendas",
    "lucro_bruto",
    "despesas_adm_com",
    "lucro_operacional",
    "resultado_financeiro",
    "pcld",
    "resultado_contabil_antes_impostos",
    "resultado_gerencial_antes_impostos",
    "resultado_exercicio",
    "provisoes_fiscais",
    "depreciacao",
    "ebitda",
    "margem_bruta_valor",
}

BALANCETE_BRANCH_NAMES = {
    "1": "Goiatuba",
    "2": "Campo Alegre",
    "3": "Gurupi",
    "4": "Lagoa",
    "5": "Porto",
    "6": "Araguari",
    "7": "Monte Carmelo",
    "8": "Alvorada",
    "9": "Piracanjuba",
}

# Nomes de mês em português, para rotular a coluna do ano parcial na aba
# Planejamento (ex.: "maio/2026" quando --data-fim corta em 31/05/2026).
# Hardcoded em vez de locale.setlocale("pt_BR") porque esse locale costuma
# não estar instalado no Windows, e o nome do mês não pode depender disso.
MESES_PT = [
    "", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]


@dataclass(frozen=True)
class BranchInfo:
    id: str
    nome: str
    situacao: str = "A"
    ativa: bool = True
    origem: str = "api"

    @property
    def label(self) -> str:
        return f"{self.nome} ({self.id})" if self.id not in str(self.nome) else self.nome


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Relatório DRE/BP no estilo do controller, gerado pela ActionAPI (sem template)."
    )
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--anos", help="Intervalo/lista de anos. Ex.: 2021-2025 ou 2023,2024,2025.")
    parser.add_argument(
        "--data-fim",
        help=(
            "Corta o exercício mais recente numa data específica em vez de 31/12, "
            "no formato brasileiro DD/MM/AAAA (ex.: --data-fim 31/05/2026 para um "
            "relatório parcial do ano corrente até 31/05/2026). Também aceita "
            "DDMMAAAA sem separador (ex.: 31052026). O ano dessa data deve ser o "
            "mais recente do intervalo; se não estiver em --anos, é adicionado "
            "automaticamente."
        ),
    )
    parser.add_argument("--arquivo", help="Arquivo .xlsx de saída.")
    parser.add_argument("--historico-encerramento", default=ENCERRAMENTO_HISTORICO)
    parser.add_argument(
        "--nao-interativo",
        action="store_true",
        help="Nao pergunta nada; usa defaults quando algum parametro nao for informado.",
    )
    parser.add_argument(
        "--permitir-api-sem-filtro",
        action="store_true",
        help="Permite continuar mesmo se a API em execução não aplicar excluirEncerramento=true.",
    )
    return parser.parse_args()


def complete_interactive_args(args: argparse.Namespace) -> argparse.Namespace:
    """No uso manual pergunta os anos e a data final; em serviço/Docker usa o default."""
    if args.nao_interativo or not sys.stdin.isatty():
        args.anos = args.anos or "2021-2025"
        return args
    while not args.anos:
        answer = input(
            "Quais exercícios deseja gerar? Use intervalo/lista, ex.: 2021-2025 ou 2023,2024,2025 [2021-2025]: "
        ).strip() or "2021-2025"
        try:
            parse_years(answer)
        except ValueError as exc:
            print(f"  Valor inválido: {exc}", flush=True)
            continue
        args.anos = answer

    if args.data_fim is None:
        while True:
            answer = input(
                "Data final do exercício mais recente, formato DD/MM/AAAA "
                "(Enter para fechar em 31/12 normalmente, sem corte): "
            ).strip()
            if not answer:
                args.data_fim = None
                break
            try:
                parse_br_date(answer)
            except ValueError as exc:
                print(f"  Valor inválido: {exc}", flush=True)
                continue
            args.data_fim = answer
            break

    return args


# --------------------------------------------------------------------------- #
# Filiais
# --------------------------------------------------------------------------- #
def fetch_branch_registry(api_url: str, api_key: str) -> tuple[list[BranchInfo], str]:
    """Busca filiais na ActionAPI; usa lista operacional antiga como fallback."""
    try:
        payload = api_get(api_url, api_key, "/api/v1/executivo/filiais", {})
        rows = payload.get("filiais", [])
        branches = [
            BranchInfo(
                id=str(row.get("id") or row.get("codigo") or "").strip(),
                nome=str(row.get("identificacao") or row.get("fantasia") or row.get("id") or "").strip(),
                situacao=str(row.get("situacao") or "").strip() or "SEM_STATUS",
                ativa=bool(row.get("ativa")),
                origem="api",
            )
            for row in rows
            if str(row.get("id") or row.get("codigo") or "").strip()
        ]
        branches.sort(key=lambda b: (int(b.id) if b.id.isdigit() else 999999, b.id))
        if branches:
            return branches, "ActionAPI /api/v1/executivo/filiais"
    except Exception as exc:
        print(
            f"[dre-controller] aviso: nao foi possivel consultar /executivo/filiais ({exc}); "
            "usando lista operacional padrao.",
            flush=True,
        )

    fallback = [BranchInfo(id=branch_id, nome=name, origem="fallback") for branch_id, name in BRANCHES]
    return fallback, "fallback BRANCHES em relatorio_dre.py; reinicie a API para usar /executivo/filiais"


def controller_visible_branches(branches: list[BranchInfo]) -> list[BranchInfo]:
    """Filiais exibidas nas abas de DRE.

    O layout antigo do controller tinha uma lista fixa de filiais. Aqui a lista
    vem do cadastro da API para que o consolidado seja conciliado com todas as
    filiais efetivamente presentes na base. Se a rota de cadastro cair, o
    fallback continua sendo a lista operacional antiga.
    """
    return sorted(branches, key=branch_sort_key)


def branch_sort_key(branch: BranchInfo) -> tuple[int, int, str]:
    inactive = (not branch.ativa) or branch.situacao.upper() == "I"
    numeric_id = int(branch.id) if branch.id.isdigit() else 999999
    return (1 if inactive else 0, numeric_id, branch.id)


def branch_display_name(branch: BranchInfo) -> str:
    return BALANCETE_BRANCH_NAMES.get(branch.id) or branch.nome or branch.id


def branch_report_name(branch: BranchInfo) -> str:
    base = branch.label
    if (not branch.ativa) or branch.situacao.upper() == "I":
        return f"{base} - inativa"
    return base


def inactive_branches_text(branches: list[BranchInfo], source: str) -> str:
    inactive = [branch for branch in branches if (not branch.ativa) or branch.situacao.upper() == "I"]
    if not inactive:
        return f"Filiais INATIVAS: nenhuma conforme {source}."
    names = ", ".join(f"{branch.nome} ({branch.id})" for branch in inactive)
    return f"Filiais INATIVAS: {names} conforme {source}."


# --------------------------------------------------------------------------- #
# Helpers de plano de contas / formatação
# --------------------------------------------------------------------------- #
def account_description(accounts_by_id: dict[str, dict[str, Any]], account_id: str) -> str | None:
    row = accounts_by_id.get(account_id)
    return str(row.get("descricao")) if row and row.get("descricao") is not None else None


def account_grau3(accounts_by_id: dict[str, dict[str, Any]], account_id: str) -> str | None:
    if len(account_id) < 3:
        return None
    return account_description(accounts_by_id, account_id[:3])


def account_natureza(accounts_by_id: dict[str, dict[str, Any]], account_id: str) -> str | None:
    if len(account_id) < 4:
        return None
    if len(account_id) >= 6 and account_description(accounts_by_id, account_id[:6]):
        return account_description(accounts_by_id, account_id[:6])
    return account_description(accounts_by_id, account_id[:4]) or account_description(accounts_by_id, account_id)


def account_group_description(account_id: str) -> str:
    if account_id.startswith("1"):
        return "Ativo"
    if account_id.startswith("2"):
        return "Passivo"
    if account_id.startswith("3"):
        return "Receitas"
    if account_id.startswith("4"):
        return "Custo/Despesas"
    return "Outros"


def account_display_value(account_id: str, row: dict[str, Any]) -> float:
    saldo = number(row.get("saldo"))
    if account_id.startswith(("2", "3")):
        return -saldo
    return saldo


def set_number_format(cell, kind: str = "money") -> None:
    if kind == "percent":
        cell.number_format = PERCENT_FMT
    elif kind == "number":
        cell.number_format = NUMBER_FMT
    elif kind == "blank":
        cell.number_format = "General"
    else:
        cell.number_format = MONEY_FMT


def style_header(ws, row: int = 1, fill: PatternFill = HEADER_FILL) -> None:
    ws.row_dimensions[row].height = 28
    for cell in ws[row]:
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER


def apply_modern_layout(
    ws,
    header_rows: tuple[int, ...] = (1,),
    data_start_row: int = 2,
    max_style_rows: int | None = None,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    max_row = min(ws.max_row, max_style_rows) if max_style_rows else ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        row_idx = row[0].row
        is_header = row_idx in header_rows
        for cell in row:
            if not is_header:
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if row_idx >= data_start_row and row_idx % 2 == 0 and cell.fill.fill_type is None:
                    cell.fill = ZEBRA_FILL


def mark_subtotal_row(ws, row: int) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        cell.font = BOLD_FONT
        cell.fill = SUBTOTAL_FILL
        cell.border = THIN_BORDER


def average(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def parse_br_date(raw: str) -> date:
    """Converte --data-fim (formato brasileiro) para um objeto date.

    Aceita DD/MM/AAAA (padrão), DD-MM-AAAA e DDMMAAAA sem separador — nunca
    AAAA-MM-DD: o parâmetro é digitado pelo usuário no padrão do Brasil. A
    conversão para ISO (AAAA-MM-DD), exigida pela API/Postgres, é feita uma
    única vez aqui; o resto do script só circula a data já em ISO.
    """
    text = str(raw or "").strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d%m%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"--data-fim inválida: {raw!r}. Use o formato brasileiro DD/MM/AAAA (ex.: 31/05/2026).")


def resolve_year_ends(years: list[int], data_fim_iso: str | None) -> dict[int, str]:
    """Data final (dataFim) de cada exercício na consulta à API, em ISO (AAAA-MM-DD).

    Por padrão é 31/12 de cada ano. Quando data_fim_iso é informado, ele só
    vale para o ano mais recente do intervalo (relatório parcial do exercício
    corrente, ex.: até 2026-05-31) — os anos anteriores continuam fechados em
    31/12. O parâmetro já vem convertido de DD/MM/AAAA para ISO por
    parse_br_date(); a validação de que a data pertence ao ano mais recente é
    feita em generate_report(). Esta função só monta o mapa ano -> data final.
    """
    last_year = max(years)
    ends = {year: f"{year}-12-31" for year in years}
    if data_fim_iso:
        ends[last_year] = data_fim_iso
    return ends


def av_for_dre_key(key: str, values: dict[str, float]) -> float | None:
    if key not in AV_KEYS:
        return None
    line = DRE_LINE_BY_KEY.get(key)
    base_key = line.percent_base if (line and line.percent_base) else "receita_liquida"
    return safe_ratio(values.get(key, 0.0), values.get(base_key, 0.0))


def dre_label(key: str) -> str:
    line = DRE_LINE_BY_KEY.get(key)
    return line.label if line else key


def write_audit(
    audit_rows: list[list[Any]],
    sheet: str,
    cell: str,
    bloco: str,
    tipo: str,
    year: Any,
    branch: str,
    label: str,
    value: Any,
    criterio: str,
    fonte: str,
    observacao: str = "",
    *,
    referencias: str = "",
    formato: str = "money",
) -> None:
    audit_rows.append([
        sheet,
        cell,
        bloco,
        tipo,
        year,
        branch,
        label,
        value,
        criterio,
        fonte,
        referencias,
        observacao,
        formato,
    ])


def dre_source(year: int, year_ends: dict[int, str], branch: BranchInfo | None = None) -> str:
    filial = f"&filialId={branch.id}" if branch else ""
    return (
        "/api/v1/executivo/contabilidade/sintetico?"
        f"dataInicio={year}-01-01&dataFim={year_ends[year]}{filial}"
        f"&excluirEncerramento=true&historicoEncerramento={ENCERRAMENTO_HISTORICO}"
    )


def sum_dre_from_branches(
    years: list[int],
    branch_dre_by_year: dict[int, dict[str, dict[str, float]]],
    branches: list[BranchInfo],
) -> dict[int, dict[str, float]]:
    keys = [line.key for line in DRE_LINES]
    result: dict[int, dict[str, float]] = {}
    for year in years:
        result[year] = {
            key: sum(
                branch_dre_by_year.get(year, {}).get(branch.id, {}).get(key, 0.0)
                for branch in branches
            )
            for key in keys
        }
    return result


# --------------------------------------------------------------------------- #
# Coleta de dados (somente ActionAPI)
# --------------------------------------------------------------------------- #
def fetch_all_data(args: argparse.Namespace, years: list[int], branches: list[BranchInfo], year_ends: dict[int, str]):
    api_key = first_api_key()
    dre_accounts_by_year: dict[int, dict[str, dict[str, Any]]] = {}
    dre_by_year: dict[int, dict[str, float]] = {}
    bp_accounts_by_year: dict[int, dict[str, dict[str, Any]]] = {}
    indicators_by_year: dict[int, dict[str, float]] = {}
    branch_dre_by_year: dict[int, dict[str, dict[str, float]]] = {}
    branch_accounts_by_year: dict[int, dict[str, dict[str, dict[str, Any]]]] = {}
    branch_accum_by_year: dict[int, dict[str, dict[str, dict[str, Any]]]] = {}
    branch_prior_before_first: dict[str, dict[str, dict[str, Any]]] = {}

    for year in years:
        dre_accounts = fetch_synthetic(
            args.api_url,
            api_key,
            data_inicio=f"{year}-01-01",
            data_fim=year_ends[year],
            excluir_encerramento=True,
            historico_encerramento=args.historico_encerramento,
        )
        if year == max(years):
            assert_api_filter_supported(dre_accounts, args.permitir_api_sem_filtro)
        bp_accounts = fetch_synthetic(
            args.api_url,
            api_key,
            data_inicio=DATA_INICIO_PLANO_ATUAL,
            data_fim=year_ends[year],
        )
        dre_accounts_by_year[year] = dre_accounts
        dre_by_year[year] = calculate_dre(dre_accounts)
        bp_accounts_by_year[year] = bp_accounts
        indicators_by_year[year] = calculate_indicators(year, dre_by_year[year], bp_accounts)
        print(f"[dre-controller] {year} (até {year_ends[year]}): DRE e BP carregados", flush=True)

    def fetch_branch_year(year: int, branch: BranchInfo):
        accounts = fetch_synthetic(
            args.api_url,
            api_key,
            data_inicio=f"{year}-01-01",
            data_fim=year_ends[year],
            filial_id=branch.id,
            excluir_encerramento=True,
            historico_encerramento=args.historico_encerramento,
        )
        return year, branch, accounts, calculate_dre(accounts)

    branch_tasks = [(year, branch) for year in years for branch in branches]
    max_workers = min(8, max(1, len(branch_tasks)))
    for year in years:
        branch_dre_by_year[year] = {}
        branch_accounts_by_year[year] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(fetch_branch_year, year, branch) for year, branch in branch_tasks]
        for future in as_completed(futures):
            year, branch, accounts, values = future.result()
            branch_accounts_by_year[year][branch.id] = accounts
            branch_dre_by_year[year][branch.id] = values
            print(f"[dre-controller] {year} filial {branch.id} {branch.nome}: DRE carregada", flush=True)

    def fetch_branch_accum(year: int, branch: BranchInfo):
        # year_ends só tem entradas para os anos pedidos em --anos; o ano
        # anterior ao primeiro (usado só para o saldo de abertura do
        # balancete) sempre fecha em 31/12, mesmo com --data-fim.
        accounts = fetch_synthetic(
            args.api_url,
            api_key,
            data_inicio=DATA_INICIO_PLANO_ATUAL,
            data_fim=year_ends.get(year, f"{year}-12-31"),
            filial_id=branch.id,
        )
        return year, branch, accounts

    for year in years:
        branch_accum_by_year[year] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(fetch_branch_accum, year, branch) for year, branch in branch_tasks]
        for future in as_completed(futures):
            year, branch, accounts = future.result()
            branch_accum_by_year[year][branch.id] = accounts
    print("[dre-controller] saldos acumulados por filial carregados", flush=True)

    first_year = min(years)
    prior_year = first_year - 1
    with ThreadPoolExecutor(max_workers=min(8, len(branches))) as executor:
        futures = [executor.submit(fetch_branch_accum, prior_year, branch) for branch in branches]
        for future in as_completed(futures):
            _year, branch, accounts = future.result()
            branch_prior_before_first[branch.id] = accounts

    return (
        dre_accounts_by_year,
        dre_by_year,
        bp_accounts_by_year,
        indicators_by_year,
        branch_dre_by_year,
        branch_accounts_by_year,
        branch_accum_by_year,
        branch_prior_before_first,
    )


# --------------------------------------------------------------------------- #
# Abas
# --------------------------------------------------------------------------- #
def build_dre_exercicio(
    wb: Workbook,
    years: list[int],
    dre_by_year: dict[int, dict[str, float]],
    audit_rows: list[list[Any]],
    year_ends: dict[int, str],
) -> None:
    # Colunas A/B replicam os símbolos de filtro do modelo (ver bloco de
    # comentário junto a DRE_LINES.col_a/col_b — fonte única, não duplicar a
    # regra aqui). Colunas de dados intercaladas por ano, igual ao modelo:
    # [(%) A.H. ano/ano-1 | Ano | (%) A.V. ano] ... e o ano mais antigo sem A.H.
    ws = wb.create_sheet("DRE por Exercício")
    years_desc = sorted(years, reverse=True)
    n = len(years_desc)

    headers: list[Any] = [".", ".", "Contas Contábeis"]
    # tipo da coluna por posição: ("ah"|"val"|"av", year)
    col_spec: list[tuple[str, int]] = []
    for idx, y in enumerate(years_desc):
        if idx < n - 1:
            headers.append(f"(%) A.H. {y}/{years_desc[idx + 1]}")
            col_spec.append(("ah", y))
        headers.append(str(y))
        col_spec.append(("val", y))
        headers.append(f"(%) A.V. {y}")
        col_spec.append(("av", y))
    ws.append(headers)
    style_header(ws)
    data_start_col = 4  # A=1, B=2, Contas Contábeis=3

    for key in VISIBLE_DRE_KEYS:
        line = DRE_LINE_BY_KEY[key]
        record: list[Any] = [line.col_a or None, line.col_b or None, ("  " * line.level) + line.label]
        for idx, y in enumerate(years_desc):
            if idx < n - 1:
                cur = dre_by_year[y].get(key, 0.0)
                prev = dre_by_year[years_desc[idx + 1]].get(key, 0.0)
                record.append(safe_ratio(cur, prev) - 1 if prev else None)
            record.append(dre_by_year[y].get(key, 0.0))
            record.append(av_for_dre_key(key, dre_by_year[y]))
        ws.append(record)
        r = ws.max_row
        for col, (kind, _y) in enumerate(col_spec, start=data_start_col):
            set_number_format(ws.cell(r, col), "money" if kind == "val" else "percent")
        if line.bold:
            mark_subtotal_row(ws, r)
        for col, (kind, y) in enumerate(col_spec, start=data_start_col):
            cell_ref = f"{get_column_letter(col)}{r}"
            if kind == "val":
                write_audit(
                    audit_rows, ws.title, cell_ref, "DRE por exercicio", "Valor",
                    y, "Consolidado", line.label, dre_by_year[y].get(key, 0.0),
                    "Consolidado DRE = soma das filiais exibidas; " + dre_line_formula_text(line),
                    "Soma das chamadas por filial: " + dre_source(y, year_ends),
                    "Valor apresentado na aba DRE por Exercício.",
                    referencias=dre_line_components_text(line, dre_by_year[y]),
                    formato="money",
                )
            elif kind == "av":
                av_value = av_for_dre_key(key, dre_by_year[y])
                base_key = line.percent_base if line.percent_base else "receita_liquida"
                write_audit(
                    audit_rows, ws.title, cell_ref, "DRE por exercicio", "Analise vertical",
                    y, "Consolidado", line.label, av_value,
                    "A.V. = valor da linha / base de comparacao; fica em branco nas linhas de detalhe.",
                    "Valores calculados da propria DRE anual.",
                    "Linha sem A.V. no layout." if av_value is None else "Percentual apresentado na aba DRE por Exercício.",
                    referencias=(
                        f"Numerador={line.label}={format_formula_value(dre_by_year[y].get(key, 0.0))}; "
                        f"Base={dre_label(base_key)}={format_formula_value(dre_by_year[y].get(base_key, 0.0))}"
                    ),
                    formato="percent" if av_value is not None else "blank",
                )
            elif kind == "ah":
                years_desc_index = years_desc.index(y)
                previous_year = years_desc[years_desc_index + 1]
                cur = dre_by_year[y].get(key, 0.0)
                prev = dre_by_year[previous_year].get(key, 0.0)
                write_audit(
                    audit_rows, ws.title, cell_ref, "DRE por exercicio", "Analise horizontal",
                    y, "Consolidado", line.label, safe_ratio(cur, prev) - 1 if prev else None,
                    "A.H. = (ano atual / ano anterior) - 1; fica em branco quando nao ha base anterior.",
                    "Valores calculados da propria DRE anual.",
                    "Comparativo entre exercicios da mesma linha.",
                    referencias=(
                        f"{y}={format_formula_value(cur)}; "
                        f"{previous_year}={format_formula_value(prev)}"
                    ),
                    formato="percent",
                )

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 48
    for col, (kind, _y) in enumerate(col_spec, start=data_start_col):
        ws.column_dimensions[get_column_letter(col)].width = 16 if kind == "val" else 13
    apply_modern_layout(ws, header_rows=(1,), data_start_row=2)
    ws.sheet_properties.tabColor = "0B5D1E"


def build_dre_comparativa_por_ano(
    wb: Workbook,
    years: list[int],
    dre_by_year: dict[int, dict[str, float]],
    branch_dre_by_year: dict[int, dict[str, dict[str, float]]],
    branches: list[BranchInfo],
    audit_rows: list[list[Any]],
    year_ends: dict[int, str],
) -> None:
    """Comparativo por filial, filtrável por exercício (filtro nativo na coluna Ano).

    A.V. do Consolidado e "(%) no Consolidado" por filial usam a MESMA regra
    seletiva de av_for_dre_key/AV_KEYS já aplicada em build_dre_exercicio: só
    aparecem nas linhas de subtotal/resultado (ver bloco de comentário junto a
    AV_KEYS). Nas linhas de detalhe, ambas as colunas ficam em branco (None),
    igual ao modelo do controller.

    Colunas B/C replicam os símbolos de filtro do modelo (DRE_LINES.col_a/
    col_b — fonte única, não duplicar a regra aqui). A coluna "Ano" (A) é
    exclusiva desta aba, para permitir o filtro nativo do Excel por exercício.
    """
    ws = wb.create_sheet("DRE Comparativa por Ano")
    headers = ["Ano", ".", ".", "Contas Contábeis", "Consolidado", "(%) A.V."]
    for branch in branches:
        headers += [branch_report_name(branch), "(%) no Consolidado"]
    ws.append(headers)
    style_header(ws)
    branch_start_col = 7

    for year in sorted(years, reverse=True):
        consolidated = dre_by_year[year]
        for key in VISIBLE_DRE_KEYS:
            line = DRE_LINE_BY_KEY[key]
            value = consolidated.get(key, 0.0)
            is_av_line = key in AV_KEYS
            record: list[Any] = [
                year,
                line.col_a or None,
                line.col_b or None,
                ("  " * line.level) + line.label,
                value,
                av_for_dre_key(key, consolidated),
            ]
            for branch in branches:
                branch_values = branch_dre_by_year[year].get(branch.id, {})
                branch_value = branch_values.get(key, 0.0)
                share = safe_ratio(branch_value, value) if is_av_line else None
                record += [branch_value, share]
            ws.append(record)
            r = ws.max_row
            set_number_format(ws.cell(r, 5), "money")
            set_number_format(ws.cell(r, 6), "percent")
            col = branch_start_col
            for _branch in branches:
                set_number_format(ws.cell(r, col), "money")
                set_number_format(ws.cell(r, col + 1), "percent")
                col += 2
            if line.bold:
                mark_subtotal_row(ws, r)
            write_audit(
                audit_rows, ws.title, f"E{r}", "DRE comparativa", "Valor", year, "Consolidado",
                line.label, value, "Consolidado DRE = soma das filiais exibidas; " + dre_line_formula_text(line),
                "Soma das chamadas por filial: " + dre_source(year, year_ends),
                "Filtre a coluna Ano para escolher o exercicio.",
                referencias=dre_line_components_text(line, consolidated),
                formato="money",
            )
            base_key = line.percent_base if line.percent_base else "receita_liquida"
            av_value = av_for_dre_key(key, consolidated)
            write_audit(
                audit_rows, ws.title, f"F{r}", "DRE comparativa", "Analise vertical", year, "Consolidado",
                line.label, av_value,
                "A.V. = valor consolidado / base de comparacao; fica em branco nas linhas de detalhe.",
                "Valores calculados da propria DRE comparativa.",
                "Linha sem A.V. no layout." if av_value is None else "Percentual apresentado na aba DRE Comparativa por Ano.",
                referencias=(
                    f"Numerador={line.label}={format_formula_value(value)}; "
                    f"Base={dre_label(base_key)}={format_formula_value(consolidated.get(base_key, 0.0))}"
                ),
                formato="percent" if av_value is not None else "blank",
            )
            col = branch_start_col
            for branch in branches:
                branch_values = branch_dre_by_year[year].get(branch.id, {})
                branch_value = branch_values.get(key, 0.0)
                share = safe_ratio(branch_value, value) if is_av_line else None
                branch_name = branch_report_name(branch)
                write_audit(
                    audit_rows, ws.title, f"{get_column_letter(col)}{r}", "DRE comparativa", "Valor",
                    year, branch_name, line.label, branch_value, dre_line_formula_text(line),
                    dre_source(year, year_ends, branch),
                    "Valor da filial usado na soma do consolidado.",
                    referencias=dre_line_components_text(line, branch_values),
                    formato="money",
                )
                write_audit(
                    audit_rows, ws.title, f"{get_column_letter(col + 1)}{r}", "DRE comparativa",
                    "% no consolidado", year, branch_name, line.label, share,
                    "% filial = valor da filial / consolidado da mesma linha; fica em branco nas linhas de detalhe.",
                    "Valores calculados da DRE por filial e do consolidado por soma.",
                    "Linha sem percentual no layout." if share is None else "Participacao da filial na linha consolidada.",
                    referencias=(
                        f"Filial={format_formula_value(branch_value)}; "
                        f"Consolidado={format_formula_value(value)}"
                    ),
                    formato="percent" if share is not None else "blank",
                )
                col += 2

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    col = branch_start_col
    for _branch in branches:
        ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions[get_column_letter(col + 1)].width = 14
        col += 2
    apply_modern_layout(ws, header_rows=(1,), data_start_row=2)
    ws.sheet_properties.tabColor = "1D6F42"


INDICATOR_ROWS = [
    ("Liquidez Corrente", "liquidez_corrente", "Ativo Circulante / Passivo Circulante", "number"),
    ("Liquidez Imediata", "liquidez_imediata", "Disponível / Passivo Circulante", "number"),
    ("Liquidez Seca", "liquidez_seca", "(Ativo Circulante - Estoques) / Passivo Circulante", "number"),
    ("Liquidez Geral", "liquidez_geral", "(Ativo Circulante + Realizável a Longo Prazo) / (Passivo Circulante + Passivo Não Circulante)", "number"),
    ("Endividamento (Cap. Terceiros / PL)", "endividamento", "(Passivo Circulante + Passivo Não Circulante) / Patrimônio Líquido", "number"),
    ("Endividamento Geral", "endividamento_geral", "(Passivo Circulante + Passivo Não Circulante) / Ativo Total", "percent"),
    ("Empréstimos / EBITDA", "emprestimos_ebitda", "Empréstimos e Financiamentos / EBITDA", "number"),
    ("ROA", "roa", "Resultado do Exercício / Ativo Total", "percent"),
    ("ROE", "roe", "Resultado do Exercício / Patrimônio Líquido", "percent"),
    ("EBITDA", "ebitda", "Lucro Operacional + Depreciação/Amortização", "money"),
    ("Empréstimos e Financiamentos (Circulante)", "emprestimos", "Conta 211104 do BP (saldo circulante)", "money"),
]


def build_bp(
    wb: Workbook,
    years: list[int],
    bp_accounts_by_year: dict[int, dict[str, dict[str, Any]]],
    indicators_by_year: dict[int, dict[str, float]],
    inactive_text: str,
    audit_rows: list[list[Any]],
    year_ends: dict[int, str],
) -> None:
    ws = wb.create_sheet("Balanço Patrimonial")
    ws.append([inactive_text])
    ws.cell(1, 1).font = MUTED_FONT
    ws.append([])
    headers = ["Conta", "Nomenclatura"] + [str(y) for y in years]
    ws.append(headers)
    header_row = ws.max_row
    style_header(ws, header_row)

    for line in BP_LINES:
        record: list[Any] = [line.account, ("  " * line.level) + line.label]
        for year in years:
            value = bp_value(bp_accounts_by_year[year], line.account)
            record.append(value)
        ws.append(record)
        r = ws.max_row
        for idx, year in enumerate(years):
            set_number_format(ws.cell(r, 3 + idx), "money")
            write_audit(
                audit_rows, ws.title, f"{get_column_letter(3 + idx)}{r}", "Balanço Patrimonial", "Valor",
                year, "Consolidado", f"{line.account} - {line.label}", bp_value(bp_accounts_by_year[year], line.account),
                "Ativo = saldo D-C; Passivo/PL = sinal invertido para apresentação positiva.",
                f"/api/v1/executivo/contabilidade/sintetico?dataInicio={DATA_INICIO_PLANO_ATUAL}&dataFim={year_ends[year]}",
                "",
            )
        if line.bold:
            mark_subtotal_row(ws, r)

    # Bloco Ativo / Passivo / Ativo - Passivo por ano (igual ao modelo)
    ws.append([])
    for account, label in (("1", "Ativo"), ("2", "Passivo")):
        ws.append([account, label] + [bp_value(bp_accounts_by_year[y], account) for y in years])
        r = ws.max_row
        for idx, year in enumerate(years):
            set_number_format(ws.cell(r, 3 + idx), "money")
            write_audit(
                audit_rows, ws.title, f"{get_column_letter(3 + idx)}{r}", "Balanço Patrimonial",
                "Resumo BP", year, "Consolidado", label, bp_value(bp_accounts_by_year[year], account),
                "Resumo por grupo do BP; Passivo/PL com sinal invertido para apresentacao positiva.",
                f"/api/v1/executivo/contabilidade/sintetico?dataInicio={DATA_INICIO_PLANO_ATUAL}&dataFim={year_ends[year]}",
                "Bloco de conferencia Ativo/Passivo.",
                referencias=f"Conta={account}",
                formato="money",
            )
        mark_subtotal_row(ws, r)
    ws.append(["", "Ativo - Passivo"] + [
        bp_value(bp_accounts_by_year[y], "1") - bp_value(bp_accounts_by_year[y], "2") for y in years
    ])
    r = ws.max_row
    for idx, year in enumerate(years):
        set_number_format(ws.cell(r, 3 + idx), "money")
        ativo = bp_value(bp_accounts_by_year[year], "1")
        passivo = bp_value(bp_accounts_by_year[year], "2")
        write_audit(
            audit_rows, ws.title, f"{get_column_letter(3 + idx)}{r}", "Balanço Patrimonial",
            "Conferencia BP", year, "Consolidado", "Ativo - Passivo", ativo - passivo,
            "Ativo - Passivo; usado para conferir fechamento do BP.",
            "Valores calculados das linhas do proprio BP.",
            "Quanto mais proximo de zero, melhor a conciliacao do balanco.",
            referencias=f"Ativo={format_formula_value(ativo)}; Passivo={format_formula_value(passivo)}",
            formato="money",
        )
    mark_subtotal_row(ws, r)

    # Indicadores
    ws.append([])
    ws.append(["Indicadores"])
    ws.cell(ws.max_row, 1).font = BOLD_FONT
    for label, key, criterio, kind in INDICATOR_ROWS:
        record = [label, criterio] + [indicators_by_year[y].get(key, 0.0) for y in years]
        ws.append(record)
        r = ws.max_row
        for idx, year in enumerate(years):
            set_number_format(ws.cell(r, 3 + idx), kind)
            write_audit(
                audit_rows, ws.title, f"{get_column_letter(3 + idx)}{r}", "Balanço Patrimonial", "Indicador",
                year, "Consolidado", label, indicators_by_year[year].get(key, 0.0), criterio,
                "DRE consolidada por soma das filiais + BP calculado pela ActionAPI", "",
                referencias="EBITDA e resultados referenciam a DRE consolidada do relatorio." if "EBITDA" in label or key in {"roa", "roe"} else "",
                formato=kind,
            )

    ws.freeze_panes = f"C{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(2 + len(years))}{ws.max_row}"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    for idx in range(len(years)):
        ws.column_dimensions[get_column_letter(3 + idx)].width = 18
    apply_modern_layout(ws, header_rows=(header_row,), data_start_row=header_row + 1)
    ws.sheet_properties.tabColor = "7030A0"


def build_planejamento(
    wb: Workbook,
    years: list[int],
    dre_by_year: dict[int, dict[str, float]],
    branch_dre_by_year: dict[int, dict[str, dict[str, float]]],
    branches: list[BranchInfo],
    audit_rows: list[list[Any]],
    cutoff_iso: str | None,
    year_ends: dict[int, str],
) -> None:
    """Aba Planejamento.

    Estrutura de colunas FIXA (sempre as mesmas, com ou sem --data-fim) —
    consolidado tem 4 colunas, cada filial tem 5:

        Média | Último exercício fechado | Parcial | Planejado [| % Consol.]

    - "Média" e "Último exercício fechado" NUNCA incluem o ano parcial
      cortado por --data-fim — só os anos realmente fechados.
    - "Parcial" só tem valor quando existe um ano ainda não fechado
      (--data-fim informado); caso contrário a coluna existe mas fica
      vazia em todas as linhas. O rótulo dela é o mês/ano do corte em
      português (ex.: "maio/2026"); sem --data-fim, o rótulo é "(parcial)"
      e a coluna fica sem dados.
    - "% Consol." (só nas filiais) usa a mesma regra seletiva de
      av_for_dre_key/AV_KEYS das abas de DRE (só aparece em linhas de
      subtotal/resultado) e é sempre calculada com base no último
      exercício FECHADO — nunca no ano parcial.

    Colunas A/B replicam os símbolos de filtro do modelo (DRE_LINES.col_a/
    col_b — fonte única, não duplicar a regra aqui).
    """
    ws = wb.create_sheet("Planejamento")
    last_year = max(years)
    plan_year = last_year + 1
    is_partial = bool(cutoff_iso)

    closed_years = [y for y in years if not (is_partial and y == last_year)]
    if not closed_years:
        closed_years = [last_year]
    last_closed_year = max(closed_years)
    period_label = (
        f"{min(closed_years)}-{max(closed_years)}"
        if len(closed_years) > 1
        else str(closed_years[0])
    )

    if is_partial:
        cutoff_date = date.fromisoformat(cutoff_iso)
        partial_label = f"{MESES_PT[cutoff_date.month]}/{cutoff_date.year}"
    else:
        partial_label = "(parcial)"

    # Linha 1: agrupamento mesclado (CONSOLIDADO + nome de cada filial sobre o
    # respectivo bloco de colunas), igual ao modelo do controller. Linha 2:
    # cabeçalho real de cada coluna, usado pelo filtro automático do Excel.
    ws.cell(1, 4, "CONSOLIDADO")
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)
    branch_col = 8
    for branch in branches:
        ws.cell(1, branch_col, branch_report_name(branch))
        ws.merge_cells(start_row=1, start_column=branch_col, end_row=1, end_column=branch_col + 4)
        branch_col += 5
    style_header(ws, row=1, fill=GROUP_HEADER_FILL)
    for col in range(1, branch_col):
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        ".", ".", "Contas Contábeis",
        f"Média {period_label}",
        str(last_closed_year),
        partial_label,
        f"{plan_year} (Planejado)",
    ]
    for branch in branches:
        branch_name = branch_report_name(branch)
        headers += [
            f"{branch_name} Média",
            f"{branch_name} {last_closed_year}",
            f"{branch_name} {partial_label}",
            f"{branch_name} {plan_year} (Planejado)",
            f"{branch_name} % Consol.",
        ]
    ws.append(headers)
    style_header(ws, row=2)

    for key in VISIBLE_DRE_KEYS:
        line = DRE_LINE_BY_KEY[key]
        is_av_line = key in AV_KEYS
        consolidated_media = average([dre_by_year[y].get(key, 0.0) for y in closed_years])
        consolidated_closed = dre_by_year[last_closed_year].get(key, 0.0)
        consolidated_partial = dre_by_year[last_year].get(key, 0.0) if is_partial else None
        record: list[Any] = [
            line.col_a or None,
            line.col_b or None,
            ("  " * line.level) + line.label,
            consolidated_media,
            consolidated_closed,
            consolidated_partial,
            None,  # planejado em branco
        ]

        for branch in branches:
            branch_media = average([
                branch_dre_by_year.get(y, {}).get(branch.id, {}).get(key, 0.0) for y in closed_years
            ])
            branch_closed = branch_dre_by_year.get(last_closed_year, {}).get(branch.id, {}).get(key, 0.0)
            branch_partial = (
                branch_dre_by_year.get(last_year, {}).get(branch.id, {}).get(key, 0.0) if is_partial else None
            )
            share = safe_ratio(branch_closed, consolidated_closed) if is_av_line else None
            record += [branch_media, branch_closed, branch_partial, None, share]

        ws.append(record)
        r = ws.max_row
        for col in (4, 5, 6, 7):
            set_number_format(ws.cell(r, col), "money")
        col = 8
        for _branch in branches:
            set_number_format(ws.cell(r, col), "money")       # Média
            set_number_format(ws.cell(r, col + 1), "money")   # último exercício fechado
            set_number_format(ws.cell(r, col + 2), "money")   # parcial
            set_number_format(ws.cell(r, col + 3), "money")   # planejado
            set_number_format(ws.cell(r, col + 4), "percent")  # % no consolidado
            col += 5
        if line.bold:
            mark_subtotal_row(ws, r)
        media_refs = "; ".join(
            f"{year}={format_formula_value(dre_by_year[year].get(key, 0.0))}" for year in closed_years
        )
        partial_note = (
            f"Exercicio corrente parcial ate {cutoff_iso}."
            if is_partial else
            "Coluna parcial sem dados porque nenhum --data-fim foi informado."
        )
        write_audit(
            audit_rows, ws.title, f"D{r}", "Planejamento", "Media", period_label, "Consolidado",
            line.label, consolidated_media,
            f"Media = media simples dos exercicios fechados ({period_label}); consolidado = soma das filiais exibidas.",
            "DRE consolidada por soma das filiais.",
            "Valor apresentado na coluna de media do Planejamento.",
            referencias=media_refs,
            formato="money",
        )
        write_audit(
            audit_rows, ws.title, f"E{r}", "Planejamento", "Ultimo exercicio fechado", last_closed_year, "Consolidado",
            line.label, consolidated_closed,
            "Valor do ultimo exercicio fechado; consolidado = soma das filiais exibidas.",
            "DRE consolidada por soma das filiais.",
            "Valor apresentado na coluna do ultimo exercicio fechado.",
            referencias=dre_line_components_text(line, dre_by_year[last_closed_year]),
            formato="money",
        )
        write_audit(
            audit_rows, ws.title, f"F{r}", "Planejamento", "Parcial", last_year if is_partial else "",
            "Consolidado", line.label, consolidated_partial,
            "Valor do exercicio parcial quando --data-fim e informado; caso contrario fica em branco.",
            "DRE consolidada por soma das filiais.",
            partial_note,
            referencias=dre_line_components_text(line, dre_by_year[last_year]) if is_partial else "",
            formato="money" if is_partial else "blank",
        )
        write_audit(
            audit_rows, ws.title, f"G{r}", "Planejamento", "Planejado", plan_year, "Consolidado",
            line.label, None,
            "Celula reservada para planejamento manual; o script nao preenche valor.",
            "Sem fonte automatica.",
            "Celula intencionalmente em branco.",
            referencias="",
            formato="blank",
        )

        col = 8
        for branch in branches:
            branch_name = branch_report_name(branch)
            branch_media_values = [
                branch_dre_by_year.get(y, {}).get(branch.id, {}).get(key, 0.0) for y in closed_years
            ]
            branch_media = average(branch_media_values)
            branch_closed = branch_dre_by_year.get(last_closed_year, {}).get(branch.id, {}).get(key, 0.0)
            branch_partial = (
                branch_dre_by_year.get(last_year, {}).get(branch.id, {}).get(key, 0.0) if is_partial else None
            )
            branch_share = safe_ratio(branch_closed, consolidated_closed) if is_av_line else None
            branch_media_refs = "; ".join(
                f"{year}={format_formula_value(branch_dre_by_year.get(year, {}).get(branch.id, {}).get(key, 0.0))}"
                for year in closed_years
            )
            write_audit(
                audit_rows, ws.title, f"{get_column_letter(col)}{r}", "Planejamento", "Media",
                period_label, branch_name, line.label, branch_media,
                f"Media da filial = media simples dos exercicios fechados ({period_label}).",
                "DRE por filial calculada pela ActionAPI.",
                "Valor apresentado na coluna de media da filial.",
                referencias=branch_media_refs,
                formato="money",
            )
            write_audit(
                audit_rows, ws.title, f"{get_column_letter(col + 1)}{r}", "Planejamento",
                "Ultimo exercicio fechado", last_closed_year, branch_name, line.label, branch_closed,
                "Valor da filial no ultimo exercicio fechado.",
                dre_source(last_closed_year, year_ends, branch),
                "Este valor compoe a soma do consolidado.",
                referencias=dre_line_components_text(line, branch_dre_by_year.get(last_closed_year, {}).get(branch.id, {})),
                formato="money",
            )
            write_audit(
                audit_rows, ws.title, f"{get_column_letter(col + 2)}{r}", "Planejamento", "Parcial",
                last_year if is_partial else "", branch_name, line.label, branch_partial,
                "Valor parcial da filial quando --data-fim e informado; caso contrario fica em branco.",
                dre_source(last_year, year_ends, branch) if is_partial else "Sem fonte automatica.",
                partial_note,
                referencias=(
                    dre_line_components_text(line, branch_dre_by_year.get(last_year, {}).get(branch.id, {}))
                    if is_partial else ""
                ),
                formato="money" if is_partial else "blank",
            )
            write_audit(
                audit_rows, ws.title, f"{get_column_letter(col + 3)}{r}", "Planejamento", "Planejado",
                plan_year, branch_name, line.label, None,
                "Celula reservada para planejamento manual; o script nao preenche valor.",
                "Sem fonte automatica.",
                "Celula intencionalmente em branco.",
                referencias="",
                formato="blank",
            )
            write_audit(
                audit_rows, ws.title, f"{get_column_letter(col + 4)}{r}", "Planejamento",
                "% Consol.", last_closed_year, branch_name, line.label, branch_share,
                "% Consol. = filial no ultimo exercicio fechado / consolidado do ultimo exercicio fechado; fica em branco nas linhas de detalhe.",
                "Valores calculados da DRE por filial e do consolidado por soma.",
                "Linha sem percentual no layout." if branch_share is None else "Participacao da filial no consolidado fechado.",
                referencias=(
                    f"Filial={format_formula_value(branch_closed)}; "
                    f"Consolidado={format_formula_value(consolidated_closed)}"
                ),
                formato="percent" if branch_share is not None else "blank",
            )
            col += 5

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 48
    for col in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    apply_modern_layout(ws, header_rows=(1, 2), data_start_row=3)
    ws.sheet_properties.tabColor = "2F75B5"


def build_balancete_geral(
    wb: Workbook,
    years: list[int],
    branches: list[BranchInfo],
    branch_accounts_by_year: dict[int, dict[str, dict[str, dict[str, Any]]]],
    branch_accum_by_year: dict[int, dict[str, dict[str, dict[str, Any]]]],
    branch_prior_before_first: dict[str, dict[str, dict[str, Any]]],
    generated_at: str,
    audit_rows: list[list[Any]],
    year_ends: dict[int, str],
) -> None:
    ws = wb.create_sheet("Balancete Geral")
    headers = [
        "Loja", "Exercício", "Grau 1", "Grau 3", "Natureza da Conta", "Grau",
        "Conta Contábil", "Nomenclatura da Conta", "Saldo Anterior", "Débito",
        "Crédito", "Saldo do Mês", "Saldo Atual",
    ]
    ws.append(headers)
    style_header(ws)

    all_accounts: dict[str, dict[str, Any]] = {}
    for year in years:
        for branch_id in branch_accounts_by_year.get(year, {}):
            all_accounts.update(branch_accounts_by_year[year][branch_id])
        for branch_id in branch_accum_by_year.get(year, {}):
            all_accounts.update(branch_accum_by_year[year][branch_id])

    sorted_branches = sorted(branches, key=branch_sort_key)
    min_year = min(years)

    for branch in sorted_branches:
        for year in years:
            movement = branch_accounts_by_year.get(year, {}).get(branch.id, {})
            accumulated = branch_accum_by_year.get(year, {}).get(branch.id, {})
            if year - 1 in years:
                previous = branch_accum_by_year.get(year - 1, {}).get(branch.id, {})
            elif year == min_year:
                previous = branch_prior_before_first.get(branch.id, {})
            else:
                previous = {}

            account_ids = set(movement) | set(accumulated) | set(previous)
            for account_id in sorted(account_ids):
                if not account_id or not account_id[0].isdigit():
                    continue
                row_mov = movement.get(account_id, {})
                row_acc = accumulated.get(account_id, {})
                row_prev = previous.get(account_id, {})

                debit = number(row_mov.get("debitos"))
                credit = number(row_mov.get("creditos"))
                saldo_mes = account_display_value(account_id, row_mov)
                if account_id.startswith(("1", "2")):
                    saldo_anterior = account_display_value(account_id, row_prev)
                    saldo_atual = account_display_value(account_id, row_acc)
                else:
                    saldo_anterior = 0.0
                    saldo_atual = saldo_mes

                if (
                    abs(saldo_anterior) < 0.005
                    and abs(debit) < 0.005
                    and abs(credit) < 0.005
                    and abs(saldo_mes) < 0.005
                    and abs(saldo_atual) < 0.005
                ):
                    continue

                description = (
                    row_acc.get("descricao")
                    or row_mov.get("descricao")
                    or row_prev.get("descricao")
                    or account_description(all_accounts, account_id)
                )
                ws.append([
                    branch_display_name(branch),
                    year,
                    account_group_description(account_id),
                    account_grau3(all_accounts, account_id),
                    account_natureza(all_accounts, account_id),
                    len(account_id),
                    account_id,
                    description,
                    saldo_anterior,
                    debit,
                    credit,
                    saldo_mes,
                    saldo_atual,
                ])
                r = ws.max_row
                for col in range(9, 14):
                    set_number_format(ws.cell(r, col), "money")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"
    widths = {
        "A": 18, "B": 12, "C": 16, "D": 24, "E": 32, "F": 8, "G": 16,
        "H": 48, "I": 16, "J": 16, "K": 16, "L": 16, "M": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    if ws.max_row >= 2:
        balancete_audit_specs = [
            ("I", "Saldo anterior", "Saldo acumulado anterior; para contas de resultado fica zero."),
            ("J", "Debito", "Debitos do movimento do exercicio."),
            ("K", "Credito", "Creditos do movimento do exercicio."),
            ("L", "Saldo do mes", "Saldo do movimento: Ativo/Despesa = D-C; Passivo/Receita = sinal invertido para apresentacao."),
            ("M", "Saldo atual", "Saldo acumulado atual para BP; para resultado, repete o saldo do movimento."),
        ]
        for col_letter, tipo, criterio in balancete_audit_specs:
            write_audit(
                audit_rows,
                ws.title,
                f"{col_letter}2:{col_letter}{ws.max_row}",
                "Balancete Geral",
                tipo,
                "Todos",
                "Todas",
                f"Coluna {tipo}",
                None,
                criterio,
                "/api/v1/executivo/contabilidade/sintetico por filial e exercicio.",
                "Mapeamento aplicado a todas as linhas preenchidas da coluna.",
                referencias="Chaves da propria linha: Loja, Exercicio e Conta Contabil.",
                formato="blank",
            )
    apply_modern_layout(ws, header_rows=(1,), data_start_row=2, max_style_rows=5000)
    ws.sheet_properties.tabColor = "595959"


def validate_internal_consistency(
    years: list[int],
    report_dre_by_year: dict[int, dict[str, float]],
    branch_dre_by_year: dict[int, dict[str, dict[str, float]]],
    branches: list[BranchInfo],
    cutoff_iso: str | None,
) -> None:
    """Valida a lógica interna sem criar aba auxiliar no relatório."""
    tolerance = 0.01

    def ensure_match(bloco: str, period: Any, key: str, used: float, branch_sum: float) -> None:
        diff_used = used - branch_sum
        if abs(diff_used) > tolerance:
            raise RuntimeError(
                "Inconsistencia interna no DRE Controller: "
                f"{bloco} {period} / {dre_label(key)}. "
                f"Consolidado={format_formula_value(used)}; "
                f"soma filiais={format_formula_value(branch_sum)}; "
                f"diferenca={format_formula_value(diff_used)}."
            )

    for year in years:
        for key in VISIBLE_DRE_KEYS:
            used = report_dre_by_year[year].get(key, 0.0)
            branch_sum = sum(
                branch_dre_by_year.get(year, {}).get(branch.id, {}).get(key, 0.0)
                for branch in branches
            )
            ensure_match("DRE ano a ano", year, key, used, branch_sum)

    last_year = max(years)
    is_partial = bool(cutoff_iso)
    closed_years = [year for year in years if not (is_partial and year == last_year)]
    if not closed_years:
        closed_years = [last_year]
    last_closed_year = max(closed_years)
    period_label = (
        f"{min(closed_years)}-{max(closed_years)}"
        if len(closed_years) > 1
        else str(closed_years[0])
    )

    for key in VISIBLE_DRE_KEYS:
        used = average([report_dre_by_year[year].get(key, 0.0) for year in closed_years])
        branch_sum = sum(
            average([branch_dre_by_year.get(year, {}).get(branch.id, {}).get(key, 0.0) for year in closed_years])
            for branch in branches
        )
        ensure_match("Planejamento - Media", period_label, key, used, branch_sum)

        used_closed = report_dre_by_year[last_closed_year].get(key, 0.0)
        branch_sum_closed = sum(
            branch_dre_by_year.get(last_closed_year, {}).get(branch.id, {}).get(key, 0.0)
            for branch in branches
        )
        ensure_match("Planejamento - Ultimo fechado", last_closed_year, key, used_closed, branch_sum_closed)

        if is_partial:
            used_partial = report_dre_by_year[last_year].get(key, 0.0)
            branch_sum_partial = sum(
                branch_dre_by_year.get(last_year, {}).get(branch.id, {}).get(key, 0.0)
                for branch in branches
            )
            ensure_match("Planejamento - Parcial", cutoff_iso, key, used_partial, branch_sum_partial)


def build_mapa_calculo(wb: Workbook, audit_rows: list[list[Any]], generated_at: str) -> None:
    ws = wb.create_sheet("Mapa de Cálculo")
    headers = [
        "Aba", "Célula", "Bloco", "Tipo de valor", "Ano", "Filial",
        "Linha/Indicador", "Valor", "Fórmula/Critério", "Fonte",
        "Referências/Dependências", "Observação", "Formato",
    ]
    ws.append(headers)
    style_header(ws)

    for row in audit_rows:
        ws.append(row)
        value_cell = ws.cell(ws.max_row, 8)
        formato = str(ws.cell(ws.max_row, 13).value or "money").lower()
        if isinstance(value_cell.value, (int, float)):
            set_number_format(value_cell, formato)
        for col in (9, 10, 11, 12):
            ws.cell(ws.max_row, col).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 26, "B": 12, "C": 22, "D": 22, "E": 10, "F": 16,
        "G": 42, "H": 16, "I": 66, "J": 60, "K": 60, "L": 50, "M": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    apply_modern_layout(ws, header_rows=(1,), data_start_row=2, max_style_rows=5000)
    ws.sheet_properties.tabColor = "C65911"


# --------------------------------------------------------------------------- #
# Orquestração
# --------------------------------------------------------------------------- #
def save_workbook(wb: Workbook, output: Path, required: list[str]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    check = load_workbook(output, read_only=True)
    missing = [sheet for sheet in required if sheet not in check.sheetnames]
    check.close()
    if missing:
        raise RuntimeError(f"Validação falhou: abas ausentes={missing}.")


def generate_report(args: argparse.Namespace) -> Path:
    years = parse_years(args.anos)

    cutoff_iso: str | None = None
    if args.data_fim:
        cutoff = parse_br_date(args.data_fim)
        cutoff_iso = cutoff.isoformat()
        if cutoff.year not in years:
            years = sorted([*years, cutoff.year])
        if cutoff.year != max(years):
            raise ValueError(
                "--data-fim deve cair no ano mais recente do intervalo (relatório "
                f"parcial só vale para o exercício corrente). Ano em --data-fim: "
                f"{cutoff.year}; ano mais recente em --anos: {max(years)}."
            )

    if len(years) > 6:
        raise ValueError("Informe no máximo 6 exercícios por relatório.")
    year_ends = resolve_year_ends(years, cutoff_iso)

    output_suffix = f"-ate-{cutoff_iso}" if cutoff_iso else ""
    output = (
        Path(args.arquivo).resolve()
        if args.arquivo
        else ROOT / "relatorios" / f"dre-controller-{min(years)}-{max(years)}{output_suffix}.xlsx"
    )
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    if cutoff_iso:
        print(f"[dre-controller] exercício {max(years)} cortado em {args.data_fim} ({cutoff_iso}) — relatório parcial", flush=True)

    print(f"[dre-controller] consultando ActionAPI em {args.api_url}...", flush=True)
    api_key = first_api_key()
    branch_registry, branch_source = fetch_branch_registry(args.api_url, api_key)
    balancete_branches = branch_registry
    report_branches = controller_visible_branches(branch_registry)
    inactive_text = inactive_branches_text(branch_registry, branch_source)
    print(
        f"[dre-controller] filiais no comparativo: {', '.join(b.id for b in report_branches)}",
        flush=True,
    )
    print(f"[dre-controller] {inactive_text}", flush=True)

    audit_rows: list[list[Any]] = []
    (
        _dre_accounts_by_year,
        dre_by_year,
        bp_accounts_by_year,
        indicators_by_year,
        branch_dre_by_year,
        branch_accounts_by_year,
        branch_accum_by_year,
        branch_prior_before_first,
    ) = fetch_all_data(args, years, balancete_branches, year_ends)
    dre_by_year = sum_dre_from_branches(years, branch_dre_by_year, report_branches)
    indicators_by_year = {
        year: calculate_indicators(year, dre_by_year[year], bp_accounts_by_year[year])
        for year in years
    }
    validate_internal_consistency(years, dre_by_year, branch_dre_by_year, report_branches, cutoff_iso)
    print("[dre-controller] consolidado DRE recalculado pela soma das filiais exibidas", flush=True)

    # Ordem das abas igual ao modelo do controller (Planejamento, Balancete,
    # BP, Comparativa, Comparativa Exercício), com a Comparativa por Ano no lugar
    # da antiga SGA_DRE Comparativa. Mapa de Cálculo fica ao final.
    wb = Workbook()
    wb.remove(wb.active)
    build_planejamento(wb, years, dre_by_year, branch_dre_by_year, report_branches, audit_rows, cutoff_iso, year_ends)
    build_balancete_geral(
        wb, years, balancete_branches, branch_accounts_by_year, branch_accum_by_year,
        branch_prior_before_first, generated_at, audit_rows, year_ends,
    )
    build_bp(wb, years, bp_accounts_by_year, indicators_by_year, inactive_text, audit_rows, year_ends)
    build_dre_comparativa_por_ano(wb, years, dre_by_year, branch_dre_by_year, report_branches, audit_rows, year_ends)
    build_dre_exercicio(wb, years, dre_by_year, audit_rows, year_ends)
    build_mapa_calculo(wb, audit_rows, generated_at)

    required = [
        "Planejamento",
        "Balancete Geral",
        "Balanço Patrimonial",
        "DRE Comparativa por Ano",
        "DRE por Exercício",
        "Mapa de Cálculo",
    ]
    save_workbook(wb, output, required)
    print(f"[dre-controller] arquivo: {output}")
    return output


def main() -> None:
    try:
        generate_report(complete_interactive_args(parse_args()))
    except Exception as exc:
        print(f"[dre-controller] erro: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
