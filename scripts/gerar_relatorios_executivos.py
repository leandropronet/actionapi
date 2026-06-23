#!/usr/bin/env python3
r"""Gera relatórios executivos 360° consumindo exclusivamente a ActionAPI.

Por padrão gera sete arquivos:
  - faturamento executivo;
  - contas a receber em aberto;
  - contas a pagar em aberto;
  - contas recebidas;
  - contas pagas;
  - contabilidade;
  - visão consolidada 360°.

Exemplo:

    .\.venv\Scripts\python.exe scripts\gerar_relatorios_executivos.py `
      --data-inicio 2026-01-01 `
      --data-fim 2026-06-20
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from zipfile import ZipFile

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:
    print(
        "Dependência ausente. Execute:\n"
        "  py -m pip install -r scripts/requirements-relatorio-contas-pagar.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_API_URL = "http://127.0.0.1:3000"
PAGE_SIZE = 10_000

NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
MONEY = 'R$ #,##0.00;[Red]-R$ #,##0.00'
NUMBER = "#,##0"
DECIMAL = '#,##0.00;[Red]-#,##0.00'
PERCENT = "0.00%"
DATE = "dd/mm/yyyy"


def add_period_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--data-inicio",
        help="Data inicial: DDMMAAAA, DD/MM/AAAA ou AAAA-MM-DD.",
    )
    parser.add_argument(
        "--data-fim",
        help="Data final: DDMMAAAA, DD/MM/AAAA ou AAAA-MM-DD.",
    )
    parser.add_argument(
        "--safra",
        help="Período agrícola AAAA/AAAA: 01/07 do primeiro ano a 30/06 do segundo.",
    )
    parser.add_argument(
        "--bayer",
        help="Período Bayer AAAA/AAAA: 01/04 do primeiro ano a 30/03 do segundo.",
    )
    parser.add_argument(
        "--ano-contabil",
        help="Ano-calendário contábil com quatro dígitos, por exemplo 2025.",
    )


def parse_period_pair(value: str, option: str) -> tuple[int, int]:
    match = re.fullmatch(r"\s*(\d{4})[/-](\d{4})\s*", value or "")
    if not match:
        raise ValueError(f"{option} deve usar o formato AAAA/AAAA, por exemplo 2025/2026.")
    first, second = map(int, match.groups())
    if second != first + 1:
        raise ValueError(
            f"{option} deve conter anos consecutivos; use {first}/{first + 1}."
        )
    return first, second


def parse_user_date(value: str, option: str) -> date:
    text = str(value or "").strip()
    formats = ("%d%m%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d")
    for date_format in formats:
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError(
        f"{option} deve usar DDMMAAAA, DD/MM/AAAA ou AAAA-MM-DD."
    )


def format_date_br(value: Any) -> str:
    parsed = value if isinstance(value, date) else parse_user_date(str(value), "data")
    return parsed.strftime("%d/%m/%Y")


def period_subtitle(filters: dict[str, Any]) -> str:
    return (
        f"{format_date_br(filters['dataInicio'])} "
        f"a {format_date_br(filters['dataFim'])}"
    )


def has_period_argument(args: argparse.Namespace) -> bool:
    return any(
        getattr(args, name, None)
        for name in ("safra", "bayer", "ano_contabil", "data_inicio", "data_fim")
    )


def prompt_period_if_needed(
    args: argparse.Namespace,
    *,
    additional_period_supplied: bool = False,
) -> None:
    """Solicita o período quando ele não foi informado na linha de comando."""
    if has_period_argument(args) or additional_period_supplied or not sys.stdin.isatty():
        return

    print("\nEscolha o período do relatório:")
    print("  1 - Safra agrícola (01/07 a 30/06)")
    print("  2 - Período Bayer (01/04 a 30/03)")
    print("  3 - Ano contábil/calendário (01/01 a 31/12)")
    print("  4 - Intervalo livre")
    print("  5 - Ano atual até hoje")

    while True:
        choice = input("Opção [1-5]: ").strip()
        try:
            if choice == "1":
                args.safra = input("Safra (ex.: 2025/2026): ").strip()
            elif choice == "2":
                args.bayer = input("Período Bayer (ex.: 2025/2026): ").strip()
            elif choice == "3":
                args.ano_contabil = input("Ano contábil (ex.: 2025): ").strip()
            elif choice == "4":
                args.data_inicio = input(
                    "Data inicial (DDMMAAAA ou DD/MM/AAAA): "
                ).strip()
                args.data_fim = input(
                    "Data final (DDMMAAAA ou DD/MM/AAAA): "
                ).strip()
            elif choice == "5":
                today = date.today()
                args.data_inicio = f"01/01/{today.year}"
                args.data_fim = today.strftime("%d/%m/%Y")
            else:
                print("Opção inválida. Digite um número de 1 a 5.")
                continue

            resolve_period(args)
            return
        except ValueError as exc:
            print(f"Período inválido: {exc}")
            for name in ("safra", "bayer", "ano_contabil", "data_inicio", "data_fim"):
                setattr(args, name, None)


def resolve_period(args: argparse.Namespace) -> tuple[str, str, str]:
    selected = sum(
        bool(value)
        for value in (
            getattr(args, "safra", None),
            getattr(args, "bayer", None),
            getattr(args, "ano_contabil", None),
            getattr(args, "data_inicio", None) or getattr(args, "data_fim", None),
        )
    )
    if selected > 1:
        raise ValueError(
            "Use apenas um intervalo: --safra, --bayer, --ano-contabil "
            "ou --data-inicio com --data-fim."
        )

    if getattr(args, "safra", None):
        first, second = parse_period_pair(args.safra, "--safra")
        return f"{first}-07-01", f"{second}-06-30", f"safra-{first}-{second}"

    if getattr(args, "bayer", None):
        first, second = parse_period_pair(args.bayer, "--bayer")
        return f"{first}-04-01", f"{second}-03-30", f"bayer-{first}-{second}"

    if getattr(args, "ano_contabil", None):
        value = str(args.ano_contabil).strip()
        if not re.fullmatch(r"\d{4}", value):
            raise ValueError("--ano-contabil deve ter quatro dígitos, por exemplo 2025.")
        year = int(value)
        return f"{year}-01-01", f"{year}-12-31", f"ano-contabil-{year}"

    if getattr(args, "data_inicio", None) or getattr(args, "data_fim", None):
        if not args.data_inicio or not args.data_fim:
            raise ValueError("--data-inicio e --data-fim devem ser informadas juntas.")
        start = parse_user_date(args.data_inicio, "--data-inicio")
        end = parse_user_date(args.data_fim, "--data-fim")
        if end < start:
            raise ValueError("--data-fim não pode ser anterior a --data-inicio.")
        slug = f"{start.strftime('%d-%m-%Y')}-a-{end.strftime('%d-%m-%Y')}"
        return start.isoformat(), end.isoformat(), slug

    today = date.today()
    start = date(today.year, 1, 1)
    slug = f"{start.strftime('%d-%m-%Y')}-a-{today.strftime('%d-%m-%Y')}"
    return start.isoformat(), today.isoformat(), slug


def add_data_base_argument(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--data-base",
        help=(
            "Data-base do saldo em aberto (contas a receber/pagar): "
            "DDMMAAAA, DD/MM/AAAA ou AAAA-MM-DD. Padrão: hoje."
        ),
    )


def prompt_data_base_if_needed(args: argparse.Namespace) -> None:
    """Pergunta a data-base do saldo em aberto quando ela não foi informada.

    Saldo em aberto é uma posição "no momento": só faz sentido reproduzir
    uma data passada quando o usuário pede explicitamente — caso contrário
    assumimos hoje silenciosamente (inclusive em uso não interativo).
    """
    if getattr(args, "data_base", None) or not sys.stdin.isatty():
        return
    resposta = input(
        "Data-base do saldo em aberto (Enter = hoje, ou DD/MM/AAAA para reproduzir uma data passada): "
    ).strip()
    if resposta:
        args.data_base = resposta


def resolve_data_base(args: argparse.Namespace) -> str | None:
    value = getattr(args, "data_base", None)
    if not value:
        return None
    try:
        return parse_user_date(value, "--data-base").isoformat()
    except ValueError as exc:
        raise SystemExit(f"erro: {exc}") from exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera planilhas executivas de faturamento, caixa e contabilidade."
    )
    parser.add_argument("--api-url", default=os.getenv("ACTIONAPI_URL", DEFAULT_API_URL))
    add_period_arguments(parser)
    add_data_base_argument(parser)
    parser.add_argument("--filial-id")
    parser.add_argument("--saida-dir", default=str(ROOT / "relatorios" / "executivo"))
    parser.add_argument(
        "--relatorios",
        default="todos",
        help=(
            "todos ou lista: faturamento,a-receber,a-pagar,recebimentos,"
            "pagamentos,contabilidade,360"
        ),
    )
    return parser.parse_args()


def read_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value.startswith(("'", '"')):
            value = value[1:-1]
        values[key.strip()] = value
    return values


def first_api_key() -> str:
    env = {**read_env(ROOT / ".env"), **os.environ}
    keys = [item.strip() for item in env.get("API_KEYS", "").split(",") if item.strip()]
    if not keys:
        raise RuntimeError("API_KEYS não está configurada no .env.")
    return keys[0]


def api_get(
    base_url: str,
    api_key: str,
    endpoint: str,
    params: dict[str, Any] | None = None,
) -> dict:
    clean = {
        key: value
        for key, value in (params or {}).items()
        if value not in (None, "")
    }
    url = f"{base_url.rstrip('/')}{endpoint}"
    if clean:
        url += "?" + urlencode(clean)
    request = Request(url, headers={"X-API-Key": api_key, "Accept": "application/json"})
    try:
        with urlopen(request, timeout=180) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"ActionAPI respondeu HTTP {exc.code}: {body}") from exc
    except URLError as exc:
        raise RuntimeError(f"Não foi possível acessar a ActionAPI em {url}: {exc}") from exc


def fetch_all(
    base_url: str,
    api_key: str,
    endpoint: str,
    params: dict[str, Any],
) -> list[dict]:
    rows: list[dict] = []
    page = 1
    while True:
        payload = api_get(
            base_url,
            api_key,
            endpoint,
            {**params, "page": page, "pageSize": PAGE_SIZE},
        )
        current = payload.get("data", [])
        rows.extend(current)
        total = int(payload.get("total", len(rows)))
        if not current or len(rows) >= total:
            return rows
        page += 1


def number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def as_date(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def cnpj(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return str(value or "")


def normalized(value: Any) -> Any:
    if isinstance(value, str) and re.match(r"^\d{4}-\d{2}-\d{2}T", value):
        return as_date(value)
    return value


def rows_from_dicts(
    rows: list[dict],
    columns: list[tuple[str, str, str | None]],
) -> tuple[list[str], list[list[Any]], dict[str, str]]:
    headers = [header for _key, header, _kind in columns]
    kinds = {header: kind for _key, header, kind in columns if kind}
    output = []
    for row in rows:
        line = []
        for key, _header, kind in columns:
            value = row.get(key)
            if kind in {"money", "number", "decimal", "percent"}:
                value = number(value)
            elif kind == "date":
                value = as_date(value)
            elif kind == "cnpj":
                value = cnpj(value)
            else:
                value = normalized(value)
            line.append(value)
        output.append(line)
    return headers, output, kinds


def style_sheet(ws, kinds: dict[str, str] | None = None) -> None:
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if ws.max_row >= 2:
        ws.auto_filter.ref = ws.dimensions
    kinds = kinds or {}
    header_map = {cell.value: cell.column_letter for cell in ws[1]}
    if ws.max_row <= 20_000:
        for header, kind in kinds.items():
            letter = header_map.get(header)
            if not letter:
                continue
            fmt = {
                "money": MONEY,
                "number": NUMBER,
                "decimal": DECIMAL,
                "percent": PERCENT,
                "date": DATE,
            }.get(kind)
            if fmt:
                for cell in ws[letter][1:]:
                    cell.number_format = fmt
    sample_last = min(ws.max_row, 250)
    for cells in ws.iter_cols(min_row=1, max_row=sample_last):
        width = min(48, max(10, max(len(str(cell.value or "")) for cell in cells) + 2))
        ws.column_dimensions[cells[0].column_letter].width = width


def add_synthetic_accounts_sheet(wb: Workbook, name: str, contas: list[dict]) -> Any:
    """Plano de contas sintético + analítico, indentado por nível, com totais.

    Cada conta sintética (1, 11, 111...) soma todas as analíticas descendentes
    via prefixo do código (a hierarquia do plano é por prefixo decimal — ver
    contabilidadeSintetico em services/executivo.js). Permite conferir o
    balancete por nível, igual ao relatório nativo do SiAGRI.
    """
    print(f"[executivo] aba {name}: {len(contas)} linhas", flush=True)
    ws = wb.create_sheet(name[:31])
    headers = ["Código da Conta", "Descrição da Conta", "Débitos", "Créditos", "Saldo"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)

    levels = sorted({int(row.get("tamanho_codigo") or 0) for row in contas})
    rank = {tamanho: idx for idx, tamanho in enumerate(levels)}
    band_colors = [NAVY, BLUE, GREEN, ORANGE]

    for row in contas:
        tamanho = int(row.get("tamanho_codigo") or 0)
        nivel = rank.get(tamanho, 0)
        analitica = bool(row.get("analitica"))
        codigo = str(row.get("conta_id") or "")
        descricao = ("  " * nivel) + str(row.get("descricao") or "")
        ws.append([
            codigo, descricao,
            number(row.get("debitos")), number(row.get("creditos")), number(row.get("saldo")),
        ])
        excel_row = ws.max_row
        for col in (3, 4, 5):
            ws.cell(excel_row, col).number_format = MONEY
        if not analitica:
            color = band_colors[min(nivel, len(band_colors) - 1)]
            for col in range(1, 6):
                cell = ws.cell(excel_row, col)
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(color=WHITE, bold=True)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 52
    for letter in ("C", "D", "E"):
        ws.column_dimensions[letter].width = 20
    return ws


def add_sheet(
    wb: Workbook,
    name: str,
    headers: list[str],
    rows: list[list[Any]],
    kinds: dict[str, str] | None = None,
) -> Any:
    print(f"[executivo] aba {name}: {len(rows)} linhas", flush=True)
    ws = wb.create_sheet(name[:31])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws, kinds)
    return ws


def title_block(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A3"] = subtitle
    ws["A3"].font = Font(italic=True, color="666666")


def add_kpis(ws, kpis: list[tuple[str, float, str]], start_row: int = 5) -> None:
    ws.cell(start_row, 1, "INDICADOR")
    ws.cell(start_row, 2, "VALOR")
    for cell in ws[start_row]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    for offset, (label, value, kind) in enumerate(kpis, 1):
        row = start_row + offset
        ws.cell(row, 1, label)
        ws.cell(row, 2, number(value))
        ws.cell(row, 2).number_format = {
            "money": MONEY,
            "percent": PERCENT,
            "number": NUMBER,
            "decimal": DECIMAL,
        }.get(kind, DECIMAL)


def add_line_chart(ws, title: str, start_row: int, rows: int, value_col: int, anchor: str) -> None:
    if rows <= 0:
        return
    chart = LineChart()
    chart.title = title
    chart.height = 8
    chart.width = 15
    chart.add_data(
        Reference(ws, min_col=value_col, min_row=start_row, max_row=start_row + rows),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + rows)
    )
    ws.add_chart(chart, anchor)


def add_bar_chart(ws, title: str, start_row: int, rows: int, anchor: str) -> None:
    if rows <= 0:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.height = 8
    chart.width = 15
    chart.add_data(
        Reference(ws, min_col=2, min_row=start_row, max_row=start_row + rows),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + rows)
    )
    ws.add_chart(chart, anchor)


def add_methodology(wb: Workbook, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("Metodologia")
    ws.append(["TEMA", "METODOLOGIA"])
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 115
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def validate_xlsx(path: Path, required_sheets: list[str]) -> None:
    with ZipFile(path) as archive:
        if archive.testzip() is not None:
            raise RuntimeError(f"Arquivo XLSX inválido: {path}")
        if any(name.startswith("xl/tables/") for name in archive.namelist()):
            raise RuntimeError("O relatório não deve conter tabelas OOXML suscetíveis a reparo.")
    wb = load_workbook(path, read_only=True, data_only=False)
    missing = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
    error_cells = sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.data_type == "e"
    )
    wb.close()
    if missing or error_cells:
        raise RuntimeError(f"Validação falhou: abas ausentes={missing}, erros={error_cells}")


def save(wb: Workbook, path: Path, required_sheets: list[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    validate_xlsx(path, required_sheets)
    print(f"[executivo] arquivo validado: {path}", flush=True)
    return path


def ratio(numerator: Any, denominator: Any) -> float:
    den = number(denominator)
    return number(numerator) / den if den else 0.0


def monthly_growth(rows: list[dict], field: str) -> float:
    values = [number(row.get(field)) for row in rows]
    if len(values) < 2 or values[-2] == 0:
        return 0.0
    return values[-1] / values[-2] - 1


def generate_faturamento(
    base_url: str,
    api_key: str,
    filters: dict[str, Any],
    output: Path,
) -> tuple[Path, dict]:
    summary = api_get(base_url, api_key, "/api/v1/executivo/faturamento/resumo", filters)
    items = fetch_all(base_url, api_key, "/api/v1/executivo/faturamento", filters)
    totals = summary.get("totalizadores", {})
    monthly = summary.get("evolucao_mensal", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Painel"
    title_block(ws, "FATURAMENTO EXECUTIVO", period_subtitle(filters))
    return_rate = ratio(totals.get("devolucoes"), totals.get("vendas_brutas"))
    kpis = [
        ("Faturamento líquido", totals.get("faturamento_liquido"), "money"),
        ("Vendas brutas", totals.get("vendas_brutas"), "money"),
        ("Devoluções", totals.get("devolucoes"), "money"),
        ("Taxa de devolução", return_rate, "percent"),
        ("Notas de venda", totals.get("notas_venda"), "number"),
        ("Clientes ativos", totals.get("clientes_ativos"), "number"),
        ("Vendedores ativos", totals.get("vendedores_ativos"), "number"),
        ("Ticket médio líquido", totals.get("ticket_medio"), "money"),
        ("Crescimento último mês", monthly_growth(monthly, "faturamento_liquido"), "percent"),
    ]
    add_kpis(ws, kpis)
    ws["A17"] = "EVOLUÇÃO MENSAL"
    ws["A18"], ws["B18"], ws["C18"], ws["D18"] = (
        "Período", "Faturamento líquido", "Vendas brutas", "Devoluções"
    )
    for cell in ws[18]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    for offset, row in enumerate(monthly, 1):
        line = 18 + offset
        ws.cell(line, 1, as_date(row.get("periodo"))).number_format = DATE
        for col, key in enumerate(("faturamento_liquido", "vendas_brutas", "devolucoes"), 2):
            ws.cell(line, col, number(row.get(key))).number_format = MONEY
    add_line_chart(ws, "Faturamento líquido mensal", 18, len(monthly), 2, "F5")
    top_clients = summary.get("por_cliente", [])[:10]
    ws["A28"], ws["B28"] = "TOP 10 CLIENTES", "Vendas brutas"
    for cell in ws[28]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    for offset, row in enumerate(top_clients, 1):
        ws.cell(28 + offset, 1, row.get("cliente_nome") or row.get("cliente_id"))
        ws.cell(28 + offset, 2, number(row.get("vendas_brutas"))).number_format = MONEY
    add_bar_chart(ws, "Concentração por cliente", 28, len(top_clients), "F22")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22

    for name, key, columns in [
        ("Evolucao Mensal", "evolucao_mensal", [
            ("periodo", "Período", "date"),
            ("faturamento_liquido", "Faturamento líquido", "money"),
            ("vendas_brutas", "Vendas brutas", "money"),
            ("deducoes_saida", "Deduções de saída", "money"),
            ("devolucoes", "Devoluções", "money"),
            ("documentos", "Documentos", "number"),
        ]),
        ("Por Filial", "por_filial", [
            ("filial_id", "Filial", None),
            ("filial_identificacao", "Identificação", None),
            ("faturamento_liquido", "Faturamento líquido", "money"),
            ("vendas_brutas", "Vendas brutas", "money"),
            ("devolucoes", "Devoluções", "money"),
        ]),
        ("Por Cliente", "por_cliente", [
            ("cliente_id", "Código cliente", None),
            ("cliente_nome", "Cliente", None),
            ("quantidade_nf", "Notas", "number"),
            ("vendas_brutas", "Vendas brutas", "money"),
        ]),
        ("Por Vendedor", "por_vendedor", [
            ("vendedor_id", "Código vendedor", None),
            ("vendedor_nome", "Vendedor", None),
            ("quantidade_nf", "Notas", "number"),
            ("vendas_brutas", "Vendas brutas", "money"),
        ]),
        ("Mix Produtos", "por_grupo", [
            ("grupo_id", "Código grupo", None),
            ("grupo_descricao", "Grupo", None),
            ("quantidade_nf", "Notas", "number"),
            ("quantidade", "Quantidade", "decimal"),
            ("valor_itens", "Valor dos itens", "money"),
        ]),
    ]:
        headers, data, kinds = rows_from_dicts(summary.get(key, []), columns)
        add_sheet(wb, name, headers, data, kinds)

    item_columns = [
        ("nf_id", "Controle NF", None), ("filial_id", "Filial", None),
        ("filial_identificacao", "Identificação filial", None),
        ("data_emissao", "Data emissão", "date"), ("data_saida", "Data saída", "date"),
        ("numero_nf", "Número NF", None), ("serie", "Série", None),
        ("cliente_id", "Código cliente", None), ("cliente_nome", "Cliente", None),
        ("cliente_cnpj_cpf", "CPF/CNPJ", "cnpj"),
        ("vendedor_id", "Código vendedor", None), ("vendedor_nome", "Vendedor", None),
        ("valor_nf", "Valor NF", "money"), ("produto_id", "Código produto", None),
        ("produto_descricao", "Produto", None), ("grupo_descricao", "Grupo", None),
        ("unidade", "Unidade", None), ("quantidade", "Quantidade", "decimal"),
        ("valor_unitario", "Valor unitário", "money"), ("valor_item", "Valor item", "money"),
    ]
    headers, data, kinds = rows_from_dicts(items, item_columns)
    add_sheet(wb, "Itens Faturados", headers, data, kinds)
    notes = {}
    for row in items:
        notes.setdefault(row.get("nf_id"), row)
    note_columns = [
        ("nf_id", "Controle NF", None), ("filial_id", "Filial", None),
        ("filial_identificacao", "Identificação filial", None),
        ("data_emissao", "Data emissão", "date"), ("numero_nf", "Número NF", None),
        ("cliente_id", "Código cliente", None), ("cliente_nome", "Cliente", None),
        ("vendedor_nome", "Vendedor", None), ("valor_nf", "Valor NF", "money"),
    ]
    headers, data, kinds = rows_from_dicts(list(notes.values()), note_columns)
    add_sheet(wb, "Notas Fiscais", headers, data, kinds)
    add_methodology(wb, [
        ("Fonte", "ActionAPI: /executivo/faturamento e /executivo/faturamento/resumo."),
        ("Faturamento líquido", "Parâmetro 102 do SiAGRI: operações A menos S, combinando NOTA e devoluções em NFENTRA."),
        ("Concentração", "Clientes, vendedores e grupos são ordenados pelo valor bruto no período."),
        ("Itens", "Quantidade × valor unitário. Pode diferir do total da NF por impostos, frete e despesas."),
        ("Uso futuro", "Os mesmos endpoints foram preparados para gráficos interativos no frontend React."),
    ])
    return save(wb, output, ["Painel", "Itens Faturados", "Metodologia"]), summary


def generate_movements(
    kind: str,
    base_url: str,
    api_key: str,
    filters: dict[str, Any],
    output: Path,
) -> tuple[Path, dict]:
    endpoint = f"/api/v1/executivo/{kind}"
    summary = api_get(base_url, api_key, f"{endpoint}/resumo", filters)
    movements = fetch_all(base_url, api_key, endpoint, filters)
    totals = summary.get("totalizadores", {})
    monthly = summary.get("evolucao_mensal", [])
    receiving = kind == "recebimentos"
    title = "CONTAS RECEBIDAS" if receiving else "CONTAS PAGAS"
    partner = "Cliente" if receiving else "Fornecedor"

    wb = Workbook()
    ws = wb.active
    ws.title = "Painel"
    title_block(ws, title, period_subtitle(filters))
    kpis = [
        ("Valor líquido", totals.get("valor_liquido"), "money"),
        ("Valor principal", totals.get("valor_principal"), "money"),
        ("Juros", totals.get("juros"), "money"),
        ("Multas", totals.get("multa"), "money"),
        ("Descontos", totals.get("desconto"), "money"),
        ("Acréscimos", totals.get("acrescimo"), "money"),
        ("Valor estornado", totals.get("valor_estornado"), "money"),
        ("Movimentos normais", totals.get("quantidade_normais"), "number"),
        ("Movimentos estornados", totals.get("quantidade_estornados"), "number"),
        ("Ticket médio", totals.get("ticket_medio"), "money"),
        ("Taxa de estorno", ratio(totals.get("quantidade_estornados"), totals.get("quantidade_movimentos")), "percent"),
        ("Taxa de desconto", ratio(totals.get("desconto"), totals.get("valor_principal")), "percent"),
        ("Índice de pontualidade", totals.get("indice_pontualidade"), "percent"),
        ("Liquidações no prazo", totals.get("quantidade_no_prazo"), "number"),
        ("Liquidações em atraso", totals.get("quantidade_em_atraso"), "number"),
        ("Valor liquidado em atraso", totals.get("valor_liquido_em_atraso"), "money"),
        ("Média de dias de atraso", totals.get("media_dias_atraso"), "decimal"),
        ("Crescimento último mês", monthly_growth(monthly, "valor_liquido"), "percent"),
    ]
    add_kpis(ws, kpis)
    trend_start = 7 + len(kpis)
    ws.cell(trend_start, 1, "Período")
    ws.cell(trend_start, 2, "Valor líquido")
    for cell in ws[trend_start]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    for offset, row in enumerate(monthly, 1):
        ws.cell(trend_start + offset, 1, as_date(row.get("periodo"))).number_format = DATE
        ws.cell(trend_start + offset, 2, number(row.get("valor_liquido"))).number_format = MONEY
    add_line_chart(ws, "Evolução mensal", trend_start, len(monthly), 2, "E5")
    top = summary.get("por_parceiro", [])[:10]
    top_start = trend_start + len(monthly) + 3
    ws.cell(top_start, 1, f"TOP 10 {partner.upper()}S")
    ws.cell(top_start, 2, "Valor líquido")
    for cell in ws[top_start]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    for offset, row in enumerate(top, 1):
        ws.cell(top_start + offset, 1, row.get("parceiro_nome") or row.get("parceiro_id"))
        ws.cell(top_start + offset, 2, number(row.get("valor_liquido"))).number_format = MONEY
    add_bar_chart(
        ws,
        f"Concentração por {partner.lower()}",
        top_start,
        len(top),
        "E22",
    )
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22

    summary_columns = [
        ("periodo", "Período", "date"),
        ("quantidade_movimentos", "Movimentos", "number"),
        ("quantidade_normais", "Normais", "number"),
        ("quantidade_estornados", "Estornados", "number"),
        ("valor_principal", "Valor principal", "money"),
        ("juros", "Juros", "money"), ("multa", "Multa", "money"),
        ("desconto", "Desconto", "money"), ("acrescimo", "Acréscimo", "money"),
        ("valor_liquido", "Valor líquido", "money"),
        ("valor_estornado", "Valor estornado", "money"),
        ("quantidade_no_prazo", "Liquidações no prazo", "number"),
        ("quantidade_em_atraso", "Liquidações em atraso", "number"),
        ("quantidade_sem_vencimento", "Sem vencimento", "number"),
        ("valor_liquido_em_atraso", "Valor liquidado em atraso", "money"),
        ("media_dias_atraso", "Média dias atraso", "decimal"),
    ]
    headers, data, kinds = rows_from_dicts(monthly, summary_columns)
    add_sheet(wb, "Evolucao Mensal", headers, data, kinds)
    filial_columns = [
        ("filial_id", "Filial", None), ("filial_identificacao", "Identificação", None),
        *summary_columns[1:],
    ]
    headers, data, kinds = rows_from_dicts(summary.get("por_filial", []), filial_columns)
    add_sheet(wb, "Por Filial", headers, data, kinds)
    partner_columns = [
        ("parceiro_id", f"Código {partner.lower()}", None),
        ("parceiro_nome", partner, None), ("parceiro_cnpj_cpf", "CPF/CNPJ", "cnpj"),
        *summary_columns[1:],
    ]
    headers, data, kinds = rows_from_dicts(summary.get("por_parceiro", []), partner_columns)
    add_sheet(wb, f"Por {partner}", headers, data, kinds)

    prefix = "cliente" if receiving else "fornecedor"
    movement_columns = [
        ("movimento_id", "Controle movimento", None), ("parcela_id", "Controle parcela", None),
        ("titulo_id", "Controle título", None), ("numero_documento", "Documento", None),
        ("filial_id", "Filial", None), ("filial_identificacao", "Identificação filial", None),
        (f"{prefix}_id", f"Código {partner.lower()}", None),
        (f"{prefix}_nome", partner, None), (f"{prefix}_cnpj_cpf", "CPF/CNPJ", "cnpj"),
        ("data_emissao", "Data emissão", "date"),
        ("data_vencimento", "Data vencimento", "date"),
        ("data_movimento", "Data movimento", "date"),
        ("dias_em_relacao_vencimento", "Dias em relação ao vencimento", "number"),
        ("pontualidade", "Pontualidade", None),
        ("valor", "Valor principal", "money"),
        ("multa", "Multa", "money"), ("juros", "Juros", "money"),
        ("desconto", "Desconto", "money"), ("acrescimo", "Acréscimo", "money"),
        ("valor_complementar", "Valor complementar", "money"),
        ("valor_liquido", "Valor líquido", "money"), ("status", "Status", None),
        ("situacao", "Situação", None),
    ]
    headers, data, kinds = rows_from_dicts(movements, movement_columns)
    ws_mov = add_sheet(wb, "Movimentos", headers, data, kinds)
    status_col = headers.index("Situação") + 1
    if ws_mov.max_row > 1:
        letter = ws_mov.cell(1, status_col).column_letter
        ws_mov.conditional_formatting.add(
            f"{letter}2:{letter}{ws_mov.max_row}",
            CellIsRule(
                operator="equal",
                formula=['"ESTORNADO"'],
                fill=PatternFill("solid", fgColor=LIGHT_RED),
            ),
        )
    estornos = [row for row in movements if row.get("status") == "E"]
    headers, data, kinds = rows_from_dicts(estornos, movement_columns)
    add_sheet(wb, "Estornos", headers, data, kinds)
    atrasados = [
        row
        for row in movements
        if row.get("status") == "N" and row.get("pontualidade") == "ATRASADO"
    ]
    headers, data, kinds = rows_from_dicts(atrasados, movement_columns)
    add_sheet(wb, "Liquidados em Atraso", headers, data, kinds)
    punctuality_groups: dict[str, dict[str, float]] = defaultdict(
        lambda: {"quantidade": 0, "valor": 0.0}
    )
    for row in movements:
        status = row.get("pontualidade") or "SEM_CLASSIFICACAO"
        punctuality_groups[status]["quantidade"] += 1
        punctuality_groups[status]["valor"] += number(row.get("valor_liquido"))
    punctuality_rows = [
        [status, values["quantidade"], values["valor"]]
        for status, values in sorted(punctuality_groups.items())
    ]
    add_sheet(
        wb,
        "Pontualidade",
        ["Classificação", "Quantidade", "Valor líquido"],
        punctuality_rows,
        {"Quantidade": "number", "Valor líquido": "money"},
    )
    delay_ranking = sorted(
        summary.get("por_parceiro", []),
        key=lambda row: number(row.get("valor_liquido_em_atraso")),
        reverse=True,
    )
    headers, data, kinds = rows_from_dicts(delay_ranking, partner_columns)
    add_sheet(wb, "Ranking Atrasos", headers, data, kinds)
    add_methodology(wb, [
        ("Fonte", f"ActionAPI: {endpoint} e {endpoint}/resumo."),
        ("Movimento normal", "Status N; compõe os indicadores financeiros."),
        ("Estorno", "Status E; exibido separadamente e não somado ao caixa realizado."),
        ("Valor líquido", "Principal + multa + juros + acréscimo − desconto."),
        ("Pontualidade", "Compara a data do movimento com o vencimento original da parcela."),
        ("Atrasado", "Movimento normal realizado após o vencimento. Estornos são classificados separadamente."),
        ("Finalidade", "Análise de caixa realizado, concentração, encargos, descontos e estornos."),
    ])
    return save(
        wb,
        output,
        ["Painel", "Movimentos", "Estornos", "Pontualidade", "Metodologia"],
    ), summary


def generate_contabilidade(
    base_url: str,
    api_key: str,
    filters: dict[str, Any],
    output: Path,
) -> tuple[Path, dict]:
    summary = api_get(base_url, api_key, "/api/v1/executivo/contabilidade/resumo", filters)
    sintetico = api_get(base_url, api_key, "/api/v1/executivo/contabilidade/sintetico", filters)
    analysis = fetch_all(base_url, api_key, "/api/v1/bi/analise-contabil", filters)
    divergences = fetch_all(
        base_url,
        api_key,
        "/api/v1/conciliacao/financeiro-contabil/divergencias",
        {**filters, "tolerancia": 0.01},
    )
    reconciliation = api_get(
        base_url,
        api_key,
        "/api/v1/conciliacao/financeiro-contabil/resumo",
        {**filters, "tolerancia": 0.01},
    )
    dre = api_get(base_url, api_key, "/api/v1/dre", {})
    totals = summary.get("totalizadores", {})
    monthly = summary.get("evolucao_mensal", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Painel"
    title_block(ws, "CONTABILIDADE EXECUTIVA", period_subtitle(filters))
    store_issues = sum(1 for row in analysis if row.get("status_loja") != "OK")
    unmapped = [row for row in analysis if row.get("mapeamento_pendente")]
    unmapped_value = sum(abs(number(row.get("valor"))) for row in unmapped)
    kpis = [
        ("Débitos", totals.get("debitos"), "money"),
        ("Créditos", totals.get("creditos"), "money"),
        ("Diferença débito × crédito", totals.get("diferenca_dc"), "money"),
        ("Lançamentos", totals.get("quantidade_lancamentos"), "number"),
        ("Partidas", totals.get("quantidade_partidas"), "number"),
        ("Divergências financeiro × contábil", len(divergences), "number"),
        ("Linhas com alerta loja/CC", store_issues, "number"),
        ("Contas analisadas", len(summary.get("por_conta", [])), "number"),
        ("Contas sem mapeamento gerencial", len({row.get("conta_id") for row in unmapped}), "number"),
        ("Valor sem mapeamento gerencial (abs)", unmapped_value, "money"),
    ]
    add_kpis(ws, kpis)
    ws["A16"], ws["B16"], ws["C16"] = "Período", "Débitos", "Créditos"
    for cell in ws[16]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    for offset, row in enumerate(monthly, 1):
        ws.cell(16 + offset, 1, as_date(row.get("periodo"))).number_format = DATE
        ws.cell(16 + offset, 2, number(row.get("debitos"))).number_format = MONEY
        ws.cell(16 + offset, 3, number(row.get("creditos"))).number_format = MONEY
    add_line_chart(ws, "Débitos mensais", 16, len(monthly), 2, "E5")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22

    for name, key, columns in [
        ("Evolucao Mensal", "evolucao_mensal", [
            ("periodo", "Período", "date"),
            ("quantidade_lancamentos", "Lançamentos", "number"),
            ("quantidade_partidas", "Partidas", "number"),
            ("debitos", "Débitos", "money"), ("creditos", "Créditos", "money"),
            ("diferenca_dc", "Diferença D/C", "money"),
        ]),
        ("Balancete Executivo", "por_grupo", [
            ("grupo", "Grupo", None), ("grupo_descricao", "Descrição", None),
            ("quantidade_lancamentos", "Lançamentos", "number"),
            ("quantidade_partidas", "Partidas", "number"),
            ("debitos", "Débitos", "money"), ("creditos", "Créditos", "money"),
            ("diferenca_dc", "Diferença D/C", "money"),
        ]),
        ("Maiores Contas", "por_conta", [
            ("plano_contas", "Plano", None), ("conta", "Conta", None),
            ("conta_descricao", "Descrição", None),
            ("quantidade_lancamentos", "Lançamentos", "number"),
            ("quantidade_partidas", "Partidas", "number"),
            ("debitos", "Débitos", "money"), ("creditos", "Créditos", "money"),
            ("diferenca_dc", "Diferença D/C", "money"),
        ]),
    ]:
        headers, data, kinds = rows_from_dicts(summary.get(key, []), columns)
        add_sheet(wb, name, headers, data, kinds)

    add_synthetic_accounts_sheet(wb, "Plano de Contas", sintetico.get("contas", []))

    analysis_columns = [
        ("codigo_loja", "Código loja", None),
        ("codigo_loja_referencia_cc", "Loja referência CC", None),
        ("status_loja", "Status loja", None), ("conta_id", "Conta", None),
        ("cod_conta_contabil", "Conta formatada", None),
        ("desc_conta_contabil", "Descrição conta", None),
        ("ccusto_id", "Centro de custo", None),
        ("desc_centro_custo", "Descrição centro de custo", None),
        ("natureza_contabil", "Natureza", None),
        ("grupo_nivel_1", "Grupo nível 1", None),
        ("grupo_nivel_2", "Grupo nível 2", None),
        ("grupo_nivel_3", "Grupo nível 3", None),
        ("competencia", "Competência", "date"), ("safra", "Safra", None),
        ("valor", "Valor", "money"), ("ebitda", "Classificação EBITDA", None),
        ("qtd_partidas", "Partidas", "number"),
        ("mapeamento_pendente", "Sem mapeamento gerencial?", None),
    ]
    headers, data, kinds = rows_from_dicts(analysis, analysis_columns)
    add_sheet(wb, "Analise Gerencial", headers, data, kinds)
    quality = [row for row in analysis if row.get("status_loja") != "OK"]
    headers, data, kinds = rows_from_dicts(quality, analysis_columns)
    add_sheet(wb, "Alertas Loja CC", headers, data, kinds)
    headers, data, kinds = rows_from_dicts(unmapped, analysis_columns)
    add_sheet(wb, "Sem Mapeamento Gerencial", headers, data, kinds)

    divergence_columns = [
        ("tipo", "Tipo", None), ("titulo_id", "Título", None),
        ("filial_id", "Filial", None), ("parceiro_id", "Código parceiro", None),
        ("parceiro_nome", "Parceiro", None), ("numero_documento", "Documento", None),
        ("data_emissao", "Data emissão", "date"),
        ("valor_financeiro", "Valor financeiro", "money"),
        ("valor_contabil", "Valor contábil", "money"),
        ("diferenca_valor", "Diferença", "money"),
        ("status_conciliacao", "Status conciliação", None),
        ("regra_vinculo", "Regra de vínculo", None),
    ]
    headers, data, kinds = rows_from_dicts(divergences, divergence_columns)
    add_sheet(wb, "Divergencias Financeiro", headers, data, kinds)
    reconciliation_columns = [
        ("tipo", "Tipo", None), ("status_conciliacao", "Status", None),
        ("quantidade", "Quantidade", "number"),
        ("valor_financeiro", "Valor financeiro", "money"),
        ("valor_contabil", "Valor contábil", "money"),
        ("diferenca_valor", "Diferença", "money"),
    ]
    headers, data, kinds = rows_from_dicts(reconciliation.get("data", []), reconciliation_columns)
    add_sheet(wb, "Resumo Conciliacao", headers, data, kinds)
    dre_columns = [
        ("idre_id", "Linha DRE", None), ("descricao", "Descrição", None),
        ("nivel", "Nível", "number"), ("posicao_pai", "Linha pai", None),
        ("tipo", "Tipo", None), ("grupo", "Grupo", None), ("valor", "Valor", "money"),
    ]
    headers, data, kinds = rows_from_dicts(dre.get("linhas", []), dre_columns)
    add_sheet(wb, "DRE", headers, data, kinds)
    add_methodology(wb, [
        ("Fonte", "ActionAPI: /executivo/contabilidade/resumo, /executivo/contabilidade/sintetico, /bi/analise-contabil, /dre e /conciliacao."),
        ("Débito × crédito", "Diferença global é apresentada como indicador de qualidade e deve ser investigada quando material."),
        ("Análise gerencial", "Plano 1000002, contabilidade fiscal, excluindo origem ZR, com safra de 01/07 a 30/06."),
        ("Sem mapeamento gerencial", "Contas com lançamento mas sem classificação em analytics.conta_gerencial aparecem com Natureza/Grupo nível 1 = 'NAO CLASSIFICADO' (antes eram descartadas do relatório). Aba dedicada lista essas contas para priorizar o mapeamento."),
        ("Plano de Contas (sintético)", "Uma linha por conta do plano (sintética em destaque ou analítica), com débitos/créditos/saldo somando todas as analíticas descendentes pelo prefixo do código — mesma hierarquia do cadastro do SiAGRI (CONTASPL), permitindo conferir o balancete por nível como no relatório nativo."),
        ("Loja", "Código oficial vem do lançamento. Centro de custo serve somente como referência para detectar inconsistências."),
        ("Conciliação", "CP pela origem DP/controle; CR pela origem NE, documento, série, filial e parceiro."),
    ])
    return save(wb, output, ["Painel", "Analise Gerencial", "Plano de Contas", "DRE", "Metodologia"]), {
        **summary,
        "divergencias": divergences,
        "conciliacao": reconciliation,
    }


def generate_360(
    base_url: str,
    api_key: str,
    filters: dict[str, Any],
    output: Path,
) -> tuple[Path, dict]:
    data = api_get(base_url, api_key, "/api/v1/executivo/visao-360", filters)
    indicators = data.get("indicadores", {})
    fat = data.get("faturamento", {})
    rec = data.get("recebimentos", {})
    pag = data.get("pagamentos", {})
    top10_sales = sum(number(row.get("vendas_brutas")) for row in fat.get("por_cliente", [])[:10])
    top10_receipts = sum(number(row.get("valor_liquido")) for row in rec.get("por_parceiro", [])[:10])
    top10_payments = sum(number(row.get("valor_liquido")) for row in pag.get("por_parceiro", [])[:10])

    wb = Workbook()
    ws = wb.active
    ws.title = "Visao 360"
    title_block(ws, "VISÃO EXECUTIVA 360° — CEO/CFO", period_subtitle(filters))
    kpis = [
        ("Faturamento líquido", indicators.get("faturamento_liquido"), "money"),
        ("Vendas brutas", indicators.get("vendas_brutas"), "money"),
        ("Devoluções", indicators.get("devolucoes"), "money"),
        ("Recebimentos realizados", indicators.get("valor_recebido"), "money"),
        ("Pagamentos realizados", indicators.get("valor_pago"), "money"),
        ("Geração financeira de caixa", indicators.get("geracao_caixa_financeira"), "money"),
        ("Contas a receber em aberto", indicators.get("contas_receber_aberto"), "money"),
        ("Contas a pagar em aberto", indicators.get("contas_pagar_aberto"), "money"),
        ("Capital de giro financeiro líquido", indicators.get("capital_giro_liquido_financeiro"), "money"),
        ("Parcelas a receber abertas", indicators.get("parcelas_receber_abertas"), "number"),
        ("Parcelas a pagar abertas", indicators.get("parcelas_pagar_abertas"), "number"),
        ("Diferença débito × crédito", indicators.get("diferenca_debito_credito"), "money"),
        ("Conversão recebimento/faturamento", ratio(indicators.get("valor_recebido"), indicators.get("faturamento_liquido")), "percent"),
        ("Índice pagamentos/recebimentos", ratio(indicators.get("valor_pago"), indicators.get("valor_recebido")), "percent"),
        ("Taxa de devolução", ratio(indicators.get("devolucoes"), indicators.get("vendas_brutas")), "percent"),
        ("Concentração vendas — top 10 clientes", ratio(top10_sales, indicators.get("vendas_brutas")), "percent"),
        ("Concentração recebimentos — top 10", ratio(top10_receipts, indicators.get("valor_recebido")), "percent"),
        ("Concentração pagamentos — top 10", ratio(top10_payments, indicators.get("valor_pago")), "percent"),
    ]
    add_kpis(ws, kpis)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 24

    alerts: list[tuple[str, str, str]] = []
    if number(indicators.get("geracao_caixa_financeira")) < 0:
        alerts.append(("CRÍTICO", "Caixa", "Pagamentos superaram recebimentos no período."))
    if ratio(indicators.get("devolucoes"), indicators.get("vendas_brutas")) > 0.03:
        alerts.append(("ATENÇÃO", "Faturamento", "Taxa de devolução superior a 3%."))
    if abs(number(indicators.get("diferenca_debito_credito"))) > 0.01:
        alerts.append(("ATENÇÃO", "Contabilidade", "Há diferença material entre débitos e créditos no recorte."))
    if number(indicators.get("contas_pagar_aberto")) > number(indicators.get("contas_receber_aberto")):
        alerts.append(("ATENÇÃO", "Capital de giro", "Contas a pagar superam contas a receber em aberto."))
    if ratio(top10_sales, indicators.get("vendas_brutas")) > 0.50:
        alerts.append(("ATENÇÃO", "Concentração", "Os 10 maiores clientes representam mais de 50% das vendas."))
    if ratio(indicators.get("valor_recebido"), indicators.get("faturamento_liquido")) > 1.20:
        alerts.append((
            "INFORMAÇÃO",
            "Caixa",
            "Recebimentos superam o faturamento do período; há cobrança relevante de vendas de períodos anteriores.",
        ))
    if not alerts:
        alerts.append(("OK", "Geral", "Nenhum alerta automático ultrapassou os limites configurados."))
    add_sheet(wb, "Alertas Executivos", ["Nível", "Área", "Diagnóstico"], [list(row) for row in alerts])

    months: dict[str, dict[str, float]] = defaultdict(dict)
    for row in fat.get("evolucao_mensal", []):
        months[str(row.get("periodo"))[:10]]["faturamento"] = number(row.get("faturamento_liquido"))
    for row in rec.get("evolucao_mensal", []):
        months[str(row.get("periodo"))[:10]]["recebimentos"] = number(row.get("valor_liquido"))
    for row in pag.get("evolucao_mensal", []):
        months[str(row.get("periodo"))[:10]]["pagamentos"] = number(row.get("valor_liquido"))
    trend_rows = []
    for month in sorted(months):
        item = months[month]
        trend_rows.append([
            as_date(month),
            item.get("faturamento", 0),
            item.get("recebimentos", 0),
            item.get("pagamentos", 0),
            item.get("recebimentos", 0) - item.get("pagamentos", 0),
        ])
    ws_trend = add_sheet(
        wb,
        "Tendencias",
        ["Período", "Faturamento", "Recebimentos", "Pagamentos", "Geração de caixa"],
        trend_rows,
        {
            "Período": "date", "Faturamento": "money", "Recebimentos": "money",
            "Pagamentos": "money", "Geração de caixa": "money",
        },
    )
    add_line_chart(ws_trend, "Faturamento", 1, len(trend_rows), 2, "G3")
    indicator_rows = [[label, value, kind] for label, value, kind in kpis]
    add_sheet(wb, "Indicadores", ["Indicador", "Valor", "Tipo"], indicator_rows)
    add_methodology(wb, [
        ("Objetivo", "Visão consolidada para CEO/CFO com receita, caixa, capital de giro e qualidade contábil."),
        ("Caixa", "Recebimentos e pagamentos consideram apenas movimentos normais; estornos ficam fora do realizado."),
        ("Saldos abertos", "CR usa o snapshot oficial do SiAGRI; CP usa a reprodução validada de VALOR_ABERTO_PAGAR_DATA."),
        ("Alertas", "Regras automáticas simples, transparentes e ajustáveis no futuro frontend React."),
        ("Limitação", "Indicadores não substituem julgamento contábil, fiscal ou financeiro e devem ser analisados com os detalhes."),
    ])
    return save(wb, output, ["Visao 360", "Alertas Executivos", "Tendencias"]), data


def generate_open_financial_report(
    report: str,
    api_url: str,
    filters: dict[str, Any],
    output: Path,
    data_base: str | None = None,
) -> Path:
    """Executa os geradores avançados de saldos em aberto dentro do pacote.

    Saldo em aberto é uma posição "no momento", não um filtro por período:
    um título vencido há 2 anos ou que vence dentro de 8 meses continua em
    aberto hoje. Por isso NÃO amarramos vencimento_de/vencimento_ate ao
    período do relatório executivo — isso excluiria a maior parte do saldo
    real (vencidos antigos e parcelas futuras), como ficou evidente ao
    comparar com a planilha do controller (SGA-Fluxo Financeiro).
    """
    common = {
        "api_url": api_url,
        "arquivo": str(output),
        "vencimento_de": None,
        "vencimento_ate": None,
        "emissao_de": None,
        "emissao_ate": None,
        "filial_id": filters.get("filialId"),
        "safra": None,
        "bayer": None,
        "ano_contabil": None,
        "data_inicio": None,
        "data_fim": None,
        "data_base": data_base,
    }
    if report == "a-pagar":
        try:
            from gerar_relatorio_contas_pagar import generate_report
        except ModuleNotFoundError:
            from scripts.gerar_relatorio_contas_pagar import generate_report
        args = argparse.Namespace(
            **common,
            fornecedor_id=None,
            incluir_baixados=False,
        )
    else:
        try:
            from gerar_relatorio_contas_receber import generate_report
        except ModuleNotFoundError:
            from scripts.gerar_relatorio_contas_receber import generate_report
        args = argparse.Namespace(
            **common,
            cliente_id=None,
            tipo_documento=None,
            situacao=None,
            unidade_saldo=None,
            vendedor_id=None,
        )
    return generate_report(args)


def single_report_main(report: str) -> None:
    """Ponto de entrada compartilhado pelos scripts individuais."""
    labels = {
        "faturamento": "faturamento executivo",
        "a-receber": "contas a receber em aberto",
        "a-pagar": "contas a pagar em aberto",
        "recebimentos": "contas recebidas",
        "pagamentos": "contas pagas",
        "contabilidade": "contabilidade executiva",
        "360": "visão 360° CEO/CFO",
    }
    parser = argparse.ArgumentParser(description=f"Gera relatório de {labels[report]}.")
    parser.add_argument("--api-url", default=os.getenv("ACTIONAPI_URL", DEFAULT_API_URL))
    add_period_arguments(parser)
    if report in {"a-receber", "a-pagar"}:
        add_data_base_argument(parser)
    parser.add_argument("--filial-id")
    parser.add_argument("--arquivo", help="Arquivo .xlsx de saída.")
    args = parser.parse_args()
    prompt_period_if_needed(args)
    if report in {"a-receber", "a-pagar"}:
        prompt_data_base_if_needed(args)
        data_base = resolve_data_base(args)
    else:
        data_base = None
    try:
        data_inicio, data_fim, period_slug = resolve_period(args)
    except ValueError as exc:
        parser.error(str(exc))
    api_key = first_api_key()
    filters = {
        "dataInicio": data_inicio,
        "dataFim": data_fim,
        "filialId": args.filial_id,
    }
    suffix = period_slug
    default_names = {
        "faturamento": f"faturamento-executivo-{suffix}.xlsx",
        "a-receber": f"contas-a-receber-executivo-{suffix}.xlsx",
        "a-pagar": f"contas-a-pagar-executivo-{suffix}.xlsx",
        "recebimentos": f"contas-recebidas-executivo-{suffix}.xlsx",
        "pagamentos": f"contas-pagas-executivo-{suffix}.xlsx",
        "contabilidade": f"contabilidade-executiva-{suffix}.xlsx",
        "360": f"visao-360-ceo-cfo-{suffix}.xlsx",
    }
    output = (
        Path(args.arquivo).resolve()
        if args.arquivo
        else ROOT / "relatorios" / "executivo" / default_names[report]
    )
    if report == "faturamento":
        generate_faturamento(args.api_url, api_key, filters, output)
    elif report in {"a-receber", "a-pagar"}:
        generate_open_financial_report(report, args.api_url, filters, output, data_base=data_base)
    elif report in {"recebimentos", "pagamentos"}:
        generate_movements(report, args.api_url, api_key, filters, output)
    elif report == "contabilidade":
        generate_contabilidade(args.api_url, api_key, filters, output)
    else:
        generate_360(args.api_url, api_key, filters, output)


def main() -> None:
    args = parse_args()
    prompt_period_if_needed(args)
    try:
        data_inicio, data_fim, period_slug = resolve_period(args)
    except ValueError as exc:
        raise SystemExit(f"erro: {exc}") from exc
    api_key = first_api_key()
    filters = {
        "dataInicio": data_inicio,
        "dataFim": data_fim,
        "filialId": args.filial_id,
    }
    output_dir = Path(args.saida_dir).resolve()
    requested = {
        item.strip().lower()
        for item in args.relatorios.split(",")
        if item.strip()
    }
    if "todos" in requested:
        requested = {
            "faturamento", "a-receber", "a-pagar", "recebimentos",
            "pagamentos", "contabilidade", "360",
        }
    valid = {
        "faturamento", "a-receber", "a-pagar", "recebimentos",
        "pagamentos", "contabilidade", "360",
    }
    invalid = requested - valid
    if invalid:
        raise RuntimeError(f"Relatórios inválidos: {', '.join(sorted(invalid))}")
    if requested & {"a-receber", "a-pagar"}:
        prompt_data_base_if_needed(args)
    data_base = resolve_data_base(args)
    suffix = period_slug
    files: list[Path] = []

    if "faturamento" in requested:
        path, _ = generate_faturamento(
            args.api_url, api_key, filters,
            output_dir / f"faturamento-executivo-{suffix}.xlsx",
        )
        files.append(path)
    if "a-receber" in requested:
        path = generate_open_financial_report(
            "a-receber",
            args.api_url,
            filters,
            output_dir / f"contas-a-receber-executivo-{suffix}.xlsx",
            data_base=data_base,
        )
        files.append(path)
    if "a-pagar" in requested:
        path = generate_open_financial_report(
            "a-pagar",
            args.api_url,
            filters,
            output_dir / f"contas-a-pagar-executivo-{suffix}.xlsx",
            data_base=data_base,
        )
        files.append(path)
    if "recebimentos" in requested:
        path, _ = generate_movements(
            "recebimentos", args.api_url, api_key, filters,
            output_dir / f"contas-recebidas-executivo-{suffix}.xlsx",
        )
        files.append(path)
    if "pagamentos" in requested:
        path, _ = generate_movements(
            "pagamentos", args.api_url, api_key, filters,
            output_dir / f"contas-pagas-executivo-{suffix}.xlsx",
        )
        files.append(path)
    if "contabilidade" in requested:
        path, _ = generate_contabilidade(
            args.api_url, api_key, filters,
            output_dir / f"contabilidade-executiva-{suffix}.xlsx",
        )
        files.append(path)
    if "360" in requested:
        path, _ = generate_360(
            args.api_url, api_key, filters,
            output_dir / f"visao-360-ceo-cfo-{suffix}.xlsx",
        )
        files.append(path)

    print("[executivo] relatórios concluídos:", flush=True)
    for path in files:
        print(f"  - {path}", flush=True)


if __name__ == "__main__":
    main()
