#!/usr/bin/env python3
r"""Gera relatório avançado de Contas a Receber consumindo a ActionAPI.

Exemplo:

    .\.venv\Scripts\python.exe scripts\gerar_relatorio_contas_receber.py \
      --vencimento-de 2026-07-01 \
      --vencimento-ate 2026-12-31 \
      --arquivo relatorios\contas-receber-segundo-semestre.xlsx

O saldo em aberto vem da ActionAPI ou da reprodução histórica local. Os
recebimentos do período são consultados no PostgreSQL local via psql. A API key
é lida de API_KEYS no .env e nunca é gravada na planilha.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:
    from gerar_relatorios_executivos import (
        add_data_base_argument,
        add_period_arguments,
        has_period_argument,
        parse_user_date,
        prompt_data_base_if_needed,
        resolve_data_base,
        resolve_period,
    )
except ModuleNotFoundError:
    from scripts.gerar_relatorios_executivos import (
        add_data_base_argument,
        add_period_arguments,
        has_period_argument,
        parse_user_date,
        prompt_data_base_if_needed,
        resolve_data_base,
        resolve_period,
    )

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:
    print(
        "Dependência ausente. Execute:\n"
        "  py -m pip install -r scripts/requirements-relatorio-contas-pagar.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_API_URL = "http://127.0.0.1:3000"
PAGE_SIZE = 10_000
DEFAULT_PSQL = Path(r"C:\Program Files\PostgreSQL\16\bin\psql.exe")
LEGACY_OPEN_VALUE_KEY = "valor_em_aberto_" + "control" + "ler"

BLUE = "17365D"
LIGHT_RED = "F4CCCC"
LIGHT_ORANGE = "FCE4D6"
WHITE = "FFFFFF"
MONEY_FORMAT = 'R$ #,##0.00;[Red]-R$ #,##0.00'
NUMBER_FORMAT = "#,##0"
DECIMAL_FORMAT = '#,##0.00;[Red]-#,##0.00'
DATE_FORMAT = "dd/mm/yyyy"


DETAIL_COLUMNS = [
    ("filial_id", "Filial"),
    ("filial_identificacao", "Identificação da filial"),
    ("cliente_id", "Código cliente"),
    ("cliente_nome", "Cliente"),
    ("cliente_cnpj_cpf", "CPF/CNPJ"),
    ("data_emissao", "Data emissão"),
    ("data_vencimento", "Data vencimento"),
    ("parcela_nr", "Parcela"),
    ("valor_parcela", "Valor da parcela"),
    ("valor_baixado", "Valor recebido"),
    ("valor_em_aberto", "Valor em aberto"),
    ("unidade_saldo", "Unidade do saldo"),
    ("saldo_convertido_atual", "Saldo convertido atual"),
    ("indexador_id", "Código indexador"),
    ("indexador_descricao", "Indexador"),
    ("juros", "Juros"),
    ("multa", "Multa"),
    ("desconto", "Desconto"),
    ("acrescimo", "Acréscimo"),
    ("primeira_baixa", "Primeiro recebimento"),
    ("ultima_baixa", "Último recebimento"),
    ("vendedor_id", "Código vendedor do título"),
    ("vendedor_nome", "Vendedor do título"),
    ("vendedor_status", "Situação do vendedor"),
    ("numero_documento", "Número do documento"),
    ("serie_documento", "Série"),
    ("tipo_documento", "Tipo documento"),
    ("tipo_documento_descricao", "Descrição tipo documento"),
    ("natureza_tipo_documento", "Natureza documento"),
    ("historico", "Histórico/referência"),
    ("situacao", "Situação"),
    ("dias_atraso", "Dias em atraso"),
    ("faixa_vencimento", "Faixa de vencimento"),
    ("data_calculo", "Data do saldo"),
]

RECEBIMENTO_COLUMNS = [
    ("filial_id", "Filial"),
    ("filial_identificacao", "Identificação da filial"),
    ("cliente_id", "Código cliente"),
    ("cliente_nome", "Cliente"),
    ("cliente_cnpj_cpf", "CPF/CNPJ"),
    ("data_recebimento", "Data recebimento"),
    ("recibo_id", "Recibo"),
    ("recebimento_id", "Sequência recebimento"),
    ("data_emissao", "Data emissão"),
    ("data_vencimento", "Data vencimento"),
    ("parcela_nr", "Parcela"),
    ("valor_parcela", "Valor da parcela"),
    ("valor_baixado", "Valor recebido"),
    ("valor_liquido_baixas", "Valor líquido recebido"),
    ("juros", "Juros"),
    ("multa", "Multa"),
    ("desconto", "Desconto"),
    ("acrescimo", "Acréscimo"),
    ("valor_complementar", "Valor complementar"),
    ("unidade_saldo", "Unidade do saldo"),
    ("indexador_id", "Código indexador"),
    ("indexador_descricao", "Indexador"),
    ("primeira_baixa", "Primeiro recebimento"),
    ("ultima_baixa", "Último recebimento"),
    ("vendedor_id", "Código vendedor do título"),
    ("vendedor_nome", "Vendedor do título"),
    ("vendedor_status", "Situação do vendedor"),
    ("numero_documento", "Número do documento"),
    ("serie_documento", "Série"),
    ("tipo_documento", "Tipo documento"),
    ("tipo_documento_descricao", "Descrição tipo documento"),
    ("natureza_tipo_documento", "Natureza documento"),
    ("historico", "Histórico/referência"),
    ("status_recebimento", "Situação do recebimento"),
]

DIVERGENCE_COLUMNS = [
    ("motivo_divergencia", "Motivo provável da divergência"),
    ("diferenca_valor_em_aberto", "Diferença valor em aberto"),
    ("valor_em_aberto_snapshot", "Valor em aberto na planilha"),
    ("valor_em_aberto_recalculado", "Valor correto recalculado"),
    ("saldo_parcela", "Saldo do snapshot"),
    ("saldo_local", "Saldo local recalculado"),
    ("diferenca_saldo_local", "Diferença saldo local"),
    ("filial_id", "Filial"),
    ("filial_identificacao", "Identificação da filial"),
    ("cliente_id", "Código cliente"),
    ("cliente_nome", "Cliente"),
    ("data_emissao", "Data emissão"),
    ("data_vencimento", "Data vencimento"),
    ("parcela_nr", "Parcela"),
    ("valor_parcela", "Valor da parcela"),
    ("valor_baixado", "Valor recebido"),
    ("unidade_saldo", "Unidade do saldo"),
    ("saldo_convertido_atual", "Saldo convertido atual"),
    ("primeira_baixa", "Primeiro recebimento"),
    ("ultima_baixa", "Último recebimento"),
    ("numero_documento", "Número do documento"),
    ("serie_documento", "Série"),
    ("tipo_documento", "Tipo documento"),
    ("tipo_documento_descricao", "Descrição tipo documento"),
    ("historico", "Histórico/referência"),
    ("data_calculo", "Data do saldo"),
]

MONEY_HEADERS = {
    "Valor do título",
    "Valor da parcela",
    "Saldo convertido atual",
    "Valor recebido",
    "Valor líquido recebido",
    "Juros",
    "Multa",
    "Desconto",
    "Acréscimo",
    "Valor complementar",
    "Valor em aberto",
    "Valor em aberto na planilha",
    "Valor correto recalculado",
    "Diferença valor em aberto",
    "Saldo convertido",
    "Saldo vencido convertido",
    "Saldo a vencer convertido",
    "Recebido acumulado",
    "Em aberto",
    "Diferença",
}

DECIMAL_HEADERS = {
    "Saldo oficial",
    "Saldo reproduzido local",
    "Saldo do snapshot",
    "Saldo local recalculado",
    "Diferença saldo local",
    "Cotação de origem",
    "Cotação atual",
    "Saldo aberto",
    "Saldo vencido",
    "Saldo a vencer",
}

DATE_HEADERS = {
    "Data emissão",
    "Data vencimento",
    "Primeiro recebimento",
    "Último recebimento",
    "Data do saldo",
    "Primeiro vencimento",
    "Último vencimento",
}

INTEGER_HEADERS = {
    "Parcela",
    "Dias em atraso",
    "Quantidade de recebimentos",
    "Quantidade de parcelas",
    "Quantidade de títulos",
    "Quantidade de clientes",
    "Maior atraso em dias",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera relatório Excel avançado de Contas a Receber via ActionAPI."
    )
    parser.add_argument("--api-url", default=os.getenv("ACTIONAPI_URL", DEFAULT_API_URL))
    parser.add_argument("--arquivo", help="Caminho do arquivo .xlsx de saída.")
    parser.add_argument("--vencimento-de", help="Vencimento inicial, AAAA-MM-DD.")
    parser.add_argument("--vencimento-ate", help="Vencimento final, AAAA-MM-DD.")
    parser.add_argument("--emissao-de", help="Emissão inicial, AAAA-MM-DD.")
    parser.add_argument("--emissao-ate", help="Emissão final, AAAA-MM-DD.")
    parser.add_argument("--recebimento-de", help="Recebimento inicial, AAAA-MM-DD.")
    parser.add_argument("--recebimento-ate", help="Recebimento final, AAAA-MM-DD.")
    parser.add_argument("--filial-id", help="Código da filial.")
    parser.add_argument("--cliente-id", help="Código do cliente.")
    parser.add_argument("--tipo-documento", help="Código do tipo de documento.")
    add_period_arguments(parser)
    add_data_base_argument(parser)
    parser.add_argument(
        "--situacao",
        choices=["VENCIDA", "VENCE_HOJE", "A_VENCER", "CREDITO_EM_ABERTO"],
    )
    parser.add_argument("--unidade-saldo", help="R$, SJ$, US$ ou ER.")
    parser.add_argument("--vendedor-id", help="Código do vendedor.")
    return parser.parse_args()


def apply_period_selection(args: argparse.Namespace) -> None:
    """Aplica filtro de vencimento/emissão somente se informado explicitamente.

    Saldo em aberto é uma posição "no momento" (mesma regra do saldo
    histórico): por padrão trazemos o snapshot completo, sem perguntar
    período. O filtro só existe para quem passa --safra/--bayer/--ano-contabil/
    --data-inicio/--data-fim ou --vencimento-de/--ate/--emissao-de/--ate
    explicitamente na linha de comando.
    """
    explicit_dates = any(
        (args.vencimento_de, args.vencimento_ate, args.emissao_de, args.emissao_ate)
    )
    if has_period_argument(args):
        if explicit_dates:
            raise SystemExit(
                "erro: use o período do menu/parâmetros gerais ou os filtros "
                "específicos de vencimento/emissão, não ambos."
            )
        try:
            args.vencimento_de, args.vencimento_ate, _slug = resolve_period(args)
        except ValueError as exc:
            raise SystemExit(f"erro: {exc}") from exc
        return

    for field in ("vencimento_de", "vencimento_ate", "emissao_de", "emissao_ate"):
        value = getattr(args, field)
        if value:
            try:
                parsed = parse_user_date(value, f"--{field.replace('_', '-')}")
            except ValueError as exc:
                raise SystemExit(f"erro: {exc}") from exc
            setattr(args, field, parsed.isoformat())


def resolve_recebimento_period(args: argparse.Namespace, data_base: str) -> tuple[str, str]:
    """Resolve o período da aba Recebimentos.

    O saldo em aberto é uma posição; recebimentos são movimento. Quando o
    usuário não informa um intervalo específico, usamos o mês da data-base até
    a própria data-base.
    """
    base = parse_user_date(data_base, "data-base")
    start_value = getattr(args, "recebimento_de", None)
    end_value = getattr(args, "recebimento_ate", None)
    if not start_value and not end_value:
        return base.replace(day=1).isoformat(), base.isoformat()

    if start_value:
        start = parse_user_date(start_value, "--recebimento-de")
    elif end_value:
        end_tmp = parse_user_date(end_value, "--recebimento-ate")
        start = end_tmp.replace(day=1)
    else:
        start = base.replace(day=1)

    if end_value:
        end = parse_user_date(end_value, "--recebimento-ate")
    else:
        end = base

    if start > end:
        raise SystemExit("erro: --recebimento-de não pode ser maior que --recebimento-ate.")
    return start.isoformat(), end.isoformat()


def read_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value.startswith(("'", '"')):
            value = value[1:-1]
        values[key.strip()] = value
    return values


def psql_executable() -> str:
    configured = os.getenv("PSQL_PATH")
    if configured:
        return configured
    return str(DEFAULT_PSQL if DEFAULT_PSQL.exists() else "psql")


def pg_json_rows(query: str) -> list[dict[str, Any]]:
    env_file = read_env(ROOT / ".env")
    env = {**os.environ}
    mapping = {
        "PGHOST": env_file.get("PG_HOST"),
        "PGPORT": env_file.get("PG_PORT"),
        "PGDATABASE": env_file.get("PG_DATABASE"),
        "PGUSER": env_file.get("PG_USER"),
        "PGPASSWORD": env_file.get("PG_PASS"),
    }
    for key, value in mapping.items():
        if value and not env.get(key):
            env[key] = value
    env.setdefault("PGCLIENTENCODING", "LATIN1")
    result = subprocess.run(
        [
            psql_executable(),
            "-X",
            "-q",
            "-t",
            "-A",
            "-v",
            "ON_ERROR_STOP=1",
            "-c",
            f"COPY (SELECT row_to_json(q)::text FROM ({query}) q) TO STDOUT",
        ],
        capture_output=True,
        text=True,
        encoding="latin1",
        env=env,
        cwd=str(ROOT),
    )
    if result.returncode != 0:
        raise RuntimeError(f"Falha ao consultar PostgreSQL via psql: {result.stderr}")
    return [json.loads(line) for line in result.stdout.splitlines() if line.strip()]


def first_api_key() -> str:
    env = {**read_env(ROOT / ".env"), **os.environ}
    keys = [item.strip() for item in env.get("API_KEYS", "").split(",") if item.strip()]
    if not keys:
        raise RuntimeError("API_KEYS não está configurada no .env.")
    return keys[0]


def api_get(base_url: str, api_key: str, endpoint: str, params: dict[str, Any]) -> dict:
    clean = {key: value for key, value in params.items() if value not in (None, "")}
    url = f"{base_url.rstrip('/')}{endpoint}"
    if clean:
        url += "?" + urlencode(clean)
    request = Request(url, headers={"X-API-Key": api_key, "Accept": "application/json"})
    try:
        with urlopen(request, timeout=120) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"ActionAPI respondeu HTTP {exc.code}: {body}") from exc
    except URLError as exc:
        raise RuntimeError(f"Não foi possível acessar a ActionAPI em {url}: {exc}") from exc


NODE_SALDO_HISTORICO = ROOT / "packages" / "etl" / "src" / "scripts" / "saldo-aberto-historico.js"


def fetch_historico_cr(data_base: str) -> dict:
    """Reproduz o saldo em aberto de CR numa data-base passada/futura via PostgreSQL.

    Não usa a ActionAPI nem o Oracle: reaproveita, parametrizada, a mesma
    fórmula validada contra VALOR_ABERTO_RECEBER_DATA (zero divergências).
    Inclui também as parcelas já recebidas (saldo ≈ 0, com baixa) dos títulos
    que ainda têm saldo em aberto, para a conciliação.
    """
    print(
        f"[relatorio-receber] reproduzindo saldo em {data_base} via PostgreSQL "
        "(sem Oracle, fórmula validada)...",
        flush=True,
    )
    result = subprocess.run(
        [
            "node",
            str(NODE_SALDO_HISTORICO),
            "--tipo",
            "CR",
            "--data-base",
            data_base,
            "--incluir-baixadas",
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        cwd=str(NODE_SALDO_HISTORICO.parents[2]),
    )
    if result.returncode != 0:
        raise RuntimeError(f"Falha ao calcular saldo histórico de CR: {result.stderr}")
    payload = json.loads(result.stdout)
    return payload["cr"]


def fetch_recebidas_cr(data_base: str, titulos_abertos: set[str]) -> list[dict]:
    """Parcelas já recebidas (saldo ≈ 0) dos títulos em aberto, via PostgreSQL.

    Usado no caminho "hoje", em que a API de CR só enxerga o snapshot em aberto
    (raw.duplicatas_saldo). As recebidas vêm da reprodução local (raw.duplicatas),
    limitadas aos títulos que ainda têm saldo, para a conciliação fechar.
    """
    print(
        f"[relatorio-receber] buscando parcelas recebidas em {data_base} via PostgreSQL...",
        flush=True,
    )
    payload = fetch_historico_cr(data_base)
    return [
        row
        for row in payload["rows"]
        if abs(as_number(row.get("saldo_parcela"))) <= 0.01
        and row.get("titulo_id") in titulos_abertos
    ]


def fetch_recebimentos_cr(data_inicio: str, data_fim: str) -> list[dict]:
    """Todos os recebimentos normais do período informado."""
    for label, value in (("data inicial", data_inicio), ("data final", data_fim)):
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
            raise RuntimeError(f"Período de recebimentos inválido: {label}={value!r}")

    print(
        f"[relatorio-receber] buscando recebimentos de {data_inicio} a {data_fim}...",
        flush=True,
    )
    query = f"""
      SELECT
        r.id AS recebimento_id,
        r.recibo_id,
        r.status AS status_recebimento,
        r.data_pagamento AS data_recebimento,
        r.parcela_id,
        d.nf_id AS titulo_id,
        COALESCE(r.filial_id, d.filial_id) AS filial_id,
        COALESCE(fil_cli.razao_social, fil_forn.razao_social, fil._dados->>'FANT_EMP')
          AS filial_nome,
        fil._dados->>'FANT_EMP' AS filial_fantasia,
        fil._dados->>'IDEN_EMP' AS filial_identificacao,
        COALESCE(r.cliente_id, d._dados->>'CODI_TRA') AS cliente_id,
        cli.razao_social AS cliente_nome,
        cli.cgc_cnpj AS cliente_cnpj_cpf,
        COALESCE(
          NULLIF(d._dados->>'COD1_PES', ''),
          NULLIF(nf._dados->>'COD1_PES', '')
        ) AS vendedor_id,
        vend._dados->>'NOME_PES' AS vendedor_nome,
        CASE
          WHEN COALESCE(
            NULLIF(d._dados->>'COD1_PES', ''),
            NULLIF(nf._dados->>'COD1_PES', '')
          ) IS NULL
            THEN 'NAO_INFORMADO'
          WHEN vend.id IS NULL
            THEN 'CADASTRO_NAO_SINCRONIZADO'
          WHEN vend._dados->>'SITU_PES' = 'A'
            THEN 'ATIVO'
          WHEN vend._dados->>'SITU_PES' = 'I'
            THEN 'INATIVO'
          ELSE COALESCE(vend._dados->>'SITU_PES', 'SEM_SITUACAO')
        END AS vendedor_status,
        COALESCE(d.tipo_documento, r.tipo_doc) AS tipo_documento,
        td.descricao AS tipo_documento_descricao,
        td.tipo AS natureza_tipo_documento,
        d._dados->>'HISTORICO' AS historico,
        COALESCE(ft.numero_documento, nf._dados->>'NOTA_NOT') AS numero_documento,
        COALESCE(ft.serie_documento, nf._dados->>'SERI_NOT') AS serie_documento,
        NULLIF(d._dados->>'NPAR_REC', '')::INT AS parcela_nr,
        d.data_emissao,
        d.data_vencimento,
        NULLIF(d._dados->>'VLOR_REC', '')::NUMERIC AS valor_parcela,
        COALESCE(idx.abreviatura, 'R$') AS unidade_saldo,
        COALESCE(r.indexador_id, d.indexador_id) AS indexador_id,
        idx.descricao AS indexador_descricao,
        r.valor AS valor_baixado,
        COALESCE(r.multa, 0) AS multa,
        COALESCE(r.juros, 0) AS juros,
        COALESCE(r.desconto, 0) AS desconto,
        COALESCE(r.acrescimo, 0) AS acrescimo,
        COALESCE(r.valor_complementar, 0) AS valor_complementar,
        COALESCE(r.valor, 0)
          + COALESCE(r.multa, 0)
          + COALESCE(r.juros, 0)
          + COALESCE(r.acrescimo, 0)
          + COALESCE(r.valor_complementar, 0)
          - COALESCE(r.desconto, 0) AS valor_liquido_baixas,
        r.data_pagamento AS primeira_baixa,
        r.data_pagamento AS ultima_baixa
      FROM raw.recebimentos r
      LEFT JOIN raw.duplicatas d ON d.id = r.parcela_id
      LEFT JOIN raw.faturamento nf ON nf.id = d.nf_id
      LEFT JOIN raw.financeiro_titulos ft
        ON ft.tipo = 'CR' AND ft.titulo_id = d.nf_id
      LEFT JOIN raw.clientes cli ON cli.id = COALESCE(r.cliente_id, d._dados->>'CODI_TRA')
      LEFT JOIN raw.tipos_documento td ON td.id = COALESCE(d.tipo_documento, r.tipo_doc)
      LEFT JOIN raw.indexadores idx ON idx.id = COALESCE(r.indexador_id, d.indexador_id)
      LEFT JOIN raw.vendedores vend ON vend.id = COALESCE(
        NULLIF(d._dados->>'COD1_PES', ''),
        NULLIF(nf._dados->>'COD1_PES', '')
      )
      LEFT JOIN raw.filiais fil ON fil.id = COALESCE(r.filial_id, d.filial_id)
      LEFT JOIN raw.clientes fil_cli ON fil_cli.id = fil._dados->>'COD1_TRA'
      LEFT JOIN raw.fornecedores fil_forn ON fil_forn.id = fil._dados->>'COD1_TRA'
      WHERE r.status = 'N'
        AND r.data_pagamento BETWEEN DATE '{data_inicio}' AND DATE '{data_fim}'
      ORDER BY r.data_pagamento, r.id
    """
    return pg_json_rows(query)


def fetch_all(base_url: str, api_key: str, filters: dict[str, Any]) -> list[dict]:
    rows: list[dict] = []
    page = 1
    while True:
        payload = api_get(
            base_url,
            api_key,
            "/api/v1/financeiro/contas-receber",
            {**filters, "page": page, "pageSize": PAGE_SIZE},
        )
        current = payload.get("data", [])
        rows.extend(current)
        if not current or len(rows) >= int(payload.get("total", len(rows))):
            return rows
        page += 1


def as_number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def report_open_value(row: dict[str, Any]) -> float:
    """Calcula a coluna monetária "Valor em aberto" usada no relatório.

    Em R$, o saldo já está em reais. Para títulos indexados, o valor é
    convertido pela cotação de origem do título e arredondado parcela a parcela.
    """
    if row.get("valor_em_aberto") is not None:
        return as_number(row.get("valor_em_aberto"))
    if row.get(LEGACY_OPEN_VALUE_KEY) is not None:
        return as_number(row.get(LEGACY_OPEN_VALUE_KEY))

    saldo = Decimal(str(row.get("saldo_parcela") or 0))
    if row.get("unidade_saldo") == "R$":
        value = saldo
    else:
        quotation = Decimal(str(row.get("valor_indexador_origem") or 0))
        value = saldo * quotation
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def report_open_value_from_balance(row: dict[str, Any], balance_key: str) -> float:
    saldo = Decimal(str(row.get(balance_key) or 0))
    if row.get("unidade_saldo") == "R$":
        value = saldo
    else:
        quotation = Decimal(str(row.get("valor_indexador_origem") or 0))
        value = saldo * quotation
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def enrich_open_values(
    rows: list[dict[str, Any]],
    totals: dict[str, Any],
    units: list[dict],
    clients: list[dict],
) -> None:
    by_unit: dict[str, float] = defaultdict(float)
    by_client: dict[tuple[str, str, str], float] = defaultdict(float)
    for row in rows:
        value = report_open_value(row)
        row["valor_em_aberto"] = value
        unit = row.get("unidade_saldo") or "R$"
        by_unit[unit] += value
        by_client[
            (
                str(row.get("filial_id") or ""),
                str(row.get("cliente_id") or ""),
                unit,
            )
        ] += value

    totals["valor_em_aberto"] = round(sum(by_unit.values()), 2)
    for unit in units:
        name = unit.get("unidade_saldo") or "R$"
        unit["valor_em_aberto"] = round(by_unit.get(name, 0), 2)
    for client in clients:
        key = (
            str(client.get("filial_id") or ""),
            str(client.get("cliente_id") or ""),
            client.get("unidade_saldo") or "R$",
        )
        client["valor_em_aberto"] = round(by_client.get(key, 0), 2)


def as_date(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def classify_expiry(row: dict[str, Any], data_base: str) -> None:
    """Padroniza as faixas de vencimento usadas em todas as abas do relatório."""
    due = as_date(row.get("data_vencimento"))
    base = as_date(data_base)
    if due is None or base is None:
        row["faixa_vencimento"] = "SEM_CLASSIFICACAO"
        return
    if as_number(row.get("saldo_parcela")) < -0.01:
        row["faixa_vencimento"] = "CREDITO_EM_ABERTO"
        return

    delta = (due.date() - base.date()).days
    if delta == 0:
        row["faixa_vencimento"] = "VENCE_HOJE"
        return

    prefix = "VENCE_EM" if delta > 0 else "VENCIDO"
    days = abs(delta)
    if days <= 30:
        suffix = "1_A_30_DIAS"
    elif days <= 60:
        suffix = "31_A_60_DIAS"
    elif days <= 90:
        suffix = "61_A_90_DIAS"
    elif days <= 120:
        suffix = "91_A_120_DIAS"
    elif days <= 180:
        suffix = "121_A_180_DIAS"
    elif days <= 360:
        suffix = "181_A_360_DIAS"
    else:
        row["faixa_vencimento"] = (
            "VENCE_ACIMA_360_DIAS" if delta > 0 else "VENCIDO_ACIMA_360_DIAS"
        )
        return
    row["faixa_vencimento"] = f"{prefix}_{suffix}"


def normalize_report_rows(rows: list[dict[str, Any]], data_base: str) -> None:
    for row in rows:
        classify_expiry(row, data_base)


def same_day(value: Any, reference: str) -> bool:
    parsed = as_date(value)
    ref = as_date(reference)
    return bool(parsed and ref and parsed.date() == ref.date())


def divergence_reason(row: dict[str, Any], data_base: str) -> str:
    diff = as_number(row.get("diferenca_valor_em_aberto"))
    if abs(diff) <= 0.01:
        diff = as_number(row.get("diferenca_saldo_local"))
    if abs(diff) <= 0.01:
        return "Sem divergência relevante."

    local_smaller = diff < -0.01
    local_bigger = diff > 0.01
    today_report = data_base == date.today().isoformat()
    has_payment_on_base = same_day(row.get("ultima_baixa"), data_base)
    unit = row.get("unidade_saldo") or "R$"

    if today_report and has_payment_on_base:
        if local_smaller:
            return (
                "Movimento de recebimento na própria data-base reduziu o valor "
                "recalculado. Relatórios gerados durante o dia podem mudar após "
                "novas baixas/sincronizações."
            )
        if local_bigger:
            return (
                "Movimento na própria data-base alterou o valor recalculado para "
                "maior. Verificar estorno, ajuste ou alteração do título no dia."
            )

    if today_report:
        if local_smaller:
            return (
                "Valor recalculado menor que o valor da planilha. Provável baixa, agrupamento ou "
                "alteração sincronizada após a geração do snapshot do dia."
            )
        if local_bigger:
            return (
                "Valor recalculado maior que o valor da planilha. Provável estorno, alteração de "
                "parcela ou ajuste sincronizado após a geração do snapshot do dia."
            )

    if unit != "R$":
        return (
            "Contrato indexado com diferença entre valor da planilha e valor recalculado. "
            "Verificar cotação de origem, baixas e agrupamentos."
        )

    if local_smaller:
        return "Valor recalculado menor que o valor da planilha; verificar baixas/agrupamentos aplicados."
    if local_bigger:
        return "Valor recalculado maior que o valor da planilha; verificar estornos ou alterações no título."
    return "Diferença entre valor da planilha e valor recalculado; investigar histórico da parcela."


def cnpj_format(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return str(value or "")


def converted_value(key: str, value: Any) -> Any:
    if key in {
        "valor_titulo",
        "valor_parcela",
        "saldo_parcela",
        "valor_em_aberto",
        "saldo_convertido_atual",
        "valor_indexador_origem",
        "valor_indexador_atual",
        "valor_baixado",
        "valor_liquido_baixas",
        "juros",
        "multa",
        "desconto",
        "acrescimo",
        "valor_complementar",
        "saldo_local",
        "diferenca_saldo_local",
    }:
        return as_number(value)
    if key in {
        "data_emissao",
        "data_vencimento",
        "data_recebimento",
        "primeira_baixa",
        "ultima_baixa",
        "data_calculo",
    }:
        return as_date(value)
    if key in {"parcela_nr", "dias_atraso", "qtd_baixas"}:
        return as_number(value)
    if key == "cliente_cnpj_cpf":
        return cnpj_format(value)
    return value


def selected_rows(
    rows: list[dict],
    columns: list[tuple[str, str]],
) -> list[list[Any]]:
    return [
        [converted_value(key, row.get(key)) for key, _header in columns]
        for row in rows
    ]


def detail_rows(rows: list[dict]) -> list[list[Any]]:
    return selected_rows(rows, DETAIL_COLUMNS)


def recebimento_rows(rows: list[dict]) -> list[list[Any]]:
    return selected_rows(rows, RECEBIMENTO_COLUMNS)


def divergence_rows(rows: list[dict], data_base: str) -> list[list[Any]]:
    divergences: list[dict[str, Any]] = []
    for row in rows:
        has_local_comparison = (
            row.get("saldo_local") is not None
            or row.get("diferenca_saldo_local") is not None
        )
        if not has_local_comparison:
            continue

        enriched = dict(row)
        if enriched.get("saldo_local") is None:
            enriched["saldo_local"] = (
                as_number(enriched.get("saldo_parcela"))
                + as_number(enriched.get("diferenca_saldo_local"))
            )
        snapshot_value = report_open_value_from_balance(enriched, "saldo_parcela")
        recalculated_value = report_open_value_from_balance(enriched, "saldo_local")
        value_diff = round(recalculated_value - snapshot_value, 2)
        enriched["valor_em_aberto_snapshot"] = snapshot_value
        enriched["valor_em_aberto_recalculado"] = recalculated_value
        enriched["diferenca_valor_em_aberto"] = value_diff

        if (
            abs(value_diff) <= 0.01
            and abs(as_number(enriched.get("diferenca_saldo_local"))) <= 0.01
        ):
            continue

        enriched["motivo_divergencia"] = divergence_reason(enriched, data_base)
        divergences.append(enriched)

    divergences.sort(
        key=lambda item: abs(as_number(item.get("diferenca_valor_em_aberto"))),
        reverse=True,
    )
    return selected_rows(divergences, DIVERGENCE_COLUMNS)


def client_rows(rows: list[dict]) -> list[list[Any]]:
    headers = [
        "Filial",
        "Identificação da filial",
        "Código cliente",
        "Cliente",
        "CPF/CNPJ",
        "Unidade do saldo",
        "Quantidade de parcelas",
        "Quantidade de títulos",
        "Valor das parcelas",
        "Valor recebido",
        "Valor em aberto",
        "Saldo convertido",
        "Saldo vencido",
        "Saldo vencido convertido",
        "Saldo a vencer",
        "Primeiro vencimento",
        "Último vencimento",
        "Maior atraso em dias",
    ]
    output = []
    for row in sorted(
        rows,
        key=lambda item: as_number(item.get("saldo_convertido_atual")),
        reverse=True,
    ):
        output.append(
            [
                row.get("filial_id"),
                row.get("filial_identificacao"),
                row.get("cliente_id"),
                row.get("cliente_nome"),
                cnpj_format(row.get("cliente_cnpj_cpf")),
                row.get("unidade_saldo"),
                as_number(row.get("qtd_parcelas")),
                as_number(row.get("qtd_titulos")),
                as_number(row.get("valor_parcelas")),
                as_number(row.get("valor_baixado")),
                as_number(row.get("valor_em_aberto")),
                as_number(row.get("saldo_convertido_atual")),
                as_number(row.get("saldo_vencido")),
                as_number(row.get("saldo_vencido_convertido")),
                as_number(row.get("saldo_a_vencer")),
                as_date(row.get("primeiro_vencimento")),
                as_date(row.get("ultimo_vencimento")),
                as_number(row.get("maior_atraso_dias")),
            ]
        )
    return [headers, *output]


def unit_rows(rows: list[dict]) -> list[list[Any]]:
    headers = [
        "Unidade do saldo",
        "Quantidade de parcelas",
        "Quantidade de títulos",
        "Quantidade de clientes",
        "Saldo aberto",
        "Valor em aberto",
        "Saldo convertido",
        "Saldo vencido",
        "Saldo vencido convertido",
    ]
    return [
        headers,
        *[
            [
                row.get("unidade_saldo"),
                as_number(row.get("qtd_parcelas")),
                as_number(row.get("qtd_titulos")),
                as_number(row.get("qtd_clientes")),
                as_number(row.get("saldo_aberto")),
                as_number(row.get("valor_em_aberto")),
                as_number(row.get("saldo_convertido_atual")),
                as_number(row.get("saldo_vencido")),
                as_number(row.get("saldo_vencido_convertido")),
            ]
            for row in rows
        ],
    ]


def expiry_rows(rows: list[dict]) -> list[list[Any]]:
    order = [
        "CREDITO_EM_ABERTO",
        "VENCIDO_ACIMA_360_DIAS",
        "VENCIDO_181_A_360_DIAS",
        "VENCIDO_121_A_180_DIAS",
        "VENCIDO_91_A_120_DIAS",
        "VENCIDO_61_A_90_DIAS",
        "VENCIDO_31_A_60_DIAS",
        "VENCIDO_1_A_30_DIAS",
        "VENCE_HOJE",
        "VENCE_EM_1_A_30_DIAS",
        "VENCE_EM_31_A_60_DIAS",
        "VENCE_EM_61_A_90_DIAS",
        "VENCE_EM_91_A_120_DIAS",
        "VENCE_EM_121_A_180_DIAS",
        "VENCE_EM_181_A_360_DIAS",
        "VENCE_ACIMA_360_DIAS",
    ]
    grouped: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "parcelas": 0,
            "titulos": set(),
            "valor_em_aberto": 0.0,
            "convertido": 0.0,
        }
    )
    for row in rows:
        key = (
            row.get("faixa_vencimento") or "SEM_CLASSIFICACAO",
            row.get("unidade_saldo") or "R$",
        )
        group = grouped[key]
        group["parcelas"] += 1
        group["titulos"].add(row.get("titulo_id"))
        group["valor_em_aberto"] += as_number(row.get("valor_em_aberto"))
        group["convertido"] += as_number(row.get("saldo_convertido_atual"))

    output = [[
        "Faixa de vencimento",
        "Unidade do saldo",
        "Quantidade de parcelas",
        "Quantidade de títulos",
        "Valor em aberto",
        "Saldo convertido",
    ]]
    for key in sorted(
        grouped,
        key=lambda item: (order.index(item[0]) if item[0] in order else 99, item[1]),
    ):
        group = grouped[key]
        output.append([
            key[0],
            key[1],
            group["parcelas"],
            len(group["titulos"]),
            group["valor_em_aberto"],
            group["convertido"],
        ])
    return output


def split_aberto_recebidas(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Separa parcelas em aberto (saldo ≠ 0) das já recebidas (saldo ≈ 0)."""
    abertas: list[dict] = []
    recebidas: list[dict] = []
    for row in rows:
        if abs(as_number(row.get("saldo_parcela"))) > 0.01:
            abertas.append(row)
        else:
            recebidas.append(row)
    return abertas, recebidas


def valor_assinado(row: dict) -> float:
    """Valor da parcela com o mesmo sinal do saldo (crédito = negativo)."""
    valor = as_number(row.get("valor_parcela"))
    return -valor if row.get("natureza_tipo_documento") == "C" else valor


def conciliacao_rows(rows: list[dict]) -> list[list[Any]]:
    """Uma linha por título: Valor das parcelas, Recebido acumulado e Em aberto.

    A coluna "Diferença" (= Valor − Recebido − Em aberto) fica ~0 para títulos em
    R$. Contratos indexados (SJ$, US$, ER) podem gerar diferença, pois o saldo é
    mantido na unidade e o recebido em reais — é esperado e está na Metodologia.
    """
    headers = [
        "Filial",
        "Código cliente",
        "Cliente",
        "CPF/CNPJ",
        "Controle do título",
        "Unidade do saldo",
        "Quantidade de parcelas",
        "Valor das parcelas",
        "Recebido acumulado",
        "Em aberto",
        "Diferença",
        "Situação",
    ]
    grupos: dict[Any, dict[str, Any]] = {}
    for row in rows:
        chave = (row.get("filial_id"), row.get("cliente_id"), row.get("titulo_id"))
        grupo = grupos.get(chave)
        if grupo is None:
            grupo = {
                "filial_id": row.get("filial_id"),
                "cliente_id": row.get("cliente_id"),
                "cliente_nome": row.get("cliente_nome"),
                "cliente_cnpj_cpf": row.get("cliente_cnpj_cpf"),
                "titulo_id": row.get("titulo_id"),
                "unidade_saldo": row.get("unidade_saldo") or "R$",
                "qtd": 0,
                "valor": 0.0,
                "recebido": 0.0,
                "aberto": 0.0,
            }
            grupos[chave] = grupo
        grupo["qtd"] += 1
        grupo["valor"] += valor_assinado(row)
        grupo["recebido"] += as_number(row.get("valor_baixado"))
        grupo["aberto"] += as_number(row.get("saldo_parcela"))

    output: list[list[Any]] = []
    for grupo in sorted(grupos.values(), key=lambda item: item["aberto"], reverse=True):
        aberto = grupo["aberto"]
        recebido = grupo["recebido"]
        if abs(aberto) <= 0.01:
            situacao = "RECEBIDO"
        elif recebido > 0.01:
            situacao = "PARCIAL"
        else:
            situacao = "ABERTO"
        output.append(
            [
                grupo["filial_id"],
                grupo["cliente_id"],
                grupo["cliente_nome"],
                cnpj_format(grupo["cliente_cnpj_cpf"]),
                grupo["titulo_id"],
                grupo["unidade_saldo"],
                grupo["qtd"],
                grupo["valor"],
                recebido,
                aberto,
                grupo["valor"] - recebido - aberto,
                situacao,
            ]
        )
    return [headers, *output]


def style_sheet(ws, freeze: str = "A2") -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = freeze
    if ws.max_row >= 2:
        # AutoFiltro simples é mais compatível entre versões do Excel que
        # tabelas OOXML combinadas com formatação condicional.
        ws.auto_filter.ref = ws.dimensions

    # Atribuir estilo individualmente a centenas de milhares de células torna
    # o openpyxl muito lento. Nas abas menores mantemos a formatação completa;
    # na base detalhada, os valores permanecem numéricos e as datas continuam
    # sendo gravadas como datas reais, prontas para filtros e tabelas dinâmicas.
    if ws.max_row <= 1000:
        headers = {cell.value: cell.column for cell in ws[1]}
        for header, column in headers.items():
            letter = ws.cell(1, column).column_letter
            if header in MONEY_HEADERS or header == "Valor das parcelas":
                for cell in ws[letter][1:]:
                    cell.number_format = MONEY_FORMAT
            elif header in DECIMAL_HEADERS:
                for cell in ws[letter][1:]:
                    cell.number_format = DECIMAL_FORMAT
            elif header in DATE_HEADERS:
                for cell in ws[letter][1:]:
                    cell.number_format = DATE_FORMAT
            elif header in INTEGER_HEADERS:
                for cell in ws[letter][1:]:
                    cell.number_format = NUMBER_FORMAT

    # Medir uma amostra mantém a largura útil sem percorrer centenas de
    # milhares de células repetidas nas abas de visão.
    sample_last_row = min(ws.max_row, 250)
    for column_cells in ws.iter_cols(min_row=1, max_row=sample_last_row):
        width = max(
            10,
            min(48, max(len(str(cell.value or "")) for cell in column_cells) + 2),
        )
        ws.column_dimensions[column_cells[0].column_letter].width = width


def add_data_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[Any]]) -> None:
    print(f"[relatorio-receber] montando aba {name} ({len(rows)} linhas)...", flush=True)
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)

    header_map = {cell.value: cell.column_letter for cell in ws[1]}
    situation_col = header_map.get("Situação")
    difference_col = header_map.get("Diferença saldo local")
    if situation_col and ws.max_row > 1:
        ws.conditional_formatting.add(
            f"{situation_col}2:{situation_col}{ws.max_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("VENCIDA",{situation_col}2))'],
                fill=PatternFill("solid", fgColor=LIGHT_RED),
            ),
        )
        ws.conditional_formatting.add(
            f"{situation_col}2:{situation_col}{ws.max_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("CREDITO",{situation_col}2))'],
                fill=PatternFill("solid", fgColor=LIGHT_ORANGE),
            ),
        )
    if difference_col and ws.max_row > 1:
        ws.conditional_formatting.add(
            f"{difference_col}2:{difference_col}{ws.max_row}",
            FormulaRule(
                formula=[f"ABS({difference_col}2)>0.01"],
                fill=PatternFill("solid", fgColor=LIGHT_RED),
            ),
        )


def create_dashboard(
    wb: Workbook,
    totals: dict[str, Any],
    clients: list[dict],
    units: list[dict],
    rows: list[dict],
    filters: dict[str, Any],
    recebidas: list[dict] | None = None,
    fonte: str = "ActionAPI — /api/v1/financeiro/contas-receber",
) -> None:
    recebidas = recebidas or []
    ws = wb.active
    ws.title = "Painel"
    ws.merge_cells("A1:D1")
    ws["A1"] = "RELATÓRIO AVANÇADO DE CONTAS A RECEBER"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Gerado em"
    ws["B3"] = datetime.now()
    ws["B3"].number_format = "dd/mm/yyyy hh:mm"
    ws["A4"] = "Data-base do saldo"
    ws["B4"] = as_date(totals.get("data_calculo"))
    ws["B4"].number_format = DATE_FORMAT
    ws["A5"] = "Fonte"
    ws["B5"] = fonte
    ws["A6"] = "Filtros"
    ws["B6"] = ", ".join(
        f"{key}={value}" for key, value in filters.items() if value not in (None, "")
    ) or "Posição atual completa"

    indicators = [
        (
            "Valor em aberto",
            totals.get("valor_em_aberto"),
            True,
        ),
        ("Saldo convertido atual", totals.get("saldo_convertido_atual"), True),
        ("Saldo vencido convertido", totals.get("saldo_vencido_convertido"), True),
        ("Próximos 7 dias — convertido", totals.get("saldo_proximos_7_dias_convertido"), True),
        ("Próximos 30 dias — convertido", totals.get("saldo_proximos_30_dias_convertido"), True),
        ("Quantidade de clientes", totals.get("qtd_clientes"), False),
        ("Quantidade de títulos", totals.get("qtd_titulos"), False),
        ("Quantidade de parcelas", totals.get("qtd_parcelas"), False),
        ("Parcelas vencidas", totals.get("qtd_parcelas_vencidas"), False),
        ("Parcelas indexadas", totals.get("qtd_parcelas_indexadas"), False),
        ("Parcelas ligadas a FIDC", totals.get("qtd_parcelas_fidc"), False),
        ("Saldo ligado a FIDC — convertido", totals.get("saldo_fidc_convertido"), True),
    ]
    ws["A8"] = "INDICADORES"
    ws["B8"] = "VALOR"
    for cell in ws[8]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    for row_number, (label, value, monetary) in enumerate(indicators, 9):
        ws.cell(row_number, 1, label)
        ws.cell(row_number, 2, as_number(value))
        ws.cell(row_number, 2).number_format = MONEY_FORMAT if monetary else NUMBER_FORMAT

    top = sorted(
        clients,
        key=lambda item: as_number(item.get("saldo_convertido_atual")),
        reverse=True,
    )[:10]
    start = 9 + len(indicators) + 1
    ws.cell(start, 1, "10 MAIORES CLIENTES")
    ws.cell(start, 2, "Saldo convertido")
    ws.cell(start, 3, "% do total")
    for cell in ws[start]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    total_balance = as_number(totals.get("saldo_convertido_atual"))
    for offset, client in enumerate(top, 1):
        row_number = start + offset
        ws.cell(
            row_number,
            1,
            f"{client.get('cliente_nome')} — filial {client.get('filial_id')}",
        )
        balance = as_number(client.get("saldo_convertido_atual"))
        ws.cell(row_number, 2, balance).number_format = MONEY_FORMAT
        ws.cell(row_number, 3, balance / total_balance if total_balance else 0).number_format = "0.00%"

    chart = BarChart()
    chart.title = "Concentração por cliente"
    chart.y_axis.title = "Cliente"
    chart.x_axis.title = "Saldo convertido"
    chart.add_data(
        Reference(ws, min_col=2, min_row=start, max_row=start + len(top)),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(top))
    )
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "E8")

    ws["E25"] = "Unidade"
    ws["F25"] = "Parcelas"
    for offset, unit in enumerate(units, 1):
        ws.cell(25 + offset, 5, unit.get("unidade_saldo"))
        ws.cell(25 + offset, 6, as_number(unit.get("qtd_parcelas")))
    if units:
        pie = PieChart()
        pie.title = "Parcelas por unidade"
        pie.add_data(
            Reference(ws, min_col=6, min_row=25, max_row=25 + len(units)),
            titles_from_data=True,
        )
        pie.set_categories(
            Reference(ws, min_col=5, min_row=26, max_row=25 + len(units))
        )
        pie.height = 7
        pie.width = 9
        ws.add_chart(pie, "E28")

    combinado = list(rows) + list(recebidas)
    valor_parcelas = sum(valor_assinado(r) for r in combinado)
    recebido_acumulado = sum(as_number(r.get("valor_baixado")) for r in combinado)
    base = 35
    ws.cell(base, 1, "CONCILIAÇÃO DOS TÍTULOS ABERTOS")
    ws.cell(base, 2, "VALOR")
    for cell in ws[base]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    reconciliacao = [
        ("Valor das parcelas (documento)", valor_parcelas, True),
        ("Recebido acumulado", recebido_acumulado, True),
        (
            "Valor em aberto",
            as_number(totals.get("valor_em_aberto")),
            True,
        ),
        (
            "Saldo convertido pela cotação atual",
            as_number(totals.get("saldo_convertido_atual")),
            True,
        ),
        ("Parcelas em aberto", len(rows), False),
        ("Parcelas recebidas dos títulos abertos", len(recebidas), False),
    ]
    for offset, (label, value, monetary) in enumerate(reconciliacao, 1):
        ws.cell(base + offset, 1, label)
        ws.cell(base + offset, 2, as_number(value))
        ws.cell(base + offset, 2).number_format = MONEY_FORMAT if monetary else NUMBER_FORMAT

    # Comparativo por filial (em aberto, convertido para R$), para evitar
    # filtrar e somar à mão.
    filiais: dict[Any, dict[str, Any]] = {}
    for row in rows:
        fid = row.get("filial_id")
        item = filiais.get(fid)
        if item is None:
            item = {
                "filial_id": fid,
                "nome": row.get("filial_nome") or row.get("filial_identificacao") or "",
                "saldo": 0.0,
                "vencido": 0.0,
                "qtd": 0,
            }
            filiais[fid] = item
        conv = as_number(row.get("saldo_convertido_atual"))
        item["saldo"] += conv
        item["qtd"] += 1
        if row.get("situacao") == "VENCIDA":
            item["vencido"] += conv

    fbase = base + len(reconciliacao) + 2
    ws.cell(fbase, 1, "POR FILIAL (EM ABERTO — CONVERTIDO R$)")
    ws.cell(fbase, 2, "Saldo convertido")
    ws.cell(fbase, 3, "Saldo vencido convertido")
    ws.cell(fbase, 4, "% do total")
    for cell in ws[fbase]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    total_conv = as_number(totals.get("saldo_convertido_atual"))
    for offset, item in enumerate(
        sorted(filiais.values(), key=lambda x: x["saldo"], reverse=True), 1
    ):
        linha = fbase + offset
        nome = item["nome"] or item["filial_id"]
        ws.cell(linha, 1, f"{item['filial_id']} — {nome} ({item['qtd']} parc.)")
        ws.cell(linha, 2, item["saldo"]).number_format = MONEY_FORMAT
        ws.cell(linha, 3, item["vencido"]).number_format = MONEY_FORMAT
        ws.cell(
            linha, 4, item["saldo"] / total_conv if total_conv else 0
        ).number_format = "0.00%"

    ws["A33"] = "Observação"
    ws["B33"] = (
        "Saldo oficial preserva a unidade do título. Para contratos SJ$, US$ ou ER, "
        "o painel usa a conversão atual estimada apenas para consolidação em reais."
    )
    ws["B33"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 4


def create_methodology(
    wb: Workbook,
    data_base: str,
    fonte: str,
    recebimento_de: str,
    recebimento_ate: str,
) -> None:
    ws = wb.create_sheet("Metodologia")
    content = [
        ("RELATÓRIO DE CONTAS A RECEBER — METODOLOGIA", ""),
        ("Granularidade", "Uma linha por parcela em aberto de RECEBER."),
        ("Saldo oficial", "Snapshot de VALOR_ABERTO_RECEBER_DATA, função oficial do Oracle/SiAGRI."),
        ("Fonte do título", "CABREC ligado às parcelas de RECEBER."),
        ("Cliente", "CODI_TRA enriquecido pelo cadastro de transacionadores/clientes."),
        ("Filial", "Razão social do transacionador ligado a CADEMP.COD1_TRA; fantasia e identificação ficam separadas."),
        ("Recebimentos", "CRCBAIXA com SITU_BAI='N' no período informado da aba Recebimentos."),
        ("Estornos", "CRCBAIXA com SITU_BAI='E' não reduz o saldo."),
        ("Contratos indexados", "O saldo oficial permanece em SJ$, US$ ou ER; não é automaticamente um valor em reais."),
        ("Valor em aberto", "Títulos em R$ permanecem em reais; títulos indexados são convertidos pela cotação de origem e arredondados parcela a parcela."),
        ("Conversão atual", "Estimativa adicional: saldo em unidades multiplicado pela cotação mais recente replicada."),
        ("Saldo local", "Reprodução PostgreSQL da função oficial, incluindo indexador, data de cada baixa, agrupamentos e tolerância."),
        ("Abas de saldo e conciliação", "Aba 'Em Aberto' = parcelas com saldo. Aba 'Recebidas Títulos Abertos' = parcelas já quitadas, restritas aos títulos que ainda têm saldo em aberto. Aba 'Conciliação' resume por título: Valor das parcelas (assinado), Recebido acumulado e Em aberto."),
        ("Aba Recebimentos", "Movimento completo de recebimentos normais no período informado, independente de o título ainda estar em aberto."),
        ("Coluna Diferença (Conciliação)", "Diferença = Valor das parcelas − Recebido − Em aberto. Fica ~0 para títulos em R$; contratos indexados (SJ$, US$, ER) podem gerar diferença esperada, pois o saldo é mantido na unidade e o recebido em reais."),
        ("Sinal do saldo", "Documentos de natureza crédito (ex.: adiantamento de cliente) entram negativos e abatem o que há a receber."),
        ("Data-base desta planilha", data_base),
        ("Período de recebimentos", f"{recebimento_de} a {recebimento_ate}"),
        ("Fonte desta planilha", fonte),
        ("Atenção a lançamentos retroativos", "O SiAGRI permite registrar hoje uma baixa com data de pagamento passada. Por isso, reproduções históricas podem mudar após nova sincronização; a planilha informa explicitamente a data-base e a fonte usadas."),
        ("Compatibilidade Excel", "As abas usam AutoFiltro simples, sem tabelas OOXML sobrepostas."),
    ]
    for row in content:
        ws.append(row)
    ws.merge_cells("A1:B1")
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 115
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def validate_workbook(
    path: Path,
    expected_rows: int,
    expected_converted_balance: float,
) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Em Aberto"]
    row_count = ws.max_row - 1
    headers = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}
    balance_column = headers["Saldo convertido atual"]
    balance = sum(
        as_number(row[balance_column - 1].value)
        for row in ws.iter_rows(min_row=2)
    )
    error_cells = sum(
        1
        for sheet in wb.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "e"
    )
    wb.close()
    if (
        row_count != expected_rows
        or abs(balance - expected_converted_balance) > 0.01
        or error_cells
    ):
        raise RuntimeError(
            f"Validação do XLSX falhou: linhas {row_count}/{expected_rows}, "
            f"saldo convertido {balance:.2f}/{expected_converted_balance:.2f}, "
            f"células com erro {error_cells}"
        )


def generate_report(args: argparse.Namespace) -> Path:
    data_base = resolve_data_base(args) if getattr(args, "data_base", None) else None
    historico = bool(data_base) and data_base != date.today().isoformat()
    report_data_base = data_base or date.today().isoformat()
    recebimento_de, recebimento_ate = resolve_recebimento_period(args, report_data_base)

    if args.arquivo:
        output = Path(args.arquivo).resolve()
    elif data_base and data_base != date.today().isoformat():
        output = (
            ROOT
            / "relatorios"
            / f"contas-a-receber-python-base-{data_base}-gerado-{date.today().isoformat()}.xlsx"
        )
    else:
        output = ROOT / "relatorios" / f"contas-a-receber-python-{date.today().isoformat()}.xlsx"

    if historico:
        filters = {"dataBase": data_base}
        payload = fetch_historico_cr(data_base)
        rows = payload["rows"]
        for row in rows:
            row.setdefault("data_calculo", data_base)
        clients = payload["data"]
        units = payload["unidades"]
        totals = payload["totalizadores"]
        fonte = "PostgreSQL local — reprodução histórica do saldo SiAGRI"
        abertas, recebidas = split_aberto_recebidas(rows)
        print(
            f"[relatorio-receber] {len(abertas)} parcelas em aberto + {len(recebidas)} "
            f"recebidas dos títulos abertos (reprodução local em {data_base}, sem Oracle).",
            flush=True,
        )
    else:
        api_key = first_api_key()
        filters = {
            "vencimentoDe": args.vencimento_de,
            "vencimentoAte": args.vencimento_ate,
            "emissaoDe": args.emissao_de,
            "emissaoAte": args.emissao_ate,
            "filialId": args.filial_id,
            "clienteId": args.cliente_id,
            "tipoDocumento": args.tipo_documento,
            "situacao": args.situacao,
            "unidadeSaldo": args.unidade_saldo,
            "vendedorId": args.vendedor_id,
        }
        print(f"[relatorio-receber] consultando {args.api_url}...", flush=True)
        abertas = fetch_all(args.api_url, api_key, filters)
        summary = api_get(
            args.api_url,
            api_key,
            "/api/v1/financeiro/contas-receber/resumo",
            filters,
        )
        clients = summary.get("data", [])
        units = summary.get("unidades", [])
        totals = summary.get("totalizadores", {})
        fonte = "ActionAPI — snapshot oficial raw.duplicatas_saldo"
        print(f"[relatorio-receber] API retornou {len(abertas)} parcelas em aberto.", flush=True)
        # A API de CR só enxerga o snapshot em aberto (raw.duplicatas_saldo); as
        # parcelas recebidas dos títulos abertos vêm da reprodução local para
        # apoiar a conciliação.
        titulos_abertos = {row.get("titulo_id") for row in abertas}
        recebidas = fetch_recebidas_cr(date.today().isoformat(), titulos_abertos)
        print(
            f"[relatorio-receber] {len(recebidas)} parcelas recebidas dos títulos "
            "abertos anexadas.",
            flush=True,
        )

    normalize_report_rows(abertas, report_data_base)
    normalize_report_rows(recebidas, report_data_base)
    enrich_open_values(abertas, totals, units, clients)
    filters["recebimentoDe"] = recebimento_de
    filters["recebimentoAte"] = recebimento_ate
    recebimentos = fetch_recebimentos_cr(recebimento_de, recebimento_ate)

    wb = Workbook()
    create_dashboard(
        wb,
        totals,
        clients,
        units,
        abertas,
        filters,
        recebidas=recebidas,
        fonte=fonte,
    )
    headers = [header for _key, header in DETAIL_COLUMNS]
    add_data_sheet(wb, "Em Aberto", headers, detail_rows(abertas))
    add_data_sheet(wb, "Recebidas Títulos Abertos", headers, detail_rows(recebidas))
    recebimento_headers = [header for _key, header in RECEBIMENTO_COLUMNS]
    add_data_sheet(wb, "Recebimentos", recebimento_headers, recebimento_rows(recebimentos))
    conc = conciliacao_rows(abertas + recebidas)
    add_data_sheet(wb, "Conciliação", conc[0], conc[1:])

    clients_data = client_rows(clients)
    add_data_sheet(wb, "Por Cliente", clients_data[0], clients_data[1:])
    units_data = unit_rows(units)
    add_data_sheet(wb, "Por Unidade", units_data[0], units_data[1:])
    expiry_data = expiry_rows(abertas)
    add_data_sheet(wb, "Faixas Vencimento", expiry_data[0], expiry_data[1:])
    add_data_sheet(
        wb,
        "Indexados",
        headers,
        detail_rows([row for row in abertas if row.get("unidade_saldo") != "R$"]),
    )
    add_data_sheet(
        wb,
        "Divergencias Saldo",
        [header for _key, header in DIVERGENCE_COLUMNS],
        divergence_rows(abertas, report_data_base),
    )
    add_data_sheet(
        wb,
        "FIDC",
        headers,
        detail_rows([row for row in abertas if row.get("fidc")]),
    )
    create_methodology(wb, report_data_base, fonte, recebimento_de, recebimento_ate)

    output.parent.mkdir(parents=True, exist_ok=True)
    print("[relatorio-receber] gravando arquivo XLSX...", flush=True)
    wb.save(output)
    print("[relatorio-receber] validando arquivo XLSX...", flush=True)
    expected_balance = as_number(totals.get("saldo_convertido_atual"))
    validate_workbook(output, len(abertas), expected_balance)
    print(
        f"[relatorio-receber] {len(abertas)} em aberto / {len(recebidas)} "
        f"recebidas dos títulos abertos / {len(recebimentos)} recebimentos exportados"
    )
    print(f"[relatorio-receber] saldo convertido validado: R$ {expected_balance:,.2f}")
    print(f"[relatorio-receber] arquivo: {output}")
    return output


def main() -> None:
    args = parse_args()
    prompt_data_base_if_needed(args)
    data_base = resolve_data_base(args) if getattr(args, "data_base", None) else None
    if not data_base or data_base == date.today().isoformat():
        apply_period_selection(args)
    generate_report(args)


if __name__ == "__main__":
    main()
