#!/usr/bin/env python3
"""Gera o relatório avançado de Contas a Pagar consumindo a ActionAPI.

Exemplos:

    py scripts/gerar_relatorio_contas_pagar.py

    py scripts/gerar_relatorio_contas_pagar.py \
      --vencimento-de 2026-07-01 \
      --vencimento-ate 2026-12-31 \
      --arquivo relatorios/contas-pagar-segundo-semestre.xlsx

O script não acessa Oracle nem PostgreSQL. A API key é lida de API_KEYS no
.env e nunca é gravada na planilha.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:
    from gerar_relatorios_executivos import (
        add_data_base_argument,
        add_period_arguments,
        has_period_argument,
        parse_user_date,
        prompt_data_base_if_needed,
        prompt_period_if_needed,
        resolve_data_base,
        resolve_period,
    )
except ModuleNotFoundError:
    from scripts.gerar_relatorios_executivos import (
        add_data_base_argument,
        add_period_arguments,
        has_period_argument,
        parse_user_date,
        prompt_data_base_if_needed,
        prompt_period_if_needed,
        resolve_data_base,
        resolve_period,
    )

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ModuleNotFoundError:
    print(
        "Dependência ausente. Execute:\n"
        "  py -m pip install -r scripts/requirements-relatorio-contas-pagar.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_API_URL = "http://127.0.0.1:3000"
PAGE_SIZE = 10_000

BLUE = "17365D"
LIGHT_BLUE = "D9EAF7"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
MONEY_FORMAT = 'R$ #,##0.00;[Red]-R$ #,##0.00'
NUMBER_FORMAT = "#,##0"
DATE_FORMAT = "dd/mm/yyyy"


DETAIL_COLUMNS = [
    ("filial_id", "Filial"),
    ("filial_nome", "Nome da filial"),
    ("filial_fantasia", "Fantasia da filial"),
    ("filial_identificacao", "Identificação da filial"),
    ("fornecedor_id", "Código fornecedor"),
    ("fornecedor_nome", "Fornecedor"),
    ("fornecedor_cnpj_cpf", "CPF/CNPJ"),
    ("titulo_id", "Título"),
    ("parcela_nr", "Parcela"),
    ("numero_documento", "Número da NF/documento"),
    ("nf_entrada_ids", "Controles internos NF"),
    ("tipo_documento", "Tipo documento"),
    ("tipo_documento_descricao", "Descrição tipo documento"),
    ("fidc", "FIDC?"),
    ("historico", "Histórico/referência"),
    ("data_emissao", "Data emissão"),
    ("data_vencimento", "Data vencimento"),
    ("valor_titulo", "Valor do título"),
    ("valor_parcela", "Valor da parcela"),
    ("valor_baixado", "Valor baixado"),
    ("saldo_parcela", "Saldo da parcela"),
    ("unidade_saldo", "Unidade do saldo"),
    ("indexador_id", "Código indexador"),
    ("valor_indexador_origem", "Cotação de origem"),
    ("valor_indexador_atual", "Cotação atual"),
    ("saldo_convertido_atual", "Saldo convertido atual"),
    ("situacao", "Situação"),
    ("dias_atraso", "Dias em atraso"),
    ("faixa_vencimento", "Faixa de vencimento"),
    ("qtd_pedidos", "Quantidade de pedidos"),
    ("pedidos_numeros", "Pedidos internos SiAGRI"),
    ("pedidos_fornecedor_numeros", "Pedidos do fornecedor"),
    ("fornecedores_pedido_ids", "Código fornecedor do pedido"),
    ("fornecedores_pedido_nomes", "Fornecedor do pedido"),
    ("fornecedores_pedido_cnpjs", "CPF/CNPJ fornecedor do pedido"),
    ("filiais_pedido_ids", "Filial do pedido"),
    ("pedidos_ids", "Chaves internas dos pedidos"),
    ("produtos_ids", "Produtos"),
    ("produtos_descricoes", "Descrição dos produtos"),
    ("primeira_data_pedido", "Primeira data pedido"),
    ("ultima_data_pedido", "Última data pedido"),
    ("status_vinculo_pedido", "Status vínculo pedido"),
    ("conferencia_pedido", "Conferência pedido"),
    ("divergencia_resumo", "Resumo da divergência"),
    ("divergencia_detalhe", "Detalhe da divergência"),
    ("juros", "Juros"),
    ("multa", "Multa"),
    ("desconto", "Desconto"),
    ("acrescimo", "Acréscimo"),
    ("primeira_baixa", "Primeira baixa"),
    ("ultima_baixa", "Última baixa"),
]

MONEY_HEADERS = {
    "Valor do título",
    "Valor da parcela",
    "Valor baixado",
    "Saldo da parcela",
    "Saldo convertido atual",
    "Juros",
    "Multa",
    "Desconto",
    "Acréscimo",
    "Valor das parcelas",
    "Saldo",
    "Saldo vencido",
    "Saldo próximos 7 dias",
    "Saldo próximos 30 dias",
    "Pago acumulado",
    "Em aberto",
    "Diferença",
}

DATE_HEADERS = {
    "Data emissão",
    "Data vencimento",
    "Primeira data pedido",
    "Última data pedido",
    "Primeira baixa",
    "Última baixa",
    "Primeiro vencimento",
    "Último vencimento",
}

INTEGER_HEADERS = {
    "Parcela",
    "Dias em atraso",
    "Quantidade de pedidos",
    "Quantidade de parcelas",
    "Quantidade de títulos",
    "Parcelas com pedido",
    "Parcelas sem pedido",
    "Divergências de pedido",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera relatório Excel avançado de Contas a Pagar via ActionAPI."
    )
    parser.add_argument("--api-url", default=os.getenv("ACTIONAPI_URL", DEFAULT_API_URL))
    parser.add_argument("--arquivo", help="Caminho do arquivo .xlsx de saída.")
    parser.add_argument("--vencimento-de", help="Vencimento inicial, AAAA-MM-DD.")
    parser.add_argument("--vencimento-ate", help="Vencimento final, AAAA-MM-DD.")
    parser.add_argument("--emissao-de", help="Emissão inicial, AAAA-MM-DD.")
    parser.add_argument("--emissao-ate", help="Emissão final, AAAA-MM-DD.")
    parser.add_argument("--filial-id", help="Código da filial.")
    parser.add_argument("--fornecedor-id", help="Código do fornecedor.")
    add_period_arguments(parser)
    add_data_base_argument(parser)
    parser.add_argument("--incluir-baixados", action="store_true")
    return parser.parse_args()


def apply_period_selection(args: argparse.Namespace) -> None:
    explicit_dates = any(
        (args.vencimento_de, args.vencimento_ate, args.emissao_de, args.emissao_ate)
    )
    if not has_period_argument(args) and not explicit_dates and sys.stdin.isatty():
        print("O período escolhido será aplicado à data de vencimento.")
    prompt_period_if_needed(args, additional_period_supplied=explicit_dates)
    if has_period_argument(args):
        if explicit_dates:
            raise SystemExit(
                "erro: use o período do menu/parâmetros gerais ou os filtros "
                "específicos de vencimento/emissão, não ambos."
            )
        try:
            args.vencimento_de, args.vencimento_ate, _slug = resolve_period(args)
        except ValueError as exc:
            raise SystemExit(f"erro: {exc}") from exc
        return

    for field in ("vencimento_de", "vencimento_ate", "emissao_de", "emissao_ate"):
        value = getattr(args, field)
        if value:
            try:
                parsed = parse_user_date(value, f"--{field.replace('_', '-')}")
            except ValueError as exc:
                raise SystemExit(f"erro: {exc}") from exc
            setattr(args, field, parsed.isoformat())


def read_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        value = value.strip()
        if (
            len(value) >= 2
            and value[0] == value[-1]
            and value.startswith(("'", '"'))
        ):
            value = value[1:-1]
        values[key.strip()] = value
    return values


def first_api_key() -> str:
    env = {**read_env(ROOT / ".env"), **os.environ}
    keys = [item.strip() for item in env.get("API_KEYS", "").split(",") if item.strip()]
    if not keys:
        raise RuntimeError("API_KEYS não está configurada no .env.")
    return keys[0]


def api_get(base_url: str, api_key: str, endpoint: str, params: dict[str, Any]) -> dict:
    clean = {key: value for key, value in params.items() if value not in (None, "")}
    url = f"{base_url.rstrip('/')}{endpoint}"
    if clean:
        url += "?" + urlencode(clean)
    request = Request(url, headers={"X-API-Key": api_key, "Accept": "application/json"})
    try:
        with urlopen(request, timeout=120) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"ActionAPI respondeu HTTP {exc.code}: {body}") from exc
    except URLError as exc:
        raise RuntimeError(f"Não foi possível acessar a ActionAPI em {url}: {exc}") from exc


NODE_SALDO_HISTORICO = ROOT / "packages" / "etl" / "src" / "scripts" / "saldo-aberto-historico.js"


def fetch_historico_cp(data_base: str) -> dict:
    """Reproduz o saldo em aberto de CP numa data-base passada/futura via PostgreSQL.

    Não usa a ActionAPI nem o Oracle: reaproveita, parametrizada, a mesma
    fórmula validada contra VALOR_ABERTO_PAGAR_DATA (zero divergências),
    incluindo a inversão de sinal para documentos de natureza crédito
    (ex.: adiantamento a fornecedor) — ver Metodologia no relatório gerado.
    """
    print(
        f"[relatorio-python] reproduzindo saldo em {data_base} via PostgreSQL "
        "(sem Oracle, fórmula validada)...",
        flush=True,
    )
    result = subprocess.run(
        [
            "node",
            str(NODE_SALDO_HISTORICO),
            "--tipo",
            "CP",
            "--data-base",
            data_base,
            "--incluir-baixadas",
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        cwd=str(NODE_SALDO_HISTORICO.parents[2]),
    )
    if result.returncode != 0:
        raise RuntimeError(f"Falha ao calcular saldo histórico de CP: {result.stderr}")
    payload = json.loads(result.stdout)
    return payload["cp"]


def fetch_all(base_url: str, api_key: str, filters: dict[str, Any]) -> list[dict]:
    rows: list[dict] = []
    page = 1
    while True:
        payload = api_get(
            base_url,
            api_key,
            "/api/v1/financeiro/contas-pagar",
            {**filters, "page": page, "pageSize": PAGE_SIZE},
        )
        current = payload.get("data", [])
        rows.extend(current)
        if not current or len(rows) >= int(payload.get("total", len(rows))):
            return rows
        page += 1


def as_number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def as_date(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def cnpj_format(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 14:
        return (
            f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/"
            f"{digits[8:12]}-{digits[12:]}"
        )
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return str(value or "")


def converted_value(key: str, value: Any) -> Any:
    if key in {
        "valor_titulo",
        "valor_parcela",
        "valor_baixado",
        "saldo_parcela",
        "valor_indexador_origem",
        "valor_indexador_atual",
        "saldo_convertido_atual",
        "juros",
        "multa",
        "desconto",
        "acrescimo",
    }:
        return as_number(value)
    if key in {
        "data_emissao",
        "data_vencimento",
        "primeira_data_pedido",
        "ultima_data_pedido",
        "primeira_baixa",
        "ultima_baixa",
    }:
        return as_date(value)
    if key in {"parcela_nr", "dias_atraso", "qtd_pedidos"}:
        return as_number(value)
    if key == "fornecedor_cnpj_cpf":
        return cnpj_format(value)
    return value


def detail_rows(rows: list[dict]) -> list[list[Any]]:
    return [
        [converted_value(key, row.get(key)) for key, _header in DETAIL_COLUMNS]
        for row in rows
    ]


def supplier_rows(rows: list[dict]) -> list[list[Any]]:
    headers = [
        "Filial",
        "Nome da filial",
        "Código fornecedor",
        "Fornecedor",
        "CPF/CNPJ",
        "Quantidade de parcelas",
        "Quantidade de títulos",
        "Valor das parcelas",
        "Valor baixado",
        "Saldo",
        "Saldo vencido",
        "Saldo próximos 7 dias",
        "Saldo próximos 30 dias",
        "Parcelas com pedido",
        "Parcelas sem pedido",
        "Divergências de pedido",
        "Primeiro vencimento",
        "Último vencimento",
    ]
    output = []
    for row in sorted(rows, key=lambda item: as_number(item.get("saldo")), reverse=True):
        output.append(
            [
                row.get("filial_id"),
                row.get("filial_nome"),
                row.get("fornecedor_id"),
                row.get("fornecedor_nome"),
                cnpj_format(row.get("fornecedor_cnpj_cpf")),
                as_number(row.get("qtd_parcelas")),
                as_number(row.get("qtd_titulos")),
                as_number(row.get("valor_parcelas")),
                as_number(row.get("valor_baixado")),
                as_number(row.get("saldo")),
                as_number(row.get("saldo_vencido")),
                as_number(row.get("saldo_proximos_7_dias")),
                as_number(row.get("saldo_proximos_30_dias")),
                as_number(row.get("qtd_parcelas_com_pedido")),
                as_number(row.get("qtd_parcelas_sem_pedido")),
                as_number(row.get("qtd_divergencias_pedido")),
                as_date(row.get("primeiro_vencimento")),
                as_date(row.get("ultimo_vencimento")),
            ]
        )
    return [headers, *output]


def expiry_rows(rows: list[dict]) -> list[list[Any]]:
    order = [
        "VENCIDO_ACIMA_90_DIAS",
        "VENCIDO_31_A_90_DIAS",
        "VENCIDO_1_A_30_DIAS",
        "VENCE_HOJE",
        "VENCE_EM_1_A_7_DIAS",
        "VENCE_EM_8_A_30_DIAS",
        "VENCE_EM_31_A_60_DIAS",
        "VENCE_EM_61_A_90_DIAS",
        "VENCE_ACIMA_90_DIAS",
    ]
    grouped: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"parcelas": 0, "titulos": set(), "saldo": 0.0}
    )
    for row in rows:
        group = grouped[row.get("faixa_vencimento") or "SEM_CLASSIFICACAO"]
        group["parcelas"] += 1
        group["titulos"].add(row.get("titulo_id"))
        group["saldo"] += as_number(row.get("saldo_parcela"))
    result = [["Faixa de vencimento", "Quantidade de parcelas", "Quantidade de títulos", "Saldo"]]
    for key in sorted(grouped, key=lambda item: order.index(item) if item in order else 99):
        result.append(
            [key, grouped[key]["parcelas"], len(grouped[key]["titulos"]), grouped[key]["saldo"]]
        )
    return result


def split_aberto_pagas(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Separa parcelas em aberto (saldo ≠ 0) das já quitadas (saldo ≈ 0)."""
    abertas: list[dict] = []
    pagas: list[dict] = []
    for row in rows:
        if abs(as_number(row.get("saldo_parcela"))) > 0.01:
            abertas.append(row)
        else:
            pagas.append(row)
    return abertas, pagas


def valor_assinado(row: dict) -> float:
    """Valor da parcela com o mesmo sinal do saldo (crédito = negativo).

    Mantém a coluna "Valor das parcelas" comparável ao "Valor documento" do
    controller e faz a conciliação (Valor − Pago − Em aberto) fechar para
    documentos de natureza débito e crédito não indexados.
    """
    valor = as_number(row.get("valor_parcela"))
    return -valor if row.get("natureza_tipo_documento") == "C" else valor


def conciliacao_rows(rows: list[dict]) -> list[list[Any]]:
    """Uma linha por título: Valor das parcelas, Pago acumulado e Em aberto.

    A coluna "Diferença" (= Valor − Pago − Em aberto) fica ~0 para títulos de
    natureza débito e crédito em R$. Indexadores e agrupamentos podem gerar
    diferença — é esperado e está descrito na Metodologia.
    """
    headers = [
        "Filial",
        "Nome da filial",
        "Código fornecedor",
        "Fornecedor",
        "CPF/CNPJ",
        "Título",
        "Quantidade de parcelas",
        "Valor das parcelas",
        "Pago acumulado",
        "Em aberto",
        "Diferença",
        "Situação",
    ]
    grupos: dict[Any, dict[str, Any]] = {}
    for row in rows:
        titulo = row.get("titulo_id")
        chave = (row.get("filial_id"), row.get("fornecedor_id"), titulo)
        grupo = grupos.get(chave)
        if grupo is None:
            grupo = {
                "filial_id": row.get("filial_id"),
                "filial_nome": row.get("filial_nome"),
                "fornecedor_id": row.get("fornecedor_id"),
                "fornecedor_nome": row.get("fornecedor_nome"),
                "fornecedor_cnpj_cpf": row.get("fornecedor_cnpj_cpf"),
                "titulo_id": titulo,
                "qtd": 0,
                "valor": 0.0,
                "pago": 0.0,
                "aberto": 0.0,
            }
            grupos[chave] = grupo
        grupo["qtd"] += 1
        grupo["valor"] += valor_assinado(row)
        grupo["pago"] += as_number(row.get("valor_baixado"))
        grupo["aberto"] += as_number(row.get("saldo_parcela"))

    output: list[list[Any]] = []
    for grupo in sorted(grupos.values(), key=lambda item: item["aberto"], reverse=True):
        aberto = grupo["aberto"]
        pago = grupo["pago"]
        if abs(aberto) <= 0.01:
            situacao = "BAIXADO"
        elif pago > 0.01:
            situacao = "PARCIAL"
        else:
            situacao = "ABERTO"
        output.append(
            [
                grupo["filial_id"],
                grupo["filial_nome"],
                grupo["fornecedor_id"],
                grupo["fornecedor_nome"],
                cnpj_format(grupo["fornecedor_cnpj_cpf"]),
                grupo["titulo_id"],
                grupo["qtd"],
                grupo["valor"],
                pago,
                aberto,
                grupo["valor"] - pago - aberto,
                situacao,
            ]
        )
    return [headers, *output]


def safe_table_name(name: str) -> str:
    return "Tabela" + re.sub(r"[^A-Za-z0-9]", "", name)


def style_sheet(ws, table: bool = True, freeze: str = "A2") -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = freeze

    headers = {cell.value: cell.column for cell in ws[1]}
    for header, column in headers.items():
        letter = ws.cell(1, column).column_letter
        if header in MONEY_HEADERS:
            for cell in ws[letter][1:]:
                cell.number_format = MONEY_FORMAT
        elif header in DATE_HEADERS:
            for cell in ws[letter][1:]:
                cell.number_format = DATE_FORMAT
        elif header in INTEGER_HEADERS:
            for cell in ws[letter][1:]:
                cell.number_format = NUMBER_FORMAT

    for column_cells in ws.columns:
        width = max(
            10,
            min(
                48,
                max(len(str(cell.value or "")) for cell in column_cells) + 2,
            ),
        )
        ws.column_dimensions[column_cells[0].column_letter].width = width

    if table and ws.max_row >= 2:
        table_name = safe_table_name(ws.title)
        excel_table = Table(displayName=table_name, ref=ws.dimensions)
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(excel_table)
    elif ws.max_row >= 2:
        # Uma tabela do Excel já possui seu próprio AutoFilter. Gravar também
        # um AutoFilter na planilha sobre o mesmo intervalo faz algumas versões
        # do Excel exibirem uma mensagem de reparo ao abrir o arquivo.
        ws.auto_filter.ref = ws.dimensions


def add_data_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)

    header_map = {cell.value: cell.column_letter for cell in ws[1]}
    status_col = header_map.get("Conferência pedido")
    if status_col and ws.max_row > 1:
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{ws.max_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("DIVERGENTE",{status_col}2))'],
                fill=PatternFill("solid", fgColor=LIGHT_RED),
            ),
        )
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{ws.max_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("MESMA_RAIZ",{status_col}2))'],
                fill=PatternFill("solid", fgColor=LIGHT_ORANGE),
            ),
        )


def create_dashboard(
    wb: Workbook,
    totals: dict[str, Any],
    suppliers: list[dict],
    rows: list[dict],
    filters: dict[str, Any],
    pagas: list[dict] | None = None,
) -> None:
    pagas = pagas or []
    ws = wb.active
    ws.title = "Painel"
    ws.merge_cells("A1:D1")
    ws["A1"] = "RELATÓRIO AVANÇADO DE CONTAS A PAGAR"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Gerado em"
    ws["B3"] = datetime.now()
    ws["B3"].number_format = "dd/mm/yyyy hh:mm"
    ws["A4"] = "Fonte"
    ws["B4"] = "ActionAPI — /api/v1/financeiro/contas-pagar"
    ws["A5"] = "Filtros"
    ws["B5"] = ", ".join(f"{k}={v}" for k, v in filters.items() if v not in (None, "")) or "Posição atual completa"

    indicators = [
        ("Saldo total em aberto", totals.get("saldo"), True),
        ("Saldo vencido", totals.get("saldo_vencido"), True),
        ("Saldo próximos 7 dias", totals.get("saldo_proximos_7_dias"), True),
        ("Saldo próximos 30 dias", totals.get("saldo_proximos_30_dias"), True),
        ("Quantidade de fornecedores", totals.get("qtd_fornecedores"), False),
        ("Quantidade de títulos", totals.get("qtd_titulos"), False),
        ("Quantidade de parcelas", totals.get("qtd_parcelas"), False),
        ("Parcelas com pedido", totals.get("qtd_parcelas_com_pedido"), False),
        ("Parcelas sem pedido", totals.get("qtd_parcelas_sem_pedido"), False),
        ("Divergências identificadas", totals.get("qtd_divergencias_pedido"), False),
        ("Parcelas ligadas a FIDC", totals.get("qtd_parcelas_fidc"), False),
        ("Saldo ligado a FIDC", totals.get("saldo_fidc"), True),
    ]
    ws["A7"] = "INDICADORES"
    ws["B7"] = "VALOR"
    for cell in ws[7]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    for row_number, (label, value, monetary) in enumerate(indicators, 8):
        ws.cell(row_number, 1, label)
        ws.cell(row_number, 2, as_number(value))
        ws.cell(row_number, 2).number_format = MONEY_FORMAT if monetary else NUMBER_FORMAT

    top = sorted(suppliers, key=lambda item: as_number(item.get("saldo")), reverse=True)[:10]
    start = 20
    ws.cell(start, 1, "10 MAIORES FORNECEDORES")
    ws.cell(start, 2, "Saldo")
    ws.cell(start, 3, "% do total")
    for cell in ws[start]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    total_balance = as_number(totals.get("saldo"))
    for offset, supplier in enumerate(top, 1):
        row_number = start + offset
        ws.cell(
            row_number,
            1,
            f"{supplier.get('fornecedor_nome')} — filial {supplier.get('filial_id')}",
        )
        balance = as_number(supplier.get("saldo"))
        ws.cell(row_number, 2, balance).number_format = MONEY_FORMAT
        ws.cell(row_number, 3, balance / total_balance if total_balance else 0).number_format = "0.00%"

    chart = BarChart()
    chart.title = "Concentração por fornecedor"
    chart.y_axis.title = "Fornecedor"
    chart.x_axis.title = "Saldo"
    chart.add_data(Reference(ws, min_col=2, min_row=start, max_row=start + len(top)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(top)))
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "E7")

    with_order = sum(1 for row in rows if row.get("status_vinculo_pedido") == "COM_PEDIDO")
    without_order = len(rows) - with_order
    ws["E24"] = "Vínculo"
    ws["F24"] = "Quantidade"
    ws["E25"] = "Com pedido"
    ws["F25"] = with_order
    ws["E26"] = "Sem pedido"
    ws["F26"] = without_order
    pie = PieChart()
    pie.title = "Parcelas com/sem pedido"
    pie.add_data(Reference(ws, min_col=6, min_row=24, max_row=26), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=5, min_row=25, max_row=26))
    pie.height = 7
    pie.width = 9
    ws.add_chart(pie, "E27")

    combinado = list(rows) + list(pagas)
    valor_parcelas = sum(valor_assinado(r) for r in combinado)
    pago_acumulado = sum(as_number(r.get("valor_baixado")) for r in combinado)
    saldo_aberto = as_number(totals.get("saldo"))
    base = 32
    ws.cell(base, 1, "CONCILIAÇÃO A PAGAR × PAGAS")
    ws.cell(base, 2, "VALOR")
    for cell in ws[base]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    reconciliacao = [
        ("Valor das parcelas (documento)", valor_parcelas, True),
        ("Pago acumulado", pago_acumulado, True),
        ("Saldo em aberto", saldo_aberto, True),
        ("Parcelas em aberto", len(rows), False),
        ("Parcelas pagas (de títulos abertos)", len(pagas), False),
    ]
    for offset, (label, value, monetary) in enumerate(reconciliacao, 1):
        ws.cell(base + offset, 1, label)
        ws.cell(base + offset, 2, as_number(value))
        ws.cell(base + offset, 2).number_format = MONEY_FORMAT if monetary else NUMBER_FORMAT

    # Comparativo por filial (em aberto), para evitar filtrar e somar à mão.
    filiais: dict[Any, dict[str, Any]] = {}
    for row in rows:
        fid = row.get("filial_id")
        item = filiais.get(fid)
        if item is None:
            item = {
                "filial_id": fid,
                "nome": row.get("filial_nome") or row.get("filial_fantasia") or "",
                "saldo": 0.0,
                "vencido": 0.0,
                "qtd": 0,
            }
            filiais[fid] = item
        saldo = as_number(row.get("saldo_parcela"))
        item["saldo"] += saldo
        item["qtd"] += 1
        if as_number(row.get("dias_atraso")) > 0 and saldo > 0:
            item["vencido"] += saldo

    fbase = base + len(reconciliacao) + 2
    ws.cell(fbase, 1, "POR FILIAL (EM ABERTO)")
    ws.cell(fbase, 2, "Saldo em aberto")
    ws.cell(fbase, 3, "Saldo vencido")
    ws.cell(fbase, 4, "% do total")
    for cell in ws[fbase]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    total_saldo = as_number(totals.get("saldo"))
    for offset, item in enumerate(
        sorted(filiais.values(), key=lambda x: x["saldo"], reverse=True), 1
    ):
        linha = fbase + offset
        nome = item["nome"] or item["filial_id"]
        ws.cell(linha, 1, f"{item['filial_id']} — {nome} ({item['qtd']} parc.)")
        ws.cell(linha, 2, item["saldo"]).number_format = MONEY_FORMAT
        ws.cell(linha, 3, item["vencido"]).number_format = MONEY_FORMAT
        ws.cell(
            linha, 4, item["saldo"] / total_saldo if total_saldo else 0
        ).number_format = "0.00%"

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 4


def create_methodology(wb: Workbook) -> None:
    ws = wb.create_sheet("Metodologia")
    content = [
        ("RELATÓRIO DE CONTAS A PAGAR — METODOLOGIA", ""),
        ("Granularidade", "Uma linha por parcela financeira de PAGAR."),
        ("Título", "CABPAGAR ligado às parcelas de PAGAR."),
        ("NF", "Número visível vem de NFENTRA.NUME_NFE; controle interno vem de CTRL_NFE."),
        ("Pedido interno", "INFENTRA.EMPR_PEC + NUME_PEC → PEDCOM.CODI_EMP + NUME_PEC."),
        ("Pedido do fornecedor", "PEDCOM.NUFO_PEC. Não confundir com CTRL_NFE."),
        ("Produtos", "Itens de INFENTRA ligados ao título e ao pedido."),
        ("Saldo", "Reprodução local de VALOR_ABERTO_PAGAR_DATA."),
        ("Sinal do saldo", "Documentos de natureza CRÉDITO (adiantamento a fornecedor, nota de crédito, devolução de compra) entram NEGATIVOS e abatem o que se deve, exatamente como no relatório do controller. Documentos de natureza débito entram positivos."),
        ("Abas A pagar × Pagas", "Aba 'Em Aberto' = parcelas com saldo (= relatório do controller). Aba 'Pagas' = parcelas já quitadas (saldo ≈ 0) com baixa até a data-base, restritas aos títulos que ainda têm saldo em aberto. Aba 'Conciliação' resume por título: Valor das parcelas, Pago acumulado e Em aberto."),
        ("Coluna Diferença (Conciliação)", "Diferença = Valor das parcelas − Pago acumulado − Em aberto. Fica ~0 para títulos de natureza débito em R$; indexadores, natureza crédito e agrupamentos podem gerar diferença esperada."),
        ("Baixas", "Somente CPGBAIXA normal e com data até a posição calculada. 'Pago acumulado' é o valor já baixado da parcela até a data-base."),
        ("Indexadores", "Saldo pode estar em R$, SJ$, US$ ou ER."),
        ("MESMA_RAIZ_CNPJ_ESTABELECIMENTO_DIFERENTE", "Divergência de cadastro: título/NF e pedido usam códigos e CNPJs completos diferentes, embora pertençam à mesma raiz de CNPJ."),
        ("FORNECEDOR_DIVERGENTE", "Fornecedor do pedido possui raiz de CNPJ diferente do título/NF."),
        ("Filial", "Razão social obtida pelo transacionador ligado a CADEMP.COD1_TRA; fantasia e identificação são exibidas separadamente."),
        ("Validação", "Cálculo CP comparado nas 183.656 parcelas com a função Oracle (zero divergências em 20/06/2026) e conferido contra o relatório do controller (CPG) de 21/06/2026: Valor em Aberto R$ 105.175.831,96 vs R$ 105.175.831,97 reproduzido (diferença de R$ 0,01 por arredondamento)."),
    ]
    for row in content:
        ws.append(row)
    ws.merge_cells("A1:B1")
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 115
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def validate_workbook(path: Path, expected_rows: int, expected_balance: float) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Em Aberto"]
    row_count = ws.max_row - 1
    headers = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}
    balance_column = headers["Saldo da parcela"]
    balance = sum(
        as_number(ws.cell(row, balance_column).value)
        for row in range(2, ws.max_row + 1)
    )
    wb.close()
    if row_count != expected_rows or abs(balance - expected_balance) > 0.01:
        raise RuntimeError(
            f"Validação do XLSX falhou: linhas {row_count}/{expected_rows}, "
            f"saldo {balance:.2f}/{expected_balance:.2f}"
        )


def generate_report(args: argparse.Namespace) -> Path:
    data_base = resolve_data_base(args) if getattr(args, "data_base", None) else None
    historico = bool(data_base) and data_base != date.today().isoformat()

    output = (
        Path(args.arquivo).resolve()
        if args.arquivo
        else ROOT / "relatorios" / f"contas-a-pagar-python-{date.today().isoformat()}.xlsx"
    )

    if historico:
        filters = {"dataBase": data_base}
        payload = fetch_historico_cp(data_base)
        rows = payload["rows"]
        suppliers = payload["data"]
        totals = payload["totalizadores"]
        abertas, pagas = split_aberto_pagas(rows)
        print(
            f"[relatorio-python] {len(abertas)} parcelas em aberto + {len(pagas)} pagas "
            f"(reprodução local em {data_base}, sem Oracle).",
        )
        wb = Workbook()
        create_dashboard(wb, totals, suppliers, abertas, filters, pagas=pagas)
        headers = [header for _key, header in DETAIL_COLUMNS]
        add_data_sheet(wb, "Em Aberto", headers, detail_rows(abertas))
        add_data_sheet(wb, "Pagas", headers, detail_rows(pagas))
        conc = conciliacao_rows(rows)
        add_data_sheet(wb, "Conciliação", conc[0], conc[1:])
        supplier_data = supplier_rows(suppliers)
        add_data_sheet(wb, "Por Fornecedor", supplier_data[0], supplier_data[1:])
        expiry_data = expiry_rows(abertas)
        add_data_sheet(wb, "Faixas Vencimento", expiry_data[0], expiry_data[1:])
        add_data_sheet(
            wb,
            "Vencidos",
            headers,
            detail_rows([row for row in abertas if as_number(row.get("dias_atraso")) > 0]),
        )
        add_data_sheet(
            wb,
            "FIDC",
            headers,
            detail_rows([row for row in abertas if row.get("fidc")]),
        )
        create_methodology(wb)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        expected_balance = as_number(totals.get("saldo"))
        validate_workbook(output, len(abertas), expected_balance)
        print(
            f"[relatorio-python] {len(abertas)} em aberto / {len(pagas)} pagas exportadas"
        )
        print(f"[relatorio-python] saldo validado: R$ {expected_balance:,.2f}")
        print(f"[relatorio-python] arquivo: {output}")
        return output

    api_key = first_api_key()
    base_filters = {
        "vencimentoDe": args.vencimento_de,
        "vencimentoAte": args.vencimento_ate,
        "emissaoDe": args.emissao_de,
        "emissaoAte": args.emissao_ate,
        "filialId": args.filial_id,
        "fornecedorId": args.fornecedor_id,
    }
    # Lista unificada: em aberto + pagas de títulos que ainda têm saldo (limitada
    # no servidor, sem despejar todo o histórico de baixas).
    list_filters = {**base_filters, "incluirPagasDeAbertos": "true"}
    # Totalizadores e quebra por fornecedor refletem apenas o em aberto (= controller).
    resumo_filters = {**base_filters, "somenteEmAberto": "true"}

    print(f"[relatorio-python] consultando {args.api_url}...")
    rows = fetch_all(args.api_url, api_key, list_filters)
    summary = api_get(
        args.api_url,
        api_key,
        "/api/v1/financeiro/contas-pagar/resumo",
        resumo_filters,
    )
    suppliers = summary.get("data", [])
    totals = summary.get("totalizadores", {})
    abertas, pagas = split_aberto_pagas(rows)

    wb = Workbook()
    create_dashboard(wb, totals, suppliers, abertas, list_filters, pagas=pagas)
    headers = [header for _key, header in DETAIL_COLUMNS]
    add_data_sheet(wb, "Em Aberto", headers, detail_rows(abertas))
    add_data_sheet(wb, "Pagas", headers, detail_rows(pagas))
    conc = conciliacao_rows(rows)
    add_data_sheet(wb, "Conciliação", conc[0], conc[1:])
    supplier_data = supplier_rows(suppliers)
    add_data_sheet(wb, "Por Fornecedor", supplier_data[0], supplier_data[1:])
    expiry_data = expiry_rows(abertas)
    add_data_sheet(wb, "Faixas Vencimento", expiry_data[0], expiry_data[1:])
    add_data_sheet(
        wb,
        "Vencidos",
        headers,
        detail_rows([row for row in abertas if as_number(row.get("dias_atraso")) > 0]),
    )
    add_data_sheet(
        wb,
        "Com Pedido",
        headers,
        detail_rows([row for row in abertas if row.get("status_vinculo_pedido") == "COM_PEDIDO"]),
    )
    add_data_sheet(
        wb,
        "Sem Pedido",
        headers,
        detail_rows([row for row in abertas if row.get("status_vinculo_pedido") != "COM_PEDIDO"]),
    )
    add_data_sheet(
        wb,
        "Divergencias",
        headers,
        detail_rows(
            [
                row
                for row in abertas
                if row.get("conferencia_pedido")
                not in {
                    "OK",
                    "NAO_APLICAVEL",
                }
            ]
        ),
    )
    add_data_sheet(
        wb,
        "Estabelecimentos",
        headers,
        detail_rows(
            [
                row
                for row in abertas
                if row.get("conferencia_pedido")
                == "MESMA_RAIZ_CNPJ_ESTABELECIMENTO_DIFERENTE"
            ]
        ),
    )
    add_data_sheet(
        wb,
        "FIDC",
        headers,
        detail_rows([row for row in abertas if row.get("fidc")]),
    )
    create_methodology(wb)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    expected_balance = as_number(totals.get("saldo"))
    validate_workbook(output, len(abertas), expected_balance)
    print(
        f"[relatorio-python] {len(abertas)} em aberto / {len(pagas)} pagas exportadas"
    )
    print(f"[relatorio-python] saldo validado: R$ {expected_balance:,.2f}")
    print(f"[relatorio-python] arquivo: {output}")
    return output


def main() -> None:
    args = parse_args()
    prompt_data_base_if_needed(args)
    data_base = resolve_data_base(args) if getattr(args, "data_base", None) else None
    if not data_base or data_base == date.today().isoformat():
        apply_period_selection(args)
    generate_report(args)


if __name__ == "__main__":
    main()
