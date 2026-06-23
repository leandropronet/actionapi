#!/usr/bin/env python3
r"""Relatório de Estoque x Contas a Pagar — visão macro + nota a nota.

Cruza o valor contábil do Estoque (conta sintética 115, via
/executivo/contabilidade/sintetico) com as parcelas de Contas a Pagar
vinculadas a compra de mercadoria (pedido de compra e/ou nota de entrada,
via /financeiro/contas-pagar), separando o que já foi pago do que está em
aberto.

"Fornecedor de mercadoria" = título de CP com pedido de compra e/ou nota de
entrada vinculados (status_vinculo_pedido = COM_PEDIDO ou COM_NF_SEM_PEDIDO).
Títulos sem nenhum vínculo (frete avulso, serviços, impostos, financeiro,
FIDC etc.) ficam de fora — ver aba Metodologia.

Exemplo:

    .\.venv\Scripts\python.exe scripts\gerar_relatorio_estoque_cp.py
    .\.venv\Scripts\python.exe scripts\gerar_relatorio_estoque_cp.py --data-base 2026-06-21
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from gerar_relatorios_executivos import (
        BLUE,
        MONEY,
        NAVY,
        WHITE,
        add_kpis,
        api_get,
        first_api_key,
        number,
        save,
        title_block,
    )
except ModuleNotFoundError:
    from scripts.gerar_relatorios_executivos import (
        BLUE,
        MONEY,
        NAVY,
        WHITE,
        add_kpis,
        api_get,
        first_api_key,
        number,
        save,
        title_block,
    )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:
    print(
        "Dependência ausente. Execute:\n"
        "  py -m pip install -r scripts/requirements-relatorio-contas-pagar.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_API_URL = "http://127.0.0.1:3000"
PAGE_SIZE = 10_000
DATA_INICIO_PLANO_ATUAL = "2008-01-01"
STATUS_MERCADORIA = ("COM_PEDIDO", "COM_NF_SEM_PEDIDO")

MONEY_HEADERS = {"Valor do título", "Valor das parcelas", "Valor pago", "Saldo em aberto"}
DATE_HEADERS = {"Data emissão", "Primeira baixa", "Última baixa"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera relatório Excel de Estoque x Contas a Pagar (macro + nota a nota)."
    )
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--arquivo", help="Caminho do arquivo .xlsx de saída.")
    parser.add_argument("--data-base", help="Data-base, AAAA-MM-DD (padrão: hoje).")
    return parser.parse_args()


def as_date(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def fetch_all_cp(base_url: str, api_key: str, status_vinculo: str) -> list[dict]:
    rows: list[dict] = []
    page = 1
    while True:
        payload = api_get(
            base_url, api_key, "/api/v1/financeiro/contas-pagar",
            {
                "statusVinculo": status_vinculo,
                "somenteEmAberto": "false",
                "page": page,
                "pageSize": PAGE_SIZE,
            },
        )
        current = payload.get("data", [])
        rows.extend(current)
        if not current or len(rows) >= int(payload.get("total", len(rows))):
            return rows
        page += 1


def resumo_cp(base_url: str, api_key: str, status_vinculo: str) -> dict:
    return api_get(
        base_url, api_key, "/api/v1/financeiro/contas-pagar/resumo",
        {"statusVinculo": status_vinculo, "somenteEmAberto": "false"},
    )


def fetch_estoque(base_url: str, api_key: str, data_base: str) -> list[dict]:
    sintetico = api_get(
        base_url, api_key, "/api/v1/executivo/contabilidade/sintetico",
        {"dataInicio": DATA_INICIO_PLANO_ATUAL, "dataFim": data_base},
    )
    return sintetico.get("contas", [])


def titulo_rows(parcelas: list[dict]) -> list[list[Any]]:
    """Agrupa parcelas por título — uma linha por título/NF de compra."""
    grupos: dict[str, dict[str, Any]] = {}
    for p in parcelas:
        chave = p.get("titulo_id")
        grupo = grupos.get(chave)
        if grupo is None:
            grupo = {
                "filial_id": p.get("filial_id"),
                "filial_nome": p.get("filial_nome"),
                "fornecedor_id": p.get("fornecedor_id"),
                "fornecedor_nome": p.get("fornecedor_nome"),
                "titulo_id": chave,
                "nf_entrada_ids": p.get("nf_entrada_ids"),
                "pedidos_numeros": p.get("pedidos_numeros"),
                "produtos_descricoes": p.get("produtos_descricoes"),
                "data_emissao": p.get("data_emissao"),
                "qtd_parcelas": 0,
                "valor_titulo": 0.0,
                "valor_pago": 0.0,
                "saldo_aberto": 0.0,
                "primeira_baixa": None,
                "ultima_baixa": None,
            }
            grupos[chave] = grupo
        grupo["qtd_parcelas"] += 1
        grupo["valor_titulo"] += number(p.get("valor_parcela"))
        grupo["valor_pago"] += number(p.get("valor_liquido_baixa"))
        grupo["saldo_aberto"] += number(p.get("saldo_parcela"))
        primeira = as_date(p.get("primeira_baixa"))
        if primeira and (grupo["primeira_baixa"] is None or primeira < grupo["primeira_baixa"]):
            grupo["primeira_baixa"] = primeira
        ultima = as_date(p.get("ultima_baixa"))
        if ultima and (grupo["ultima_baixa"] is None or ultima > grupo["ultima_baixa"]):
            grupo["ultima_baixa"] = ultima

    headers = [
        "Filial", "Fornecedor", "Título", "NF(s) de entrada", "Pedido(s)",
        "Produto(s)", "Data emissão", "Parcelas", "Valor do título",
        "Valor pago", "Saldo em aberto", "Situação", "Primeira baixa", "Última baixa",
    ]
    output = [headers]
    for g in sorted(grupos.values(), key=lambda item: item["saldo_aberto"], reverse=True):
        aberto = g["saldo_aberto"]
        situacao = "PAGO" if abs(aberto) <= 0.01 else ("PARCIAL" if g["valor_pago"] > 0.01 else "EM ABERTO")
        output.append([
            f"{g['filial_id']} — {g['filial_nome']}" if g["filial_nome"] else g["filial_id"],
            f"{g['fornecedor_id']} — {g['fornecedor_nome']}" if g["fornecedor_nome"] else g["fornecedor_id"],
            g["titulo_id"],
            g["nf_entrada_ids"],
            g["pedidos_numeros"],
            g["produtos_descricoes"],
            as_date(g["data_emissao"]),
            g["qtd_parcelas"],
            g["valor_titulo"],
            g["valor_pago"],
            aberto,
            situacao,
            g["primeira_baixa"],
            g["ultima_baixa"],
        ])
    return output


def add_table_sheet(wb: Workbook, name: str, rows: list[list[Any]]) -> None:
    print(f"[estoque-cp] aba {name}: {len(rows) - 1} linhas", flush=True)
    ws = wb.create_sheet(name[:31])
    headers = rows[0]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    ws.freeze_panes = "A2"
    if ws.max_row >= 2:
        ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, 1):
        letter = ws.cell(1, idx).column_letter
        if header in MONEY_HEADERS:
            for cell in ws[letter][1:]:
                cell.number_format = MONEY
        elif header in DATE_HEADERS:
            for cell in ws[letter][1:]:
                cell.number_format = "dd/mm/yyyy"
    sample = min(ws.max_row, 300)
    for cells in ws.iter_cols(min_row=1, max_row=sample):
        width = min(46, max(10, max(len(str(c.value or "")) for c in cells) + 2))
        ws.column_dimensions[cells[0].column_letter].width = width


def create_panel(wb: Workbook, data_base: str, estoque: list[dict],
                  resumo_mercadoria: dict, resumo_outras: dict) -> None:
    by_id = {c["conta_id"]: c for c in estoque}
    valor_estoque = number(by_id.get("115", {}).get("saldo")) if "115" in by_id else 0.0

    tot_merc = resumo_mercadoria.get("totalizadores", {})
    tot_outras = resumo_outras.get("totalizadores", {})
    valor_titulos_merc = number(tot_merc.get("valor_parcelas"))
    valor_pago_merc = number(tot_merc.get("valor_baixado"))
    saldo_aberto_merc = number(tot_merc.get("saldo"))

    ws = wb.active
    ws.title = "Painel"
    title_block(ws, "ESTOQUE × CONTAS A PAGAR", f"Posição em {data_base}")

    kpis = [
        ("Valor contábil do Estoque (conta 115)", valor_estoque, "money"),
        ("CP de mercadoria — valor das parcelas (histórico)", valor_titulos_merc, "money"),
        ("CP de mercadoria — já pago", valor_pago_merc, "money"),
        ("CP de mercadoria — saldo em aberto", saldo_aberto_merc, "money"),
        ("CP de mercadoria — % do valor já pago", (valor_pago_merc / valor_titulos_merc) if valor_titulos_merc else 0, "percent"),
        ("Títulos de mercadoria", tot_merc.get("qtd_titulos"), "number"),
        ("Parcelas de mercadoria", tot_merc.get("qtd_parcelas"), "number"),
        ("Fornecedores de mercadoria", tot_merc.get("qtd_fornecedores"), "number"),
        ("CP sem vínculo de mercadoria — saldo em aberto", number(tot_outras.get("saldo")), "money"),
    ]
    add_kpis(ws, kpis)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22

    obs_row = 5 + len(kpis) + 2
    ws.cell(obs_row, 1, "Observação")
    ws.cell(obs_row, 2, (
        "O Estoque contábil e o saldo de CP de mercadoria NÃO têm vínculo 1:1 — "
        "produto já vendido sai do estoque mas a parcela de compra pode continuar "
        "em aberto. Os dois números juntos dão a leitura de financiamento "
        "(quanto do estoque/compras ainda não foi pago), não uma conciliação exata."
    ))
    ws.cell(obs_row, 2).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=obs_row, start_column=2, end_row=obs_row, end_column=4)


def create_methodology(wb: Workbook) -> None:
    ws = wb.create_sheet("Metodologia")
    content = [
        ("ESTOQUE × CONTAS A PAGAR — METODOLOGIA", ""),
        ("Estoque", "Saldo contábil acumulado da conta sintética 115 (ESTOQUES), via /executivo/contabilidade/sintetico — mesma fonte do relatório de Patrimônio."),
        ("Fornecedor de mercadoria", "Título de CP com pedido de compra e/ou nota de entrada vinculados (status_vinculo_pedido = COM_PEDIDO ou COM_NF_SEM_PEDIDO, já calculado em /financeiro/contas-pagar). Títulos sem nenhum vínculo (frete avulso, serviços, impostos, despesas financeiras, FIDC etc.) ficam fora — aparecem só como contraste no Painel."),
        ("Pago × em aberto", "Pago = soma de valor_baixado das parcelas (CRCBAIXA/baixas já registradas). Em aberto = soma de saldo_parcela (saldo oficial, já com o sinal de adiantamentos/créditos aplicado)."),
        ("Granularidade da aba Nota a Nota", "Uma linha por título de CP (CABPAGAR), agregando as parcelas. NF(s)/Pedido(s)/Produto(s) vêm do vínculo automático título → pedido de compra → nota de entrada já existente na ActionAPI."),
        ("Limitação", "Não há vínculo direto entre uma baixa de estoque (saída por venda) e uma parcela de CP específica — por isso o relatório cruza totais (visão macro) e o histórico de compras (nota a nota), mas não afirma 'este item do estoque foi pago por esta parcela'."),
        ("Cobertura", "Cerca de 31% dos títulos de CP têm nota de entrada/pedido vinculado (dado de 2026-06-23); o restante são despesas sem relação direta com mercadoria."),
    ]
    for row in content:
        ws.append(row)
    ws.merge_cells("A1:B1")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def generate_report(args: argparse.Namespace) -> Path:
    data_base = args.data_base or date.today().isoformat()
    output = (
        Path(args.arquivo).resolve()
        if args.arquivo
        else ROOT / "relatorios" / f"estoque-contas-a-pagar-{data_base}.xlsx"
    )

    api_key = first_api_key()
    print(f"[estoque-cp] consultando {args.api_url}...", flush=True)
    estoque = fetch_estoque(args.api_url, api_key, data_base)

    parcelas_mercadoria: list[dict] = []
    for status in STATUS_MERCADORIA:
        parcelas_mercadoria.extend(fetch_all_cp(args.api_url, api_key, status))
    print(f"[estoque-cp] {len(parcelas_mercadoria)} parcelas de CP vinculadas a mercadoria.", flush=True)

    resumo_mercadoria_totais = {"qtd_titulos": 0, "qtd_parcelas": 0, "qtd_fornecedores": set(),
                                 "valor_parcelas": 0.0, "valor_baixado": 0.0, "saldo": 0.0}
    fornecedores_vistos: set[str] = set()
    titulos_vistos: set[str] = set()
    for p in parcelas_mercadoria:
        resumo_mercadoria_totais["qtd_parcelas"] += 1
        titulos_vistos.add(p.get("titulo_id"))
        fornecedores_vistos.add(p.get("fornecedor_id"))
        resumo_mercadoria_totais["valor_parcelas"] += number(p.get("valor_parcela"))
        resumo_mercadoria_totais["valor_baixado"] += number(p.get("valor_liquido_baixa"))
        resumo_mercadoria_totais["saldo"] += number(p.get("saldo_parcela"))
    resumo_mercadoria = {
        "totalizadores": {
            "qtd_titulos": len(titulos_vistos),
            "qtd_parcelas": resumo_mercadoria_totais["qtd_parcelas"],
            "qtd_fornecedores": len(fornecedores_vistos),
            "valor_parcelas": resumo_mercadoria_totais["valor_parcelas"],
            "valor_baixado": resumo_mercadoria_totais["valor_baixado"],
            "saldo": resumo_mercadoria_totais["saldo"],
        }
    }
    resumo_outras = resumo_cp(args.api_url, api_key, "SEM_NF_E_SEM_PEDIDO")

    wb = Workbook()
    create_panel(wb, data_base, estoque, resumo_mercadoria, resumo_outras)

    por_fornecedor: dict[str, dict[str, Any]] = defaultdict(lambda: {
        "fornecedor_nome": None, "qtd_titulos": set(), "valor_parcelas": 0.0,
        "valor_pago": 0.0, "saldo_aberto": 0.0,
    })
    for p in parcelas_mercadoria:
        item = por_fornecedor[p.get("fornecedor_id")]
        item["fornecedor_nome"] = p.get("fornecedor_nome")
        item["qtd_titulos"].add(p.get("titulo_id"))
        item["valor_parcelas"] += number(p.get("valor_parcela"))
        item["valor_pago"] += number(p.get("valor_liquido_baixa"))
        item["saldo_aberto"] += number(p.get("saldo_parcela"))
    fornecedor_rows = [["Fornecedor", "Códigos", "Títulos", "Valor das parcelas", "Valor pago", "Saldo em aberto"]]
    for fornecedor_id, item in sorted(por_fornecedor.items(), key=lambda kv: kv[1]["saldo_aberto"], reverse=True):
        fornecedor_rows.append([
            item["fornecedor_nome"], fornecedor_id, len(item["qtd_titulos"]),
            item["valor_parcelas"], item["valor_pago"], item["saldo_aberto"],
        ])
    add_table_sheet(wb, "Por Fornecedor", fornecedor_rows)

    add_table_sheet(wb, "Nota a Nota", titulo_rows(parcelas_mercadoria))

    create_methodology(wb)

    output.parent.mkdir(parents=True, exist_ok=True)
    save(wb, output, ["Painel", "Por Fornecedor", "Nota a Nota", "Metodologia"])
    print(f"[estoque-cp] arquivo: {output}")
    return output


def main() -> None:
    generate_report(parse_args())


if __name__ == "__main__":
    main()
