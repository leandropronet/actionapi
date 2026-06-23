#!/usr/bin/env python3
"""Gera o relatório de Fluxo de Caixa (a receber x a pagar) espelhando a
planilha do controller (SGA - Cash Flow).

Arquitetura (igual ao controller): uma base "BD Fluxo" com TODAS as parcelas em
aberto de Contas a Receber e Contas a Pagar (saldo em aberto, com sinal, e o
valor convertido para R$), e a partir dela:
  - projeção por mês de vencimento (A Receber, A Pagar, Líquido, Acumulado);
  - aging em 13 faixas (vencido/a vencer) para CR e CP, como no Demonstrativo
    Financeiro do controller.

Fonte: reprodução local validada (packages/etl/.../saldo-aberto-historico.js),
sem acessar Oracle. Não usa a ActionAPI. Para reproduzir uma data passada, use
--data-base AAAA-MM-DD (padrão: hoje).

Exemplos:
    py scripts/gerar_relatorio_fluxo_caixa.py
    py scripts/gerar_relatorio_fluxo_caixa.py --data-base 2026-06-21
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ModuleNotFoundError:
    print(
        "Dependência ausente. Execute:\n"
        "  py -m pip install -r scripts/requirements-relatorio-contas-pagar.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


ROOT = Path(__file__).resolve().parents[1]
NODE_SALDO_HISTORICO = ROOT / "packages" / "etl" / "src" / "scripts" / "saldo-aberto-historico.js"

BLUE = "17365D"
GREEN = "70AD47"
RED = "C00000"
WHITE = "FFFFFF"
MONEY_FORMAT = 'R$ #,##0.00;[Red]-R$ #,##0.00'
NUMBER_FORMAT = "#,##0"
DATE_FORMAT = "dd/mm/yyyy"

# Faixas de aging, calibradas contra o Demonstrativo Financeiro do controller.
AGING_ORDER = [
    "01-Vencido acima de 360 dias",
    "02-Vencido entre 180 e 360 dias",
    "03-Vencido entre 90 e 180 dias",
    "04-Vencido entre 60 e 90 dias",
    "05-Vencido entre 30 e 60 dias",
    "06-Vencido até 30 dias",
    "07-Vence em até 30 dias",
    "08-Vence entre 30 e 60 dias",
    "09-Vence entre 60 e 90 dias",
    "10-Vence entre 90 e 120 dias",
    "11-Vence entre 120 e 180 dias",
    "12-Vence entre 180 e 360 dias",
    "13-Vence acima de 360 dias",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera relatório Excel de Fluxo de Caixa.")
    parser.add_argument("--arquivo", help="Caminho do arquivo .xlsx de saída.")
    parser.add_argument("--data-base", help="Data-base do saldo, AAAA-MM-DD (padrão: hoje).")
    return parser.parse_args()


def as_number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def as_date(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def fetch_ambos(data_base: str) -> dict:
    """Reproduz CR e CP em aberto na data-base via PostgreSQL (sem Oracle)."""
    print(
        f"[fluxo-caixa] reproduzindo CR e CP em {data_base} via PostgreSQL "
        "(sem Oracle, fórmula validada)...",
        flush=True,
    )
    result = subprocess.run(
        ["node", str(NODE_SALDO_HISTORICO), "--tipo", "AMBOS", "--data-base", data_base],
        capture_output=True,
        text=True,
        encoding="utf-8",
        cwd=str(NODE_SALDO_HISTORICO.parents[2]),
    )
    if result.returncode != 0:
        raise RuntimeError(f"Falha ao calcular saldos em aberto: {result.stderr}")
    return json.loads(result.stdout)


def faixa_aging(atraso_dias: int) -> str:
    """Classifica pela quantidade de dias de atraso (positivo = vencido)."""
    a = atraso_dias
    if a > 360:
        return AGING_ORDER[0]
    if a > 180:
        return AGING_ORDER[1]
    if a > 90:
        return AGING_ORDER[2]
    if a > 60:
        return AGING_ORDER[3]
    if a > 30:
        return AGING_ORDER[4]
    if a > 0:
        return AGING_ORDER[5]
    falta = -a
    if falta <= 30:
        return AGING_ORDER[6]
    if falta <= 60:
        return AGING_ORDER[7]
    if falta <= 90:
        return AGING_ORDER[8]
    if falta <= 120:
        return AGING_ORDER[9]
    if falta <= 180:
        return AGING_ORDER[10]
    if falta <= 360:
        return AGING_ORDER[11]
    return AGING_ORDER[12]


def style_header(ws, row: int) -> None:
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[Any]],
                    money_cols: set[int], date_cols: set[int] | None = None) -> None:
    date_cols = date_cols or set()
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws, 1)
    ws.freeze_panes = "A2"
    for col in money_cols:
        for cell in ws[ws.cell(1, col).column_letter][1:]:
            cell.number_format = MONEY_FORMAT
    for col in date_cols:
        for cell in ws[ws.cell(1, col).column_letter][1:]:
            cell.number_format = DATE_FORMAT
    for column_cells in ws.iter_cols(min_row=1, max_row=min(ws.max_row, 200)):
        width = max(10, min(46, max(len(str(c.value or "")) for c in column_cells) + 2))
        ws.column_dimensions[column_cells[0].column_letter].width = width
    if ws.max_row >= 2:
        table = Table(displayName="T" + name.replace(" ", ""), ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(table)


def build_bd(cr_rows: list[dict], cp_rows: list[dict]) -> list[list[Any]]:
    out = []
    for tipo, rows, nome_key, cod_key in (
        ("CONTAS A RECEBER", cr_rows, "cliente_nome", "cliente_id"),
        ("CONTAS A PAGAR", cp_rows, "fornecedor_nome", "fornecedor_id"),
    ):
        for r in rows:
            out.append([
                tipo,
                r.get("filial_id"),
                r.get("filial_nome") or r.get("filial_identificacao"),
                r.get(cod_key),
                r.get(nome_key),
                r.get("tipo_documento_descricao"),
                r.get("numero_documento") or r.get("parcela_nr"),
                as_date(r.get("data_emissao")),
                r.get("unidade_saldo") or "R$",
                as_number(r.get("saldo_parcela")),
                as_number(r.get("saldo_convertido_atual")),
                as_date(r.get("data_vencimento")),
                r.get("situacao"),
            ])
    return out


def monthly_projection(cr_rows: list[dict], cp_rows: list[dict], base: date) -> list[list[Any]]:
    receber: dict[str, float] = defaultdict(float)
    pagar: dict[str, float] = defaultdict(float)

    def bucket(venc: date) -> str:
        return "VENCIDO" if venc < base else f"{venc.year:04d}-{venc.month:02d}"

    for r in cr_rows:
        v = as_date(r.get("data_vencimento"))
        if v:
            receber[bucket(v.date())] += as_number(r.get("saldo_convertido_atual"))
    for r in cp_rows:
        v = as_date(r.get("data_vencimento"))
        if v:
            pagar[bucket(v.date())] += as_number(r.get("saldo_convertido_atual"))

    chaves = sorted(set(receber) | set(pagar), key=lambda k: ("" if k == "VENCIDO" else k))
    out = []
    acumulado = 0.0
    for k in chaves:
        rec = receber.get(k, 0.0)
        pag = pagar.get(k, 0.0)
        liq = rec - pag
        acumulado += liq
        out.append([k, rec, pag, liq, acumulado])
    return out


def aging_table(rows: list[dict], base: date) -> list[list[Any]]:
    g: dict[str, dict[str, Any]] = {f: {"qtd": 0, "saldo": 0.0} for f in AGING_ORDER}
    for r in rows:
        v = as_date(r.get("data_vencimento"))
        if not v:
            continue
        atraso = (base - v.date()).days
        faixa = faixa_aging(atraso)
        g[faixa]["qtd"] += 1
        g[faixa]["saldo"] += as_number(r.get("saldo_convertido_atual"))
    return [[f, g[f]["qtd"], g[f]["saldo"]] for f in AGING_ORDER]


def create_panel(wb: Workbook, base: date, cr_rows: list[dict], cp_rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Painel"
    ws.merge_cells("A1:D1")
    ws["A1"] = "FLUXO DE CAIXA — POSIÇÃO CONSOLIDADA"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Data-base do saldo"
    ws["B3"] = base
    ws["B3"].number_format = DATE_FORMAT
    ws["A4"] = "Gerado em"
    ws["B4"] = datetime.now()
    ws["B4"].number_format = "dd/mm/yyyy hh:mm"
    ws["A5"] = "Fonte"
    ws["B5"] = "Reprodução local validada (VALOR_ABERTO_RECEBER/PAGAR_DATA)"

    total_receber = sum(as_number(r.get("saldo_convertido_atual")) for r in cr_rows)
    total_pagar = sum(as_number(r.get("saldo_convertido_atual")) for r in cp_rows)
    indicators = [
        ("Total a Receber (convertido R$)", total_receber),
        ("Total a Pagar (convertido R$)", total_pagar),
        ("Líquido (Receber − Pagar)", total_receber - total_pagar),
        ("Parcelas a receber", len(cr_rows)),
        ("Parcelas a pagar", len(cp_rows)),
    ]
    ws["A7"] = "INDICADORES"
    ws["B7"] = "VALOR"
    style_header(ws, 7)
    for i, (label, value) in enumerate(indicators, 8):
        ws.cell(i, 1, label)
        cell = ws.cell(i, 2, as_number(value))
        cell.number_format = MONEY_FORMAT if i <= 10 else NUMBER_FORMAT

    # Aging lado a lado (CR e CP), como no Demonstrativo do controller.
    cr_aging = aging_table(cr_rows, base)
    cp_aging = aging_table(cp_rows, base)
    start = 15
    ws.cell(start, 1, "AGING — CONTAS A RECEBER")
    ws.cell(start, 2, "Saldo (R$)")
    ws.cell(start, 4, "AGING — CONTAS A PAGAR")
    ws.cell(start, 5, "Saldo (R$)")
    for col in (1, 2, 4, 5):
        ws.cell(start, col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(start, col).font = Font(color=WHITE, bold=True)
    for i, (faixa, _qtd, saldo) in enumerate(cr_aging, 1):
        ws.cell(start + i, 1, faixa)
        ws.cell(start + i, 2, saldo).number_format = MONEY_FORMAT
    for i, (faixa, _qtd, saldo) in enumerate(cp_aging, 1):
        ws.cell(start + i, 4, faixa)
        ws.cell(start + i, 5, saldo).number_format = MONEY_FORMAT
    ws.cell(start + len(AGING_ORDER) + 1, 1, "Total Geral")
    ws.cell(start + len(AGING_ORDER) + 1, 2, total_receber).number_format = MONEY_FORMAT
    ws.cell(start + len(AGING_ORDER) + 1, 4, "Total Geral")
    ws.cell(start + len(AGING_ORDER) + 1, 5, total_pagar).number_format = MONEY_FORMAT

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 20


def create_methodology(wb: Workbook) -> None:
    ws = wb.create_sheet("Metodologia")
    content = [
        ("FLUXO DE CAIXA — METODOLOGIA", ""),
        ("Base", "Todas as parcelas EM ABERTO de Contas a Receber e Contas a Pagar na data-base, com saldo em aberto (com sinal) e valor convertido para R$."),
        ("Sinal", "Documentos de natureza crédito (adiantamento de cliente/fornecedor) entram negativos, igual ao controller."),
        ("Projeção mensal", "A Receber e A Pagar agregados pelo mês de vencimento; parcelas já vencidas vão para o bucket VENCIDO. Líquido = Receber − Pagar; Acumulado é o caixa projetado."),
        ("Aging", "13 faixas calibradas contra o Demonstrativo Financeiro do controller. As faixas 'a vencer' (07–13) reproduzem o controller; o vencido profundo (01–03) pode divergir por diferença de data de referência/método do controller para itens muito antigos."),
        ("Conversão", "Contratos indexados (SJ$, US$, ER) usam a cotação mais recente para consolidar em reais."),
        ("Validação", "Totais conferidos contra o controller (CR 159,29 mi / CP 105,18 mi em 21/06/2026)."),
    ]
    for row in content:
        ws.append(row)
    ws.merge_cells("A1:B1")
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def generate_report(args: argparse.Namespace) -> Path:
    data_base = args.data_base or date.today().isoformat()
    base = datetime.fromisoformat(data_base).date()
    payload = fetch_ambos(data_base)
    cr_rows = payload["cr"]["rows"]
    cp_rows = payload["cp"]["rows"]
    print(f"[fluxo-caixa] {len(cr_rows)} parcelas a receber + {len(cp_rows)} a pagar.", flush=True)

    output = (
        Path(args.arquivo).resolve()
        if args.arquivo
        else ROOT / "relatorios" / f"fluxo-de-caixa-{data_base}.xlsx"
    )

    wb = Workbook()
    create_panel(wb, base, cr_rows, cp_rows)

    proj = monthly_projection(cr_rows, cp_rows, base)
    add_table_sheet(
        wb, "Fluxo Mensal",
        ["Período", "A Receber", "A Pagar", "Líquido", "Acumulado"],
        proj, money_cols={2, 3, 4, 5},
    )

    bd = build_bd(cr_rows, cp_rows)
    add_table_sheet(
        wb, "BD Fluxo",
        ["Tipo", "Filial", "Filial nome", "Código", "Cliente/Fornecedor",
         "Tipo documento", "Documento", "Emissão", "Moeda", "VL Aberto (unidade)",
         "VL Aberto (R$)", "Vencimento", "Situação"],
        bd, money_cols={10, 11}, date_cols={8, 12},
    )

    cr_aging = aging_table(cr_rows, base)
    cp_aging = aging_table(cp_rows, base)
    add_table_sheet(
        wb, "Aging CR",
        ["Faixa", "Quantidade", "Saldo (R$)"], cr_aging, money_cols={3},
    )
    add_table_sheet(
        wb, "Aging CP",
        ["Faixa", "Quantidade", "Saldo (R$)"], cp_aging, money_cols={3},
    )
    create_methodology(wb)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    total_receber = sum(as_number(r.get("saldo_convertido_atual")) for r in cr_rows)
    total_pagar = sum(as_number(r.get("saldo_convertido_atual")) for r in cp_rows)
    print(f"[fluxo-caixa] a receber R$ {total_receber:,.2f} | a pagar R$ {total_pagar:,.2f} "
          f"| líquido R$ {total_receber - total_pagar:,.2f}")
    print(f"[fluxo-caixa] arquivo: {output}")
    return output


def main() -> None:
    generate_report(parse_args())


if __name__ == "__main__":
    main()
