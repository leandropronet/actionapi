#!/usr/bin/env python3
r"""Gera o relatório de Patrimônio (Balanço Patrimonial sintético + analítico).

Reaproveita /api/v1/executivo/contabilidade/sintetico, que soma o saldo
acumulado de cada conta do plano de contas (sintética ou analítica) a partir
de todos os lançamentos contábeis até a data-base.

Exemplo:

    .\.venv\Scripts\python.exe scripts\gerar_relatorio_patrimonio.py
    .\.venv\Scripts\python.exe scripts\gerar_relatorio_patrimonio.py --data-base 2026-06-21

O script não acessa Oracle nem PostgreSQL diretamente. A API key é lida de
API_KEYS no .env e nunca é gravada na planilha.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path
from typing import Any

try:
    from gerar_relatorios_executivos import (
        NAVY,
        WHITE,
        add_kpis,
        add_synthetic_accounts_sheet,
        api_get,
        first_api_key,
        number,
        save,
        title_block,
    )
except ModuleNotFoundError:
    from scripts.gerar_relatorios_executivos import (
        NAVY,
        WHITE,
        add_kpis,
        add_synthetic_accounts_sheet,
        api_get,
        first_api_key,
        number,
        save,
        title_block,
    )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:
    print(
        "Dependência ausente. Execute:\n"
        "  py -m pip install -r scripts/requirements-relatorio-contas-pagar.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_API_URL = "http://127.0.0.1:3000"
# Início real do plano de contas atual (1000002) — datas anteriores são
# irrelevantes para a consulta (o endpoint só soma esse plano), mas usamos
# uma margem segura.
DATA_INICIO_PLANO_ATUAL = "2008-01-01"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera relatório Excel de Patrimônio (Balanço).")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--arquivo", help="Caminho do arquivo .xlsx de saída.")
    parser.add_argument("--data-base", help="Data-base do balanço, AAAA-MM-DD (padrão: hoje).")
    return parser.parse_args()


def create_panel(wb: Workbook, data_base: str, contas: list[dict]) -> None:
    by_id = {c["conta_id"]: c for c in contas}

    def saldo(conta_id: str) -> float:
        row = by_id.get(conta_id)
        return number(row.get("saldo")) if row else 0.0

    total_ativo = saldo("1")
    ativo_circulante = saldo("11")
    ativo_nao_circulante = saldo("12")
    # Passivo é de natureza credora — saldo sai negativo na convenção D-C; o
    # balanço tradicional mostra como positivo.
    total_passivo = -saldo("2")
    passivo_circulante = -saldo("21")
    passivo_nao_circulante = -saldo("22")
    patrimonio_liquido = -saldo("23")
    diferenca = total_ativo - total_passivo

    ws = wb.active
    ws.title = "Painel"
    title_block(ws, "PATRIMÔNIO — BALANÇO PATRIMONIAL", f"Saldo acumulado até {data_base}")

    kpis = [
        ("Total do Ativo", total_ativo, "money"),
        ("  Ativo Circulante", ativo_circulante, "money"),
        ("  Ativo Não Circulante", ativo_nao_circulante, "money"),
        ("Total do Passivo + Patrimônio Líquido", total_passivo, "money"),
        ("  Passivo Circulante", passivo_circulante, "money"),
        ("  Passivo Não Circulante", passivo_nao_circulante, "money"),
        ("  Patrimônio Líquido", patrimonio_liquido, "money"),
        ("Diferença (Ativo − Passivo, deveria ser ≈0)", diferenca, "money"),
        ("Diferença em % do Ativo", (diferenca / total_ativo) if total_ativo else 0, "percent"),
    ]
    add_kpis(ws, kpis)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22

    alert_row = 5 + len(kpis) + 2
    if total_ativo and abs(diferenca / total_ativo) > 0.02:
        ws.cell(alert_row, 1, "ATENÇÃO: diferença acima de 2% do Ativo — investigar antes de usar.")
        ws.cell(alert_row, 1).font = Font(bold=True, color="C00000")


def create_methodology(wb: Workbook) -> None:
    ws = wb.create_sheet("Metodologia")
    content = [
        ("PATRIMÔNIO — METODOLOGIA", ""),
        ("Fonte", "ActionAPI: /executivo/contabilidade/sintetico (soma raw.contabil por conta, com rollup hierárquico pelo prefixo do código)."),
        ("Saldo", "Acumulado de TODOS os lançamentos do plano de contas atual (1000002) até a data-base — não é um filtro de período, é o saldo patrimonial na data."),
        ("Convenção de sinal", "Saldo = Débitos − Créditos. Contas de natureza credora (Passivo, Patrimônio Líquido, Receitas) saem negativas nessa convenção; o painel inverte o sinal para exibir como positivo, como num balanço tradicional."),
        ("Cobertura histórica", "Plano 1000002 começou em 31/12/2014 e está 100% sincronizado desde então (conferido contra o Oracle em 23/06/2026: 2.078.107 partidas, 0 divergência)."),
        ("Planos antigos NÃO incluídos", "Existiram 2 planos de contas anteriores (código '1', 2010-2014, e '1000001', 2008-2010), já sincronizados em raw.contabil, mas com estrutura de códigos de conta DIFERENTE e sem mapeamento validado para o plano atual (o mesmo código pequeno, ex. '12', significa contas diferentes em cada plano). Por isso este relatório não soma o histórico anterior a 2014-12-31 — fazer isso exigiria mapear conta a conta entre os 3 planos, decisão de produto pendente."),
        ("Diferença residual Ativo × Passivo", "Pequena diferença (tipicamente <1% do Ativo) é esperada: lançamentos com conta contábil fora da hierarquia do plano atual, ajustes de tipo Fiscal×Societário (CABLANCTB.TIPO_CLC) e o próprio efeito de não ter o saldo de abertura dos planos antigos. Investigar se ultrapassar 2%."),
        ("Hierarquia", "Cada conta soma todas as suas descendentes pelo prefixo do código (ex.: conta '11' soma tudo que começa com '11') — mesma lógica do cadastro CONTASPL do SiAGRI."),
    ]
    for row in content:
        ws.append(row)
    ws.merge_cells("A1:B1")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def generate_report(args: argparse.Namespace) -> Path:
    data_base = args.data_base or date.today().isoformat()
    output = (
        Path(args.arquivo).resolve()
        if args.arquivo
        else ROOT / "relatorios" / f"patrimonio-{data_base}.xlsx"
    )

    api_key = first_api_key()
    print(f"[patrimonio] consultando {args.api_url}...", flush=True)
    sintetico = api_get(
        args.api_url, api_key, "/api/v1/executivo/contabilidade/sintetico",
        {"dataInicio": DATA_INICIO_PLANO_ATUAL, "dataFim": data_base},
    )
    contas = sintetico.get("contas", [])
    print(f"[patrimonio] {len(contas)} contas do plano retornadas.", flush=True)

    wb = Workbook()
    create_panel(wb, data_base, contas)

    ativo = [c for c in contas if c["conta_id"].startswith("1")]
    passivo = [c for c in contas if c["conta_id"].startswith("2")]
    add_synthetic_accounts_sheet(wb, "Ativo", ativo)
    add_synthetic_accounts_sheet(wb, "Passivo e PL", passivo)
    create_methodology(wb)

    output.parent.mkdir(parents=True, exist_ok=True)
    save(wb, output, ["Painel", "Ativo", "Passivo e PL", "Metodologia"])
    print(f"[patrimonio] arquivo: {output}")
    return output


def main() -> None:
    generate_report(parse_args())


if __name__ == "__main__":
    main()
